import streamlit as st
import pandas as pd
import random
import os
import re
import json
import urllib.parse
from datetime import datetime, timedelta, date
from io import BytesIO

# Importar Scraper de la Reunión
try:
    from jw_scraper import JWScraper
    JW_AVAILABLE = True
except ImportError:
    JW_AVAILABLE = False

# Importar librerías de generación de archivos
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet

# ==========================================
# CONFIGURACIÓN DE PÁGINA Y DISEÑO PREMIUM
# ==========================================
st.set_page_config(
    page_title="Coordinación VMC - El Araguaney",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo CSS premium inyectado para mejorar la experiencia de usuario móvil y de escritorio
st.markdown("""
<style>
    /* Estilos globales y paleta de colores teocráticos */
    .stApp {
        background-color: #f7f9fc;
    }
    h1, h2, h3 {
        color: #4A148C;
        font-family: 'Outfit', 'Segoe UI', sans-serif;
    }
    
    /* Métrica de Roster */
    .metric-card {
        background-color: white;
        border-radius: 12px;
        padding: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border-left: 5px solid #7B1FA2;
        text-align: center;
        margin-bottom: 10px;
    }
    .metric-card h4 {
        margin: 0;
        color: #555;
        font-size: 0.9rem;
    }
    .metric-card h2 {
        margin: 5px 0 0 0;
        color: #7B1FA2;
        font-size: 1.8rem;
    }

    /* Cards de Semanas */
    .semana-card {
        background-color: white;
        border-radius: 16px;
        padding: 20px;
        box-shadow: 0 6px 12px rgba(0,0,0,0.04);
        margin-bottom: 25px;
        border: 1px solid #e1e8ed;
    }
    
    /* Botones de acción */
    div.stButton > button {
        border-radius: 8px !important;
        font-weight: bold !important;
        transition: all 0.3s ease !important;
    }
    
    /* WhatsApp buttons */
    .whatsapp-btn {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        background-color: #25D366;
        color: white !important;
        text-decoration: none;
        padding: 8px 12px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85rem;
        margin-right: 5px;
        margin-top: 5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        transition: background-color 0.2s;
    }
    .whatsapp-btn:hover {
        background-color: #128C7E;
    }
    
    .whatsapp-web-btn {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        background-color: #128C7E;
        color: white !important;
        text-decoration: none;
        padding: 8px 12px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85rem;
        margin-top: 5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        transition: background-color 0.2s;
    }
    .whatsapp-web-btn:hover {
        background-color: #075E54;
    }
</style>
""", unsafe_allow_html=True)


# ==========================================
# SISTEMA DE CONFIGURACIÓN Y PERSISTENCIA
# ==========================================
EXCEL_FILE = "Congregacion_Araguaney.xlsx"
HISTORIAL_FILE = "vmc_historial.json"

@st.cache_data
def cargar_roster_excel(file_path):
    """Carga y procesa la información de publicadores desde el Excel.
    Acepta ruta de archivo (str) o un objeto UploadedFile de Streamlit.
    """
    try:
        # Si es un objeto UploadedFile de Streamlit, leerlo directamente
        if hasattr(file_path, 'read'):
            df = pd.read_excel(file_path)
        else:
            if not os.path.exists(file_path):
                return None
            df = pd.read_excel(file_path)

        # Filtros de exclusión clásicos
        df = df[~df['Nombre'].str.contains('Spolzino|Saucedo', case=False, na=False)]

        # Rellenar columnas esenciales si faltan
        if 'Es_Menor' not in df.columns:
            df['Es_Menor'] = 'No'
        df['Es_Menor'] = df['Es_Menor'].astype(str).str.strip().str.title()

        if 'Telefono' not in df.columns:
            df['Telefono'] = ''
        df['Telefono'] = df['Telefono'].astype(str).str.strip()

        # Asegurar todas las habilidades si no están presentes
        habilidades = [
            'Hab_Tes_Discurso', 'Hab_Tes_Perlas', 'Hab_Lectura',
            'Hab_Mae_Conversacion_Enc', 'Hab_Mae_Conversacion_Ayu',
            'Hab_Mae_Revisita_Enc', 'Hab_Mae_Revisita_Ayu',
            'Hab_Mae_Discipulos_Enc', 'Hab_Mae_Discipulos_Ayu',
            'Hab_Mae_Creencias_Esc_Enc', 'Hab_Mae_Creencias_Esc_Ayu',
            'Hab_Mae_Creencias_Dis', 'Hab_Mae_Discurso',
            'Hab_Oracion', 'Hab_Vida_Parte1', 'Hab_Vida_Parte2', 'Hab_Vida_Locales',
            'Hab_Estudio_Conductor', 'Hab_Estudio_Lector',
            'Hab_Sonido', 'Hab_Mics', 'Hab_Plataforma', 'Hab_Acomodador'
        ]
        for hab in habilidades:
            if hab not in df.columns:
                df[hab] = 'No'
                # Inteligencia por defecto basada en privilegios / género si no hay Excel de capacidades completo
                if hab == 'Hab_Oracion':
                    df.loc[df['Genero'] == 'M', hab] = 'Si'
                elif hab == 'Hab_Lectura':
                    df.loc[df['Genero'] == 'M', hab] = 'Si'
                elif hab.startswith('Hab_Mae_'):
                    df[hab] = 'Si'
                elif hab == 'Hab_Estudio_Conductor':
                    df.loc[df['Privilegio'] == 'Anciano', hab] = 'Si'
                elif hab == 'Hab_Estudio_Lector':
                    df.loc[df['Genero'] == 'M', hab] = 'Si'

        return df
    except Exception as e:
        st.error(f"Error al leer el archivo Excel: {e}")
        return None


def cargar_historial():
    """Carga el historial de la base de datos local"""
    if os.path.exists(HISTORIAL_FILE):
        try:
            with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def guardar_historial(historial):
    """Guarda el historial a nivel local"""
    try:
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(historial, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"Error al guardar historial: {e}")


# ==========================================
# CORE DE ASIGNACIONES (ALGORITMO VMC)
# ==========================================
class AsignadorVMC:
    def __init__(self, df_roster):
        self.df = df_roster
        self.historial_maestros = cargar_historial()
        
        # Listas básicas
        self.ancianos = self.df[(self.df['Privilegio'] == 'Anciano') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        self.siervos = self.df[(self.df['Privilegio'] == 'Siervo Min.') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        self.todos_varones = self.df[self.df['Genero'] == 'M']['Nombre'].tolist()
        self.hermanas = self.df[self.df['Genero'] == 'F']['Nombre'].tolist()
        
        self.inicializar_pools()

    def inicializar_pools(self):
        # Exclusión explícita de Rafael Torrealba de Tesoros y Vida Cristiana
        candidatos_tesoros = [v for v in self.todos_varones if v != 'Rafael Torrealba']
        random.shuffle(candidatos_tesoros)
        self.pool_tesoros = list(candidatos_tesoros)
        
        candidatos_vida = [v for v in (self.ancianos + self.siervos) if v != 'Rafael Torrealba']
        random.shuffle(candidatos_vida)
        self.pool_vida = list(candidatos_vida)

    def es_menor(self, nombre):
        match = self.df[self.df['Nombre'] == nombre]
        if not match.empty:
            return match.iloc[0].get('Es_Menor', 'No') == 'Si'
        return False
        
    def comparten_apellido(self, nombre1, nombre2):
        if not nombre1 or not nombre2 or nombre1 == "__________________" or nombre2 == "__________________":
            return False
        ap1 = nombre1.split()[-1].lower() if len(nombre1.split()) > 1 else nombre1.lower()
        ap2 = nombre2.split()[-1].lower() if len(nombre2.split()) > 1 else nombre2.lower()
        return ap1 == ap2

    def candidato_valido_maestros(self, persona, semana_num):
        ultima = self.historial_maestros.get(persona, -99)
        return (semana_num - ultima) >= 6

    def asignar_desde_pool(self, pool, lista_candidatos_base, asignados_semana, evitar=None):
        disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
        if not disponibles:
            base = [v for v in lista_candidatos_base if v != 'Rafael Torrealba']
            random.shuffle(base)
            pool.extend(base)
            disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
            
        if not disponibles:
            return "__________________"
            
        seleccionado = disponibles[0]
        pool.remove(seleccionado)
        asignados_semana.add(seleccionado)
        return seleccionado

    def asignar_persona(self, lista_candidatos, asignados_semana, evitar=None):
        disponibles = [p for p in lista_candidatos if p not in asignados_semana and p != evitar]
        if not disponibles:
            return "__________________"
        seleccionado = random.choice(disponibles)
        asignados_semana.add(seleccionado)
        return seleccionado

    def asignar_pareja_generica(self, asignados_semana, semana_num, pool_titulares, pool_ayudantes_base, permite_familiar_opuesto=False):
        titulares_disp = [t for t in pool_titulares if t not in asignados_semana and self.candidato_valido_maestros(t, semana_num)]
        if not titulares_disp:
            titulares_disp = [t for t in pool_titulares if t not in asignados_semana]
            
        if not titulares_disp: 
            return "__________________ // __________________"
        titular = random.choice(titulares_disp)
        
        pool_ayudantes = list(pool_ayudantes_base)
        if permite_familiar_opuesto and random.random() < 0.3:
            es_mujer = titular in self.hermanas
            pool_opuesto = self.todos_varones if es_mujer else self.hermanas
            fam_opuestos = [p for p in pool_opuesto if self.comparten_apellido(titular, p)]
            if fam_opuestos:
                pool_ayudantes.extend(fam_opuestos * 3)
        
        ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular and self.candidato_valido_maestros(a, semana_num)]
        if not ayudantes_disp:
            ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular]
            
        ayudantes_validos = []
        for a in ayudantes_disp:
            if self.es_menor(a) and not self.es_menor(titular):
                if not self.comparten_apellido(titular, a):
                    continue
            ayudantes_validos.append(a)
            
        if not ayudantes_validos:
            asignados_semana.add(titular)
            self.historial_maestros[titular] = semana_num
            return f"{titular} // __________________"
            
        ayudante = random.choice(ayudantes_validos)
        asignados_semana.add(titular)
        asignados_semana.add(ayudante)
        self.historial_maestros[titular] = semana_num
        self.historial_maestros[ayudante] = semana_num
        
        return f"{titular} // {ayudante}"

    def asignar_estudiante_solo_con_pool(self, asignados_semana, semana_num, pool):
        disp = [p for p in pool if p not in asignados_semana and self.candidato_valido_maestros(p, semana_num)]
        if not disp:
            disp = [p for p in pool if p not in asignados_semana]
        if not disp: 
            return "__________________"
        p = random.choice(disp)
        asignados_semana.add(p)
        self.historial_maestros[p] = semana_num
        return p

    def generar_semana(self, index_semana, fecha, lectura, asigs_maestros):
        asignados = set()
        
        # TESOROS (Pool)
        presidente = self.asignar_persona(self.ancianos, asignados)
        
        # Filtros de habilidades
        list_oracion = self.df[self.df['Hab_Oracion'] == 'Si']['Nombre'].tolist()
        oracion = self.asignar_persona(list_oracion if list_oracion else self.todos_varones, asignados)
        
        list_tes_discurso = self.df[(self.df['Hab_Tes_Discurso'] == 'Si') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        list_tes_perlas = self.df[(self.df['Hab_Tes_Perlas'] == 'Si') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        list_lectores_biblia = self.df[self.df['Hab_Lectura'] == 'Si']['Nombre'].tolist()
        
        num1_tesoros = self.asignar_desde_pool(self.pool_tesoros, list_tes_discurso if list_tes_discurso else self.pool_tesoros, asignados)
        num2_tesoros = self.asignar_desde_pool(self.pool_tesoros, list_tes_perlas if list_tes_perlas else self.pool_tesoros, asignados)
        lectura_biblia = self.asignar_desde_pool(self.pool_tesoros, list_lectores_biblia if list_lectores_biblia else self.pool_tesoros, asignados)
        
        # MINISTERIO (Maestros)
        maestros_asignaciones = []
        for asig in asigs_maestros:
            asig_lower = asig.lower()
            
            # Habilidades específicas
            if "convers" in asig_lower:
                p_enc = self.df[self.df['Hab_Mae_Conversacion_Enc'] == 'Si']['Nombre'].tolist()
                p_ayu = self.df[self.df['Hab_Mae_Conversacion_Ayu'] == 'Si']['Nombre'].tolist()
            elif "revisita" in asig_lower:
                p_enc = self.df[self.df['Hab_Mae_Revisita_Enc'] == 'Si']['Nombre'].tolist()
                p_ayu = self.df[self.df['Hab_Mae_Revisita_Ayu'] == 'Si']['Nombre'].tolist()
            elif "discipulo" in asig_lower or "discípulo" in asig_lower:
                p_enc = self.df[self.df['Hab_Mae_Discipulos_Enc'] == 'Si']['Nombre'].tolist()
                p_ayu = self.df[self.df['Hab_Mae_Discipulos_Ayu'] == 'Si']['Nombre'].tolist()
            elif "escenificaci" in asig_lower:
                p_enc = self.df[self.df['Hab_Mae_Creencias_Esc_Enc'] == 'Si']['Nombre'].tolist()
                p_ayu = self.df[self.df['Hab_Mae_Creencias_Esc_Ayu'] == 'Si']['Nombre'].tolist()
            else:
                p_enc, p_ayu = None, None

            if p_enc:
                usar_hermanas = (random.random() < 0.8) and bool(self.hermanas)
                f_enc = [n for n in p_enc if (n in self.hermanas if usar_hermanas else n in self.todos_varones)]
                f_ayu = [n for n in p_ayu if (n in self.hermanas if usar_hermanas else n in self.todos_varones)]
                if not f_enc: f_enc = p_enc
                if not f_ayu: f_ayu = p_ayu
                
                res = self.asignar_pareja_generica(asignados, index_semana, f_enc, f_ayu, permite_familiar_opuesto=True)
                maestros_asignaciones.append((asig, res))
            elif "creencias" in asig_lower:
                pool = self.df[self.df['Hab_Mae_Creencias_Dis'] == 'Si']['Nombre'].tolist()
                res = self.asignar_estudiante_solo_con_pool(asignados, index_semana, pool if pool else self.todos_varones)
                maestros_asignaciones.append((asig, res))
            else:
                pool = self.df[self.df['Hab_Mae_Discurso'] == 'Si']['Nombre'].tolist()
                res = self.asignar_estudiante_solo_con_pool(asignados, index_semana, pool if pool else self.todos_varones)
                maestros_asignaciones.append((asig, res))
                
        # VIDA CRISTIANA (Pool)
        list_vida_p1 = self.df[self.df['Hab_Vida_Parte1'] == 'Si']['Nombre'].tolist()
        num1_vida = self.asignar_desde_pool(self.pool_vida, list_vida_p1 if list_vida_p1 else self.pool_vida, asignados)
        
        list_estudio_cond = self.df[self.df['Hab_Estudio_Conductor'] == 'Si']['Nombre'].tolist()
        list_estudio_lect = self.df[self.df['Hab_Estudio_Lector'] == 'Si']['Nombre'].tolist()
        
        estudio_biblico = self.asignar_persona(list_estudio_cond if list_estudio_cond else self.ancianos, asignados)
        lector = self.asignar_persona(list_estudio_lect if list_estudio_lect else self.todos_varones, asignados, evitar=estudio_biblico)
        presidencia_aux = self.asignar_persona(self.ancianos, asignados)
        
        # SERVICIOS
        list_sonido = self.df[self.df['Hab_Sonido'] == 'Si']['Nombre'].tolist()
        list_plat = self.df[self.df['Hab_Plataforma'] == 'Si']['Nombre'].tolist()
        list_mics = self.df[self.df['Hab_Mics'] == 'Si']['Nombre'].tolist()
        list_acom = self.df[self.df['Hab_Acomodador'] == 'Si']['Nombre'].tolist()
        
        sonido = self.asignar_persona(list_sonido if list_sonido else self.todos_varones, asignados)
        plataforma = self.asignar_persona(list_plat if list_plat else self.todos_varones, asignados)
        mic1 = self.asignar_persona(list_mics if list_mics else self.todos_varones, asignados)
        mic2 = self.asignar_persona(list_mics if list_mics else self.todos_varones, asignados, evitar=mic1)
        acomodador1 = self.asignar_persona(list_acom if list_acom else self.todos_varones, asignados)
        acomodador2 = self.asignar_persona(list_acom if list_acom else self.todos_varones, asignados, evitar=acomodador1)
        
        return {
            'semana': index_semana,
            'fecha': fecha,
            'lectura': lectura,
            'presidente': presidente,
            'oracion': oracion,
            'num1_tesoros': num1_tesoros,
            'num2_tesoros': num2_tesoros,
            'lectura_biblia': lectura_biblia,
            'maestros': maestros_asignaciones,
            'num1_vida': num1_vida,
            'estudio_biblico': estudio_biblico,
            'lector': lector,
            'presidencia_aux': presidencia_aux,
            'sonido': sonido,
            'plataforma': plataforma,
            'microfonos': f"{mic1} / {mic2}",
            'acomodadores': f"{acomodador1} / {acomodador2}",
            'total_asignados': len(asignados)
        }


# ==========================================
# MOTOR DE HUMANIZACIÓN DE MENSAJES
# ==========================================
def humanizar_mensaje(nombre, parte, fecha, seccion):
    """Genera una versión humanizada aleatoria para evitar detección de bots y sonar cálido"""
    saludos = [
        "Hola {nombre}, 🙂",
        "Hola, hermano {nombre}. Espero que te encuentres muy bien. 🌟",
        "Saludos, {nombre} 👋",
        "¡Hola, {nombre}! Qué gusto saludarte. 😊",
        "Buen día, {nombre}. ✨",
        "Hola {nombre}, ¿cómo estás? Espero que todo esté excelente. 🤗",
        "Estimado {nombre}, un saludo muy cordial. 🌸",
        "Hola {nombre}. ¡Espero que tengas un feliz día! ☀️",
        "Saludos cariñosos, {nombre}. ❤️"
    ]
    
    introducciones = [
        "Te informamos que tienes la siguiente asignación para la reunión:",
        "Te escribo para comentarte que tienes la siguiente asignación:",
        "Paso por aquí a recordarte tu asignación programada:",
        "Te comparto los detalles de tu asignación para esta semana:",
        "Aquí tienes la información sobre tu parte en la reunión:",
        "Queremos notificarte de tu participación para la reunión de esta semana:",
        "Te dejamos por acá los detalles de tu asignación VMC:"
    ]
    
    despedidas = [
        "¡Muchas gracias por tu servicio! 🙏",
        "¡Muchas gracias por tu excelente disposición de siempre! 🙏",
        "Que Jehová bendiga mucho tu buena actitud y tus esfuerzos. ✨",
        "Agradecemos de corazón tu valioso apoyo. ¡Un gran abrazo! 🤗",
        "Gracias por tu buena disposición de siempre. 👍",
        "¡Que tengas una bendecida y excelente semana! Saludos. 😊",
        "Agradecemos mucho tu valioso servicio a favor de la congregación. 🤝",
        "Que Jehová te acompañe en tu preparación. ¡Saludos! 📖"
    ]
    
    saludo = random.choice(saludos).format(nombre=nombre)
    intro = random.choice(introducciones)
    despedida = random.choice(despedidas)
    
    disenos = [
        f"{saludo}\n\n{intro}\n\n📋 *{parte}*\n📅 *{fecha}*\n📖 Sección: {seccion}\n\n{despedida}",
        f"{saludo}\n\n{intro}\n\n👉 *{parte}*\n📅 _Fecha: {fecha}_\n📌 Sección: {seccion}\n\n{despedida}",
        f"{saludo}\n\n{intro}\n\n✨ Asignación: *{parte}*\n📅 Semana: *{fecha}*\n📖 Sección: *{seccion}*\n\n{despedida}"
    ]
    
    return random.choice(disenos)


# ==========================================
# DETECTOR DE SECCIÓN PARA EL WHATSAPP
# ==========================================
def obtener_seccion_parte(parte):
    parte_lower = parte.lower()
    if any(x in parte_lower for x in ["tesoro", "perlas", "lectura de la biblia", "num 1", "num 2", "num 3"]):
        return "💎 Tesoros de la Biblia"
    elif any(x in parte_lower for x in ["convers", "revis", "discip", "creencia", "discurso", "num 4", "num 5", "num 6"]):
        return "🤝 Seamos Mejores Maestros"
    else:
        return "🏠 Nuestra Vida Cristiana"


# ==========================================
# INTERFAZ DE NAVEGACIÓN Y CARGA DE DATOS
# ==========================================
st.sidebar.markdown("<h2 style='text-align: center; color: #4A148C;'>⚙️ Panel de Control</h2>", unsafe_allow_html=True)

# 1. Selector/Cargador de Archivo Roster
excel_cargado = False
roster_df = None

# Auto-detectar si ya existe el archivo en la carpeta
if os.path.exists(EXCEL_FILE):
    roster_df = cargar_roster_excel(EXCEL_FILE)
    excel_cargado = True
    st.sidebar.success("📂 Excel de Congregación cargado de forma automática.")
else:
    st.sidebar.warning("⚠️ No se encontró Congregacion_Araguaney.xlsx en la carpeta.")

# Permitir subir otro archivo si lo desea
uploaded_file = st.sidebar.file_uploader("Subir base de datos Excel (Opcional)", type=["xlsx"])
if uploaded_file is not None:
    roster_df = cargar_roster_excel(uploaded_file)
    excel_cargado = True
    st.sidebar.success("📂 Nuevo Excel cargado con éxito.")

# Detener ejecución si no hay datos
if not excel_cargado or roster_df is None:
    st.title("📅 Coordinación VMC - El Araguaney")
    st.info("👋 ¡Bienvenido a la Web App de Coordinación! Por favor, coloque el archivo `Congregacion_Araguaney.xlsx` en el directorio de la aplicación o súbalo en el menú de la izquierda para comenzar.")
    st.stop()

# Inicializar listas del roster para Selectboxes
ancianos_list = sorted(roster_df[(roster_df['Privilegio'] == 'Anciano') & (roster_df['Genero'] == 'M')]['Nombre'].tolist())
siervos_list = sorted(roster_df[(roster_df['Privilegio'] == 'Siervo Min.') & (roster_df['Genero'] == 'M')]['Nombre'].tolist())
varones_list = sorted(roster_df[roster_df['Genero'] == 'M']['Nombre'].tolist())
hermanas_list = sorted(roster_df[roster_df['Genero'] == 'F']['Nombre'].tolist())
todos_list = sorted(roster_df['Nombre'].tolist())

# Panel de Ajustes en Sidebar
st.sidebar.markdown("---")
st.sidebar.markdown("### 🛠️ Ajustes del Programador")
segundos_delay = st.sidebar.number_input("Demora de envío masivo (segundos):", min_value=1, max_value=20, value=3)
humanizar_active = st.sidebar.toggle("Humanizar Mensajes automáticamente", value=True)

# Plantilla Fija de Respaldo
template_respaldo = st.sidebar.text_area(
    "Plantilla fija (si se desactiva Humanizar):",
    value="Hola {nombre}, 🙂\n\nTe informamos que tienes la siguiente asignación:\n\n📋 *{parte}*\n📅 *{fecha}*\n📖 Sección: {seccion}\n\n¡Muchas gracias por tu servicio! 🙏",
    height=120
)

# Descargador JW.org
st.sidebar.markdown("---")
st.sidebar.markdown("### 📥 Sincronización JW.org")
current_year = datetime.now().year
año_seleccionado = st.sidebar.selectbox("Año:", [current_year - 1, current_year, current_year + 1], index=1)
bimestres_nombres = [b[0] for b in JWScraper.get_bimestres()] if JW_AVAILABLE else ["Mayo-Junio"]
nombre_actual_bim, _ = JWScraper.detectar_bimestre_actual() if JW_AVAILABLE else ("Mayo-Junio", 5)
bimestre_seleccionado = st.sidebar.selectbox("Bimestre:", bimestres_nombres, index=bimestres_nombres.index(nombre_actual_bim) if nombre_actual_bim in bimestres_nombres else 0)


# ==========================================
# CONTROL DEL ESTADO DE LA APLICACIÓN
# ==========================================
if "semanas_config" not in st.session_state:
    st.session_state["semanas_config"] = []
if "bimestre_data" not in st.session_state:
    st.session_state["bimestre_data"] = []
if "descargado_exito" not in st.session_state:
    st.session_state["descargado_exito"] = False


# ==========================================
# INTERFAZ PRINCIPAL - CABECERA Y METRICAS
# ==========================================
st.markdown("<h1 style='text-align: center; margin-bottom: 5px;'>📋 Vida y Ministerio Cristianos</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; color: #666; font-size: 1.1rem; margin-bottom: 25px;'>Herramienta Web Bimestral • Congregación 'El Araguaney'</p>", unsafe_allow_html=True)

# Dashboard de Roster
m_anc, m_sie, m_var, m_her = st.columns(4)
with m_anc:
    st.markdown(f"<div class='metric-card'><h4>👴 Ancianos</h4><h2>{len(ancianos_list)}</h2></div>", unsafe_allow_html=True)
with m_sie:
    st.markdown(f"<div class='metric-card'><h4>📖 Siervos Ministeriales</h4><h2>{len(siervos_list)}</h2></div>", unsafe_allow_html=True)
with m_var:
    st.markdown(f"<div class='metric-card'><h4>👨 Todos los Varones</h4><h2>{len(varones_list)}</h2></div>", unsafe_allow_html=True)
with m_her:
    st.markdown(f"<div class='metric-card'><h4>👩 Hermanas</h4><h2>{len(hermanas_list)}</h2></div>", unsafe_allow_html=True)

# Tabs Principales
tab_config, tab_editor, tab_exports = st.tabs([
    "🗓️ 1. Configurar y Sincronizar", 
    "✍️ 2. Vista Previa y Edición", 
    "💾 3. Descargar y Exportar"
])


# ==========================================
# TAB 1: CONFIGURACIÓN Y SINCRONIZACIÓN
# ==========================================
with tab_config:
    st.subheader("📥 Cargar Guía Teocrática")
    
    col_btn, col_status = st.columns([1, 2])
    with col_btn:
        if st.button("⬇ DESCARGAR DESDE JW.ORG", type="primary", use_container_width=True):
            if not JW_AVAILABLE:
                st.error("El módulo jw_scraper no está disponible en el servidor.")
            else:
                with st.spinner("Conectando con wol.jw.org y descargando semanas..."):
                    try:
                        scraper = JWScraper()
                        mes_inicio = 5
                        for nombre, mes in JWScraper.get_bimestres():
                            if nombre == bimestre_seleccionado:
                                mes_inicio = mes
                                break
                        
                        semanas_iso = JWScraper.calcular_semanas_bimestre(año_seleccionado, mes_inicio)
                        resultados = []
                        
                        for i, (iso_year, iso_week) in enumerate(semanas_iso):
                            try:
                                datos = scraper.obtener_semana(iso_year, iso_week)
                                resultados.append(datos)
                            except Exception as e:
                                resultados.append({
                                    'error': str(e), 'fecha': f'Semana {iso_week} (error)',
                                    'lectura_biblica': '', 'maestros': [],
                                })
                        
                        # Generar lista de configuración
                        semanas_config = []
                        for idx, r in enumerate(resultados):
                            if 'error' in r:
                                semanas_config.append({
                                    'numero': idx + 1,
                                    'fecha': f"Semana {idx+1}",
                                    'lectura': "",
                                    'maestros_parts': ["Empiece conversaciones", "Haga revisitas", "Haga discípulos"]
                                })
                            else:
                                # Extraer partes de maestros
                                partes = []
                                for m in r.get('maestros', []):
                                    partes.append(f"{m['numero']}. {m['tipo']} ({m['mins']} min)")
                                if not partes:
                                    partes = ["Empiece conversaciones", "Haga revisitas", "Haga discípulos"]
                                
                                semanas_config.append({
                                    'numero': idx + 1,
                                    'fecha': r.get('fecha', f"Semana {idx+1}"),
                                    'lectura': r.get('lectura_biblica', ""),
                                    'maestros_parts': partes
                                })
                                
                        st.session_state["semanas_config"] = semanas_config
                        st.session_state["descargado_exito"] = True
                        st.success(f"¡Sincronización Completada! Se descargaron {len(semanas_config)} semanas correctamente.")
                    except Exception as e:
                        st.error(f"Error al descargar la guía: {e}")
                        
    # Configuración Manual / Visualización de Semanas
    if st.session_state["descargado_exito"]:
        st.markdown("### 📋 Semanas Configuradas para el Bimestre")
        
        # Permitir cambiar u organizar partes antes de la generación
        for s in st.session_state["semanas_config"]:
            with st.expander(f"📅 Semana {s['numero']}: {s['fecha']}", expanded=True):
                col_f, col_l = st.columns(2)
                with col_f:
                    s['fecha'] = st.text_input(f"Fecha (Semana {s['numero']})", value=s['fecha'], key=f"fecha_{s['numero']}")
                with col_l:
                    s['lectura'] = st.text_input(f"Lectura Bíblica (Semana {s['numero']})", value=s['lectura'], key=f"lectura_{s['numero']}")
                
                # Checkboxes de Maestros
                st.write("**Asignaciones de la sección Seamos Mejores Maestros:**")
                checked_parts = []
                for p in s['maestros_parts']:
                    if st.checkbox(p, value=True, key=f"chk_{s['numero']}_{p}"):
                        checked_parts.append(p)
                s['maestros_parts_selected'] = checked_parts

        st.markdown("---")
        if st.button("📊 GENERAR ASIGNACIONES AUTOMÁTICAS", type="primary", use_container_width=True):
            with st.spinner("Ejecutando algoritmo de emparejamiento teocrático circular..."):
                try:
                    asignador = AsignadorVMC(roster_df)
                    resultados_bimestre = []
                    
                    for s in st.session_state["semanas_config"]:
                        parts_to_gen = s.get('maestros_parts_selected', s['maestros_parts'])
                        res_semana = asignador.generar_semana(s['numero'], s['fecha'], s['lectura'], parts_to_gen)
                        resultados_bimestre.append(res_semana)
                        
                    st.session_state["bimestre_data"] = resultados_bimestre
                    st.success("🎉 ¡Asignaciones del bimestre generadas con total éxito! Ve a la pestaña 'Edición' para previsualizar.")
                except Exception as e:
                    st.error(f"Error durante la generación: {e}")


# ==========================================
# TAB 2: VISTA PREVIA Y EDICIÓN DE CAMBIOS
# ==========================================
with tab_editor:
    if not st.session_state["bimestre_data"]:
        st.info("👋 Aún no has generado las asignaciones. Ve a la pestaña '1. Configurar y Sincronizar' para comenzar.")
    else:
        st.markdown("### ✍️ Ajustes manuales y Envío a WhatsApp")
        st.caption("Puedes cambiar cualquier asignación usando los selectores. Al lado de cada publicador tienes el botón de WhatsApp para enviarle la asignación individualmente.")
        
        # Enviar masivo asíncrono simulado
        st.write("---")
        st.write("**📱 Envío Masivo Automatizado**")
        if st.button("📱 ABRIR ENVIADOR MASIVO", type="secondary"):
            # Buscar todos los que tienen teléfono
            con_tel = []
            para_enviar = []
            for sem in st.session_state["bimestre_data"]:
                # Agregar presidente
                con_tel.append((sem['presidente'], f"Presidente de la reunión del {sem['fecha']}", sem['fecha']))
                # Oración
                con_tel.append((sem['oracion'], f"Oración de la reunión del {sem['fecha']}", sem['fecha']))
                # Tesoros
                con_tel.append((sem['num1_tesoros'], f"Parte 1: Perlas de la Biblia ({sem['lectura']})", sem['fecha']))
                con_tel.append((sem['num2_tesoros'], f"Parte 2: Perlas de la Biblia ({sem['lectura']})", sem['fecha']))
                con_tel.append((sem['lectura_biblia'], f"Lectura de la Biblia", sem['fecha']))
                # Maestros
                for tipo, val in sem['maestros']:
                    if "//" in val:
                        partes = val.split("//")
                        con_tel.append((partes[0].strip(), tipo, sem['fecha']))
                        con_tel.append((partes[1].strip(), f"Ayudante de: {tipo}", sem['fecha']))
                    else:
                        con_tel.append((val, tipo, sem['fecha']))
                # Vida
                con_tel.append((sem['num1_vida'], f"Primera sección de Vida Cristiana", sem['fecha']))
                con_tel.append((sem['estudio_biblico'], f"Estudio Bíblico de Congregación", sem['fecha']))
                con_tel.append((sem['lector'], f"Lector del Estudio Bíblico", sem['fecha']))
            
            # Filtrar válidos con teléfono
            validos = []
            for nombre, parte, fecha in con_tel:
                if nombre and nombre != "__________________" and "______" not in nombre:
                    match = roster_df[roster_df['Nombre'] == nombre]
                    if not match.empty:
                        tel = str(match.iloc[0].get('Telefono', '')).strip()
                        if tel and tel.lower() not in ['nan', 'none', '']:
                            validos.append({'nombre': nombre, 'tel': tel, 'parte': parte, 'fecha': fecha})
            
            if not validos:
                st.warning("No se encontraron números de teléfono asignados válidos en el Excel.")
            else:
                st.success(f"Se detectaron {len(validos)} asignaciones con teléfonos cargados.")
                
                # Barra de progreso masivo
                progress_bar = st.progress(0)
                status_txt = st.empty()
                
                for idx, v in enumerate(validos):
                    secc = obtener_seccion_parte(v['parte'])
                    if humanizar_active:
                        msg = humanizar_mensaje(v['nombre'], v['parte'], v['fecha'], secc)
                    else:
                        msg = template_respaldo.format(nombre=v['nombre'], parte=v['parte'], fecha=v['fecha'], seccion=secc)
                    
                    # Generar URLs
                    tel_clean = "".join(filter(str.isdigit, v['tel']))
                    texto_enc = urllib.parse.quote(msg)
                    url_app = f"whatsapp://send?phone={tel_clean}&text={texto_enc}"
                    url_web = f"https://web.whatsapp.com/send?phone={tel_clean}&text={texto_enc}"
                    
                    # Mostrar card interactiva para abrir los chats de uno en uno de forma humana controlable
                    with st.chat_message("user", avatar="📱"):
                        st.markdown(f"**Enviar a:** {v['nombre']} • **Parte:** {v['parte']}")
                        col_a, col_b = st.columns(2)
                        with col_a:
                            st.markdown(f"<a href='{url_app}' target='_blank' class='whatsapp-btn'>📱 Enviar por WhatsApp App</a>", unsafe_allow_html=True)
                        with col_b:
                            st.markdown(f"<a href='{url_web}' target='_blank' class='whatsapp-web-btn'>💻 WhatsApp Web</a>", unsafe_allow_html=True)
                    
                    progress_bar.progress((idx + 1) / len(validos))
                    status_txt.write(f"Procesando {idx+1}/{len(validos)}: {v['nombre']}")
        
        st.write("---")
        
        # Grid interactivo por cada semana
        for w_idx, sem in enumerate(st.session_state["bimestre_data"]):
            st.markdown(f"<div style='background-color:#E8EAF6; padding:10px; border-radius:10px; margin-top:20px;'><h3 style='margin:0;'>📅 Semana {w_idx+1}: {sem['fecha']} • {sem['lectura']}</h3></div>", unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("<h4 style='color:#808080;'>💎 TESOROS</h4>", unsafe_allow_html=True)
                
                # Presidente
                sem['presidente'] = st.selectbox("Presidente:", ancianos_list, index=ancianos_list.index(sem['presidente']) if sem['presidente'] in ancianos_list else 0, key=f"pres_{w_idx}")
                # Oración
                sem['oracion'] = st.selectbox("Oración:", varones_list, index=varones_list.index(sem['oracion']) if sem['oracion'] in varones_list else 0, key=f"or_{w_idx}")
                # Tesoros 1
                sem['num1_tesoros'] = st.selectbox("Tesoro 1:", varones_list, index=varones_list.index(sem['num1_tesoros']) if sem['num1_tesoros'] in varones_list else 0, key=f"t1_{w_idx}")
                # Tesoros 2
                sem['num2_tesoros'] = st.selectbox("Tesoro 2:", varones_list, index=varones_list.index(sem['num2_tesoros']) if sem['num2_tesoros'] in varones_list else 0, key=f"t2_{w_idx}")
                # Lectura Biblia
                sem['lectura_biblia'] = st.selectbox("Lectura Biblia:", varones_list, index=varones_list.index(sem['lectura_biblia']) if sem['lectura_biblia'] in varones_list else 0, key=f"lb_{w_idx}")
                
            with col2:
                st.markdown("<h4 style='color:#FF8C00;'>🤝 MINISTERIO</h4>", unsafe_allow_html=True)
                
                # Lista de maestros editables
                nuevas_maestros = []
                for m_idx, (tipo, val) in enumerate(sem['maestros']):
                    if "//" in val:
                        # Pareja
                        t, a = val.split("//")
                        t, a = t.strip(), a.strip()
                        
                        col_t, col_a = st.columns(2)
                        with col_t:
                            new_t = st.selectbox(f"{tipo} (Titular):", todos_list, index=todos_list.index(t) if t in todos_list else 0, key=f"t_mae_{w_idx}_{m_idx}")
                        with col_a:
                            new_a = st.selectbox(f"{tipo} (Ayudante):", todos_list, index=todos_list.index(a) if a in todos_list else 0, key=f"a_mae_{w_idx}_{m_idx}")
                        
                        nuevas_maestros.append((tipo, f"{new_t} // {new_a}"))
                    else:
                        # Estudiante solo
                        new_val = st.selectbox(f"{tipo}:", varones_list, index=varones_list.index(val) if val in varones_list else 0, key=f"s_mae_{w_idx}_{m_idx}")
                        nuevas_maestros.append((tipo, new_val))
                sem['maestros'] = nuevas_maestros
                
            with col3:
                st.markdown("<h4 style='color:#B22222;'>🏠 VIDA CRISTIANA</h4>", unsafe_allow_html=True)
                
                # Parte 1
                sem['num1_vida'] = st.selectbox("Vida Parte 1:", varones_list, index=varones_list.index(sem['num1_vida']) if sem['num1_vida'] in varones_list else 0, key=f"v1_{w_idx}")
                # Conductor Estudio
                sem['estudio_biblico'] = st.selectbox("Conductor Estudio:", ancianos_list, index=ancianos_list.index(sem['estudio_biblico']) if sem['estudio_biblico'] in ancianos_list else 0, key=f"cond_{w_idx}")
                # Lector Estudio
                sem['lector'] = st.selectbox("Lector Estudio:", varones_list, index=varones_list.index(sem['lector']) if sem['lector'] in varones_list else 0, key=f"lect_{w_idx}")
                # Consejero
                sem['presidencia_aux'] = st.selectbox("Consejero Auxiliar:", ancianos_list, index=ancianos_list.index(sem['presidencia_aux']) if sem['presidencia_aux'] in ancianos_list else 0, key=f"cons_{w_idx}")
                
            # Fila de servicios
            with st.expander("⚙️ Servicios y Logística de la semana"):
                c_s1, c_s2, c_s3, c_s4 = st.columns(4)
                with c_s1:
                    sem['sonido'] = st.selectbox("Sonido:", varones_list, index=varones_list.index(sem['sonido']) if sem['sonido'] in varones_list else 0, key=f"son_{w_idx}")
                with c_s2:
                    sem['plataforma'] = st.selectbox("Plataforma:", varones_list, index=varones_list.index(sem['plataforma']) if sem['plataforma'] in varones_list else 0, key=f"plat_{w_idx}")
                with c_s3:
                    sem['microfonos'] = st.text_input("Micrófonos:", value=sem['microfonos'], key=f"mics_{w_idx}")
                with c_s4:
                    sem['acomodadores'] = st.text_input("Acomodadores:", value=sem['acomodadores'], key=f"acom_{w_idx}")


# ==========================================
# TAB 3: EXPORTACIONES EXCEL / PDF / S-89
# ==========================================
with tab_exports:
    if not st.session_state["bimestre_data"]:
        st.info("👋 Aún no has generado las asignaciones. Ve a la pestaña '1. Configurar y Sincronizar' para comenzar.")
    else:
        st.markdown("### 💾 Descarga tus documentos terminados")
        st.caption("Todos los archivos son generados en tiempo real con las modificaciones que hayas hecho en la pestaña de edición.")
        
        # 1. GENERACIÓN DE EXCEL PROFESIONAL
        def generate_excel_in_memory():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bimestre VMC"
            
            # Anchos
            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 48
            ws.column_dimensions['C'].width = 40
            
            font_title = Font(name='Arial', size=18, bold=True)
            font_date = Font(name='Arial', size=12, bold=True)
            font_header_blanco = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            font_normal = Font(name='Arial', size=10)
            
            fill_purple = PatternFill(start_color='5E005E', end_color='5E005E', fill_type='solid')
            fill_tesoros = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            fill_maestros = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
            fill_vida = PatternFill(start_color='B22222', end_color='B22222', fill_type='solid')
            fill_light_grey = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
            
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                                 
            current_row = 1
            for sem in st.session_state["bimestre_data"]:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell_title = ws.cell(row=current_row, column=1, value=f"VIDA Y MINISTERIO CRISTIANOS  |  {sem['lectura'].upper()}")
                cell_title.font = font_title
                cell_title.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_date = ws.cell(row=current_row, column=3, value=sem['fecha'])
                cell_date.font = font_date
                cell_date.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
                
                # Franja morada
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).fill = fill_purple
                current_row += 1
                
                # Presidente / Oración
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell_pres = ws.cell(row=current_row, column=1, value=f"PRESIDENTE DE LA REUNIÓN: {sem['presidente']}")
                cell_pres.font = font_normal
                cell_pres.border = border_thin
                
                cell_ora = ws.cell(row=current_row, column=3, value=f"ORACIÓN: {sem['oracion']}")
                cell_ora.font = font_normal
                cell_ora.border = border_thin
                current_row += 1
                
                # Encabezados
                headers = [("TESOROS DE LA BIBLIA", fill_tesoros), ("SEAMOS MEJORES MAESTROS", fill_maestros), ("NUESTRA VIDA CRISTIANA", fill_vida)]
                for col, (text, fill) in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col, value=text)
                    cell.font = font_header_blanco
                    cell.fill = fill
                    cell.border = border_thin
                current_row += 1
                
                # Datos
                col_t = [f"NUM 1: {sem['num1_tesoros']}", f"NUM 2: {sem['num2_tesoros']}", f"NUM 3: {sem['lectura_biblia']}"]
                col_m = []
                num_actual = 4
                for tipo, val in sem['maestros']:
                    col_m.append(f"NUM {num_actual}: {val}")
                    num_actual += 1
                col_v = [f"NUM {num_actual}: {sem['num1_vida']}", f"NUM {num_actual+1}: (Estudio) {sem['estudio_biblico']}", f"LECTOR ESTUDIO: {sem['lector']}"]
                
                max_r = max(len(col_t), len(col_m), len(col_v))
                for i in range(max_r):
                    ws.cell(row=current_row, column=1, value=col_t[i] if i < len(col_t) else "").border = border_thin
                    ws.cell(row=current_row, column=2, value=col_m[i] if i < len(col_m) else "").border = border_thin
                    ws.cell(row=current_row, column=3, value=col_v[i] if i < len(col_v) else "").border = border_thin
                    current_row += 1
                    
                # Servicios
                servs = [
                    (f"Sonido: {sem['sonido']}", f"Plataforma: {sem['plataforma']}", f"Acomodadores: {sem['acomodadores']}"),
                    (f"Cons Aux: {sem['presidencia_aux']}", f"Micrófonos: {sem['microfonos']}", "")
                ]
                for r in servs:
                    for col, val in enumerate(r, start=1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        cell.font = Font(name='Arial', size=9)
                        cell.fill = fill_light_grey
                        cell.border = border_thin
                    current_row += 1
                current_row += 2 # Separación
                
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return out
            
        # 2. GENERACIÓN DE PDF COMPLETO
        def generate_pdf_in_memory():
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            story = []
            
            for sem in st.session_state["bimestre_data"]:
                data = []
                data.append([f"VIDA Y MINISTERIO CRISTIANOS", "", f"{sem['fecha']}  |  {sem['lectura']}"])
                data.append(["", "", ""])
                data.append([f"PRESIDENTE DE LA REUNIÓN: {sem['presidente']}", "", f"ORACIÓN: {sem['oracion']}"])
                data.append(["TESOROS DE LA BIBLIA", "SEAMOS MEJORES MAESTROS", "NUESTRA VIDA CRISTIANA"])
                
                col_t = [f"NUM 1: {sem['num1_tesoros']}", f"NUM 2: {sem['num2_tesoros']}", f"NUM 3: {sem['lectura_biblia']}"]
                col_m = []
                num_actual = 4
                for tipo, val in sem['maestros']:
                    col_m.append(f"NUM {num_actual}: {val}")
                    num_actual += 1
                col_v = [f"NUM {num_actual}: {sem['num1_vida']}", f"NUM {num_actual+1}: (Estudio) {sem['estudio_biblico']}", f"LECTOR ESTUDIO: {sem['lector']}"]
                
                max_r = max(len(col_t), len(col_m), len(col_v))
                for i in range(max_r):
                    data.append([col_t[i] if i < len(col_t) else "", col_m[i] if i < len(col_m) else "", col_v[i] if i < len(col_v) else ""])
                    
                data.append([f"Sonido: {sem['sonido']}", f"Plataforma: {sem['plataforma']}", f"Acomodadores: {sem['acomodadores']}"])
                data.append([f"Cons Aux: {sem['presidencia_aux']}", f"Micrófonos: {sem['microfonos']}", ""])
                
                table = Table(data, colWidths=[3.5*inch, 3.5*inch, 3.5*inch])
                table.setStyle(TableStyle([
                    ('SPAN', (0,0), (1,0)),
                    ('FONTNAME', (0,0), (1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (1,0), 16),
                    ('ALIGN', (0,0), (1,0), 'CENTER'),
                    ('ALIGN', (2,0), (2,0), 'RIGHT'),
                    ('SPAN', (0,1), (2,1)),
                    ('BACKGROUND', (0,1), (2,1), colors.HexColor('#5E005E')),
                    ('SPAN', (0,2), (1,2)),
                    ('FONTNAME', (0,2), (-1,2), 'Helvetica'),
                    ('FONTSIZE', (0,2), (-1,2), 10),
                    ('BACKGROUND', (0,3), (0,3), colors.HexColor('#808080')),
                    ('BACKGROUND', (1,3), (1,3), colors.HexColor('#FF8C00')),
                    ('BACKGROUND', (2,3), (2,3), colors.HexColor('#B22222')),
                    ('TEXTCOLOR', (0,3), (-1,3), colors.whitesmoke),
                    ('FONTNAME', (0,3), (-1,3), 'Helvetica-Bold'),
                    ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor('#EAEAEA')),
                    ('GRID', (0,2), (-1,-1), 0.5, colors.black),
                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                ]))
                story.append(table)
                story.append(Spacer(1, 0.4*inch))
                
            doc.build(story)
            buffer.seek(0)
            return buffer

        # 3. GENERACIÓN DE VALES S-89
        def draw_s89(c, x, y, w, h, asig, fecha):
            c.saveState()
            c.setLineWidth(0.8)
            c.rect(x, y, w, h)
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(x + w/2, y + h - 0.8*cm, "ASIGNACIÓN PARA LA REUNIÓN")
            c.drawCentredString(x + w/2, y + h - 1.3*cm, "VIDA Y MINISTERIO CRISTIANOS")
            
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x + 0.6*cm, y + h - 2.5*cm, "Nombre:")
            c.drawString(x + 0.6*cm, y + h - 3.5*cm, "Ayudante:")
            c.drawString(x + 0.6*cm, y + h - 4.5*cm, "Fecha:")
            c.drawString(x + 0.6*cm, y + h - 5.8*cm, "Intervención núm.:")
            
            c.setFont("Helvetica", 10)
            c.drawString(x + 2.3*cm, y + h - 2.5*cm, asig['nombre'])
            c.drawString(x + 2.3*cm, y + h - 3.5*cm, asig['ayudante'] or "---------------------------------------")
            c.drawString(x + 2.3*cm, y + h - 4.5*cm, fecha)
            c.drawString(x + 4.0*cm, y + h - 5.8*cm, asig['num'])
            
            c.setDash(1, 2)
            c.line(x + 2.2*cm, y + h - 2.6*cm, x + w - 0.6*cm, y + h - 2.6*cm)
            c.line(x + 2.2*cm, y + h - 3.6*cm, x + w - 0.6*cm, y + h - 3.6*cm)
            c.line(x + 2.2*cm, y + h - 4.6*cm, x + w - 0.6*cm, y + h - 4.6*cm)
            c.line(x + 3.9*cm, y + h - 5.9*cm, x + w - 0.6*cm, y + h - 5.9*cm)
            c.setDash([])
            
            c.setFont("Helvetica-Oblique", 11)
            tipo_txt = re.sub(r'\(.*?\)', '', asig['tipo']).strip()
            c.drawCentredString(x + w/2, y + h - 7.0*cm, f"( {tipo_txt} )")
            
            c.setFont("Helvetica-Bold", 9.5)
            c.drawString(x + 0.6*cm, y + h - 8.0*cm, "Se presentará en:")
            c.rect(x + 1.2*cm, y + h - 8.6*cm, 0.35*cm, 0.35*cm)
            c.drawString(x + 1.3*cm, y + h - 8.5*cm, "✓")
            c.drawString(x + 1.8*cm, y + h - 8.6*cm, "Sala principal")
            
            c.rect(x + 1.2*cm, y + h - 9.2*cm, 0.35*cm, 0.35*cm)
            c.drawString(x + 1.8*cm, y + h - 9.2*cm, "Sala auxiliar núm. 1")
            
            c.rect(x + 1.2*cm, y + h - 9.8*cm, 0.35*cm, 0.35*cm)
            c.drawString(x + 1.8*cm, y + h - 9.8*cm, "Sala auxiliar núm. 2")
            
            # Nota
            style = getSampleStyleSheet()['BodyText']
            style.fontSize = 8
            style.leading = 9
            texto_nota = "<b>Nota al estudiante:</b> En la <i>Guía de actividades</i> encontrará la información que necesita. Repase las indicaciones S-38."
            p = Paragraph(texto_nota, style)
            p.wrap(w - 1.2*cm, 3*cm)
            p.drawOn(c, x + 0.6*cm, y + 1.6*cm)
            
            c.setFont("Helvetica", 7.5)
            c.drawString(x + 0.6*cm, y + 0.6*cm, "S-89-S 11/23")
            c.restoreState()

        def generate_s89_in_memory():
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            vale_w, vale_h = 10.5*cm, 13.5*cm
            margin_x = (width - 2*vale_w) / 3
            margin_y = (height - 2*vale_h) / 3
            
            vales_por_p = 4
            count = 0
            
            for sem in st.session_state["bimestre_data"]:
                fecha = sem['fecha']
                asigs = []
                
                # Lectura Biblia
                if sem['lectura_biblia'] and sem['lectura_biblia'] != "__________________":
                    asigs.append({'nombre': sem['lectura_biblia'], 'ayudante': '', 'num': '3', 'tipo': 'Lectura de la Biblia'})
                
                # Maestros
                for idx, (tipo_orig, val) in enumerate(sem['maestros']):
                    num = str(idx+4)
                    nombre = val
                    ayudante = ""
                    if "//" in val:
                        partes = val.split("//")
                        nombre, ayudante = partes[0].strip(), partes[1].strip()
                    
                    if nombre and nombre != "__________________":
                        asigs.append({'nombre': nombre, 'ayudante': ayudante, 'num': num, 'tipo': tipo_orig})
                
                for asig in asigs:
                    if count >= vales_por_p:
                        c.showPage()
                        count = 0
                    col = count % 2
                    row = 1 - (count // 2)
                    x_pos = margin_x + col * (vale_w + margin_x)
                    y_pos = margin_y + row * (vale_h + margin_y)
                    
                    draw_s89(c, x_pos, y_pos, vale_w, vale_h, asig, fecha)
                    count += 1
                    
            c.save()
            buffer.seek(0)
            return buffer

        # Grid de Descargas
        col_excel, col_pdf, col_s89 = st.columns(3)
        
        with col_excel:
            st.info("📊 **Excel del Bimestre**\n\nPrograma completo en 3 columnas estilizadas.")
            excel_data = generate_excel_in_memory()
            st.download_button(
                label="⬇️ Descargar Excel (.xlsx)",
                data=excel_data,
                file_name=f"bimestre_vmc_{bimestre_seleccionado.replace('-','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with col_pdf:
            st.info("📄 **PDF de la Reunión**\n\nFormato idéntico al Excel, listo para imprimir.")
            pdf_data = generate_pdf_in_memory()
            st.download_button(
                label="⬇️ Descargar PDF de Reunión",
                data=pdf_data,
                file_name=f"programa_vmc_{bimestre_seleccionado.replace('-','_')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
            
        with col_s89:
            st.info("🎟️ **Vales S-89 Impresores**\n\nBoletas individuales S-89 listas para repartir.")
            s89_data = generate_s89_in_memory()
            st.download_button(
                label="⬇️ Descargar Vales S-89 (.pdf)",
                data=s89_data,
                file_name=f"vales_S89_{bimestre_seleccionado.replace('-','_')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
