import customtkinter as ctk
import pandas as pd
import random
from tkinter import messagebox, filedialog, ttk
import os
from datetime import datetime
import threading
try:
    from jw_scraper import JWScraper
    JW_AVAILABLE = True
except ImportError:
    JW_AVAILABLE = False
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

# Configuración
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class CoordinacionVMC:
    def __init__(self, root):
        self.root = root
        self.root.title("Coordinación VMC - El Araguaney (Bimestral)")
        self.root.geometry("1300x800")
        
        # Datos de la congregación
        self.df = None
        self.ancianos = []
        self.siervos = []
        self.publicadores_varones = []
        self.hermanas = []
        self.todos_varones = []
        
        # Datos del bimestre
        self.semanas = []
        self.bimestre_data = []
        
        # Historial y reglas globales
        self.historial_maestros = {}
        self.pool_tesoros = []
        self.pool_vida = []
        
        # Datos de JW.org
        self.datos_jw = []
        self.scraper = JWScraper() if JW_AVAILABLE else None
        
        # Cargar datos
        self.cargar_datos()
        
        # Crear interfaz
        self.crear_widgets()
        
    def cargar_datos(self):
        """Carga el archivo Excel de la congregación"""
        try:
            # Construir la ruta absoluta basada en la ubicación del script
            directorio_script = os.path.dirname(os.path.abspath(__file__))
            archivo = os.path.join(directorio_script, "Congregacion_Araguaney.xlsx")
            
            if not os.path.exists(archivo):
                self.crear_datos_ejemplo()
                return
            
            self.df = pd.read_excel(archivo)
            
            # 1. Exclusiones: Eliminar familia Spolzino y Saucedo
            self.df = self.df[~self.df['Nombre'].str.contains('Spolzino|Saucedo', case=False, na=False)]
            
            # 2. Agregar columna Es_Menor si no existe
            if 'Es_Menor' not in self.df.columns:
                self.df['Es_Menor'] = 'No'
            
            self.df['Es_Menor'] = self.df['Es_Menor'].astype(str).str.strip().str.title()
            
            self.ancianos = self.df[(self.df['Privilegio'] == 'Anciano') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.siervos = self.df[(self.df['Privilegio'] == 'Siervo Min.') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.publicadores_varones = self.df[(self.df['Privilegio'] == 'Publicador') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.hermanas = self.df[self.df['Genero'] == 'F']['Nombre'].tolist()
            self.todos_varones = self.ancianos + self.siervos + self.publicadores_varones
            
            messagebox.showinfo("Éxito", f"Datos cargados:\n👴 Ancianos: {len(self.ancianos)}\n📖 Siervos: {len(self.siervos)}\n👨 Varones: {len(self.publicadores_varones)}\n👩 Hermanas: {len(self.hermanas)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            
    def crear_datos_ejemplo(self):
        """Crea datos de ejemplo"""
        datos = {
            'Nombre': ['Gilmer González', 'David Cordero', 'Raúl Cordero', 'Jhovanny Suárez',
                       'Gonzalo Sayago', 'Engelberth Oviol',
                       'Javier Alvarado', 'Gilmer de Jesús González', 'Rafael Torrealba',
                       'Willians Sivira', 'Samuel Sivira', 'Alís Fernández'],
            'Privilegio': ['Anciano','Anciano','Anciano','Anciano','Anciano','Anciano',
                          'Siervo Min.','Siervo Min.','Siervo Min.',
                          'Publicador','Publicador','Publicador'],
            'Genero': ['M','M','M','M','M','M','M','M','M','M','M','M'],
            'Es_Menor': ['No','No','No','No','No','No','No','No','No','No','Si','Si']
        }
        # Agregar hermanas
        for i in range(20):
            datos['Nombre'].append(f'Hermana {i+1}')
            datos['Privilegio'].append('Publicador')
            datos['Genero'].append('F')
            datos['Es_Menor'].append('No')
        
        self.df = pd.DataFrame(datos)
        self.ancianos = ['Gilmer González', 'David Cordero', 'Raúl Cordero', 'Jhovanny Suárez',
                        'Gonzalo Sayago', 'Engelberth Oviol']
        self.siervos = ['Javier Alvarado', 'Gilmer de Jesús González', 'Rafael Torrealba']
        self.publicadores_varones = ['Willians Sivira', 'Samuel Sivira', 'Alís Fernández']
        self.hermanas = [f'Hermana {i+1}' for i in range(20)]
        self.todos_varones = self.ancianos + self.siervos + self.publicadores_varones
        
        messagebox.showinfo("Datos de ejemplo", f"Datos de ejemplo creados:\n👴 Ancianos: 6\n📖 Siervos: 3\n👨 Varones: {len(self.todos_varones)}\n👩 Hermanas: 20")
        
    def crear_widgets(self):
        # Frame principal
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(padx=15, pady=15, fill="both", expand=True)
        
        # Título
        titulo = ctk.CTkLabel(main_frame, text="VIDA Y MINISTERIO CRISTIANOS", 
                             font=("Arial", 26, "bold"))
        titulo.pack(pady=10)
        
        subtitulo = ctk.CTkLabel(main_frame, text="Herramienta Bimestral - Congregación 'El Araguaney'", 
                                font=("Arial", 14))
        subtitulo.pack(pady=5)
        
        # Frame para JW.org download
        jw_frame = ctk.CTkFrame(main_frame)
        jw_frame.pack(pady=5, fill="x", padx=10)
        
        ctk.CTkLabel(jw_frame, text="📥 Descargar desde JW.org:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        
        current_year = datetime.now().year
        self.year_combo = ctk.CTkComboBox(jw_frame, values=[str(current_year - 1), str(current_year), str(current_year + 1)], width=80)
        self.year_combo.pack(side="left", padx=5)
        self.year_combo.set(str(current_year))
        
        bimestres_nombres = [b[0] for b in JWScraper.get_bimestres()] if JW_AVAILABLE else ["Mayo-Junio"]
        self.bimestre_combo = ctk.CTkComboBox(jw_frame, values=bimestres_nombres, width=170)
        self.bimestre_combo.pack(side="left", padx=5)
        # Auto-detectar bimestre actual
        if JW_AVAILABLE:
            nombre_actual, _ = JWScraper.detectar_bimestre_actual()
            self.bimestre_combo.set(nombre_actual)
        
        self.btn_descargar = ctk.CTkButton(jw_frame, text="⬇ DESCARGAR GUÍA", 
                     command=self.descargar_desde_jw,
                     fg_color="#7B1FA2", hover_color="#4A148C",
                     font=("Arial", 12, "bold"))
        self.btn_descargar.pack(side="left", padx=10)
        
        self.status_label = ctk.CTkLabel(jw_frame, text="", font=("Arial", 11))
        self.status_label.pack(side="left", padx=10)
        
        # Frame para configuración del bimestre
        config_frame = ctk.CTkFrame(main_frame)
        config_frame.pack(pady=5, fill="x", padx=10)
        
        ctk.CTkLabel(config_frame, text="Semanas:", font=("Arial", 12)).pack(side="left", padx=5)
        self.num_semanas = ctk.CTkComboBox(config_frame, values=["8", "9", "10"], width=80)
        self.num_semanas.pack(side="left", padx=5)
        self.num_semanas.set("8")
        
        ctk.CTkButton(config_frame, text="+ CONFIGURAR MANUAL", 
                     command=self.configurar_semanas,
                     fg_color="#1565C0", hover_color="#0D47A1").pack(side="left", padx=10)
        
        ctk.CTkButton(config_frame, text="📊 GENERAR BIMESTRE", 
                     command=self.generar_bimestre_completo,
                     fg_color="#2E7D32", hover_color="#1B5E20",
                     font=("Arial", 13, "bold")).pack(side="left", padx=10)
        
        ctk.CTkButton(config_frame, text="📁 EXPORTAR EXCEL", 
                     command=self.exportar_excel,
                     fg_color="#FF6F00", hover_color="#E65100").pack(side="left", padx=5)
        
        ctk.CTkButton(config_frame, text="📄 EXPORTAR PDF", 
                     command=self.exportar_pdf,
                     fg_color="#C62828", hover_color="#B71C1C").pack(side="left", padx=5)
        
        # Barra de progreso
        self.progress_bar = ctk.CTkProgressBar(main_frame, width=400)
        self.progress_bar.pack(pady=3, padx=10)
        self.progress_bar.set(0)
        
        # Frame para las semanas (scrollable)
        self.semanas_frame = ctk.CTkScrollableFrame(main_frame, height=450)
        self.semanas_frame.pack(pady=5, fill="both", expand=True, padx=10)

    def descargar_desde_jw(self):
        """Descarga datos de JW.org en un hilo separado"""
        if not JW_AVAILABLE or not self.scraper:
            messagebox.showerror("Error", "El módulo jw_scraper no está disponible.\nVerifique que jw_scraper.py esté en la misma carpeta.")
            return
        
        self.btn_descargar.configure(state="disabled", text="Descargando...")
        self.status_label.configure(text="Conectando a JW.org...", text_color="orange")
        self.progress_bar.set(0)
        
        def _descargar():
            try:
                year = int(self.year_combo.get())
                bimestre_nombre = self.bimestre_combo.get()
                
                # Buscar el mes de inicio del bimestre seleccionado
                mes_inicio = 5  # default
                for nombre, mes in JWScraper.get_bimestres():
                    if nombre == bimestre_nombre:
                        mes_inicio = mes
                        break
                
                semanas_iso = JWScraper.calcular_semanas_bimestre(year, mes_inicio)
                
                def update_progress(current, total, info):
                    self.progress_bar.set(current / total)
                    self.status_label.configure(text=f"Descargando {current}/{total}: {info}")
                
                resultados = []
                total = len(semanas_iso)
                for i, (iso_year, iso_week) in enumerate(semanas_iso):
                    try:
                        datos = self.scraper.obtener_semana(iso_year, iso_week)
                        resultados.append(datos)
                    except Exception as e:
                        resultados.append({
                            'error': str(e), 'fecha': f'Semana {iso_week} (error)',
                            'lectura_biblica': '', 'maestros': [],
                        })
                    self.root.after(0, update_progress, i + 1, total, 
                                   resultados[-1].get('fecha', f'Semana {iso_week}'))
                
                self.datos_jw = resultados
                self.root.after(0, self._aplicar_datos_jw)
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error al descargar:\n{str(e)}"))
            finally:
                self.root.after(0, lambda: self.btn_descargar.configure(state="normal", text="⬇ DESCARGAR GUÍA"))
        
        threading.Thread(target=_descargar, daemon=True).start()
    
    def _aplicar_datos_jw(self):
        """Aplica los datos descargados a los formularios de semanas"""
        if not self.datos_jw:
            return
        
        # Actualizar número de semanas y configurar
        num = len(self.datos_jw)
        if num <= 10:
            self.num_semanas.set(str(num))
        
        # Crear formularios con datos pre-llenados
        self._crear_formularios_semanas(self.datos_jw)
        
        exitosas = sum(1 for d in self.datos_jw if 'error' not in d)
        self.status_label.configure(
            text=f"✅ {exitosas}/{num} semanas descargadas", text_color="green"
        )
        self.progress_bar.set(1)
    
    def configurar_semanas(self):
        """Crea formularios vacíos para entrada manual"""
        self._crear_formularios_semanas(None)
    
    def _crear_formularios_semanas(self, datos_precargados=None):
        """Crea los formularios para cada semana del bimestre"""
        # Limpiar frame
        for widget in self.semanas_frame.winfo_children():
            widget.destroy()
        
        self.semanas = []
        num = len(datos_precargados) if datos_precargados else int(self.num_semanas.get())
        
        for i in range(num):
            jw_data = datos_precargados[i] if datos_precargados and i < len(datos_precargados) else None
            tiene_error = jw_data and 'error' in jw_data
            
            semana_frame = ctk.CTkFrame(self.semanas_frame)
            semana_frame.pack(pady=8, fill="x", padx=10)
            
            # Encabezado de semana
            header_text = f"📅 SEMANA {i+1}"
            if jw_data and not tiene_error:
                header_text += f"  •  {jw_data.get('fecha', '')}"
                if jw_data.get('cancion_inicial'):
                    header_text += f"  |  🎵 {jw_data['cancion_inicial']}, {jw_data.get('cancion_intermedia','')}, {jw_data.get('cancion_final','')}"
            
            header_color = "red" if tiene_error else None
            ctk.CTkLabel(semana_frame, text=header_text, font=("Arial", 13, "bold"),
                        text_color=header_color).grid(row=0, column=0, columnspan=6, pady=5, sticky="w")
            
            # Fecha
            ctk.CTkLabel(semana_frame, text="Fecha:").grid(row=1, column=0, padx=5, pady=2, sticky="e")
            fecha_entry = ctk.CTkEntry(semana_frame, width=180, placeholder_text="Ej: 2-8 de octubre")
            fecha_entry.grid(row=1, column=1, padx=5, pady=2, sticky="w")
            if jw_data and not tiene_error:
                fecha_entry.insert(0, jw_data.get('fecha', ''))
            
            # Lectura
            ctk.CTkLabel(semana_frame, text="Lectura:").grid(row=1, column=2, padx=5, pady=2, sticky="e")
            lectura_entry = ctk.CTkEntry(semana_frame, width=180, placeholder_text="Ej: Isaías 58, 59")
            lectura_entry.grid(row=1, column=3, padx=5, pady=2, sticky="w")
            if jw_data and not tiene_error:
                lectura_entry.insert(0, jw_data.get('lectura_biblica', ''))
            
            # Asignaciones de Maestros
            ctk.CTkLabel(semana_frame, text="SEAMOS MEJORES MAESTROS:", 
                        font=("Arial", 11, "bold")).grid(row=2, column=0, columnspan=6, pady=(8,2), sticky="w")
            
            nombres_maestros = [
                "Empiece conversaciones", "Haga revisitas", "Haga discípulos",
                "Explique sus creencias", "Discurso"
            ]
            
            # Determinar cuáles activar basado en datos de JW.org
            tipos_activos = set()
            if jw_data and not tiene_error and jw_data.get('maestros'):
                for m in jw_data['maestros']:
                    tipos_activos.add(m['tipo'])
            
            vars_maestros = []
            for idx, asig in enumerate(nombres_maestros):
                if jw_data and not tiene_error and jw_data.get('maestros'):
                    # Activar según lo que vino de JW.org
                    val = asig in tipos_activos
                else:
                    val = True if idx < 3 else False
                var = ctk.BooleanVar(value=val)
                chk = ctk.CTkCheckBox(semana_frame, text=asig, variable=var)
                chk.grid(row=3, column=idx, padx=10, pady=2, sticky="w")
                vars_maestros.append(var)
            
            # Mostrar detalle de asignaciones si hay datos de JW.org
            if jw_data and not tiene_error and jw_data.get('maestros'):
                detalles = "  |  ".join([f"{m['tipo']} ({m['mins']}min)" for m in jw_data['maestros']])
                ctk.CTkLabel(semana_frame, text=f"→ {detalles}", 
                            font=("Arial", 10), text_color="gray").grid(
                    row=4, column=0, columnspan=6, padx=20, sticky="w")
            
            # Mostrar info de Vida Cristiana si hay datos
            if jw_data and not tiene_error and jw_data.get('vida_cristiana'):
                vc_info = "  |  ".join([f"{v['titulo']} ({v['mins']}min)" for v in jw_data['vida_cristiana']])
                ctk.CTkLabel(semana_frame, text=f"VIDA CRISTIANA: {vc_info}", 
                            font=("Arial", 10), text_color="#555").grid(
                    row=5, column=0, columnspan=6, padx=20, pady=(2,0), sticky="w")
            
            self.semanas.append({
                'numero': i+1,
                'fecha_entry': fecha_entry,
                'lectura_entry': lectura_entry,
                'maestros_vars': vars_maestros,
                'frame': semana_frame,
                'jw_data': jw_data,
            })
        
        modo = "desde JW.org" if datos_precargados else "manual"
        ctk.CTkLabel(self.semanas_frame, text=f"✅ {num} semanas configuradas ({modo}). Presione 'GENERAR BIMESTRE'", 
                    font=("Arial", 12, "bold"), text_color="green").pack(pady=10)
    
    def es_menor(self, nombre):
        if self.df is None or self.df.empty:
            return False
        match = self.df[self.df['Nombre'] == nombre]
        if not match.empty:
            return match.iloc[0].get('Es_Menor', 'No') == 'Si'
        return False
        
    def comparten_apellido(self, nombre1, nombre2):
        if not nombre1 or not nombre2 or nombre1 == "__________________" or nombre2 == "__________________":
            return False
        ap1 = nombre1.split()[-1].lower() if len(nombre1.split()) > 1 else nombre1.lower()
        ap2 = nombre2.split()[-1].lower() if len(nombre2.split()) > 1 else nombre2.lower()
        return ap1 == ap2

    def inicializar_pools(self):
        # Excluir explícitamente a Rafael Torrealba de Tesoros y Vida Cristiana
        elegibles_tesoros = [v for v in self.todos_varones if v != 'Rafael Torrealba']
        random.shuffle(elegibles_tesoros)
        self.pool_tesoros = list(elegibles_tesoros)
        
        # Vida Cristiana solo debe tener Ancianos y Siervos
        elegibles_vida = [v for v in (self.ancianos + self.siervos) if v != 'Rafael Torrealba']
        random.shuffle(elegibles_vida)
        self.pool_vida = list(elegibles_vida)

    def asignar_desde_pool(self, pool, lista_candidatos_base, asignados_semana, evitar=None):
        """Extrae del pool garantizando que todos pasen antes de repetir."""
        # Filtrar pool actual
        disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
        
        if not disponibles:
            # Si se agotaron los elegibles, rellenar con la lista base (excluyendo a Torrealba para Tesoros/Vida)
            base = [v for v in lista_candidatos_base if v != 'Rafael Torrealba']
            random.shuffle(base)
            pool.extend(base)
            disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
            
        if not disponibles:
            return "__________________"
            
        seleccionado = disponibles[0]
        pool.remove(seleccionado)
        asignados_semana.add(seleccionado)
        return seleccionado

    def asignar_persona(self, lista_candidatos, asignados_semana, evitar=None):
        """Asigna aleatorio normal (sin pool)"""
        disponibles = [p for p in lista_candidatos if p not in asignados_semana and p != evitar]
        if not disponibles:
            return "__________________"
        seleccionado = random.choice(disponibles)
        asignados_semana.add(seleccionado)
        return seleccionado

    def candidato_valido_maestros(self, persona, semana_num):
        ultima = self.historial_maestros.get(persona, -99)
        return (semana_num - ultima) >= 6

    def asignar_pareja_generica(self, asignados_semana, semana_num, es_hermanas=True, permite_familiar_opuesto=False):
        separador = " // "
        
        # Titular
        pool_titulares = self.hermanas if es_hermanas else [p for p in self.publicadores_varones]
        # Ayudante base
        pool_ayudantes_base = self.hermanas if es_hermanas else self.todos_varones
        
        # 1. Buscar titular con cooldown de 6 semanas
        titulares_disp = [t for t in pool_titulares if t not in asignados_semana and self.candidato_valido_maestros(t, semana_num)]
        if not titulares_disp:
            titulares_disp = [t for t in pool_titulares if t not in asignados_semana]
            
        if not titulares_disp: return "__________________ // __________________"
        titular = random.choice(titulares_disp)
        
        # 2. Determinar pool de ayudantes permitido
        pool_ayudantes = list(pool_ayudantes_base)
        
        if permite_familiar_opuesto:
            # 30% de probabilidad de buscar familiar de sexo opuesto (como dice el manual: mismo sexo o familiar)
            if random.random() < 0.3:
                pool_opuesto = self.todos_varones if es_hermanas else self.hermanas
                fam_opuestos = [p for p in pool_opuesto if self.comparten_apellido(titular, p)]
                if fam_opuestos:
                    pool_ayudantes.extend(fam_opuestos * 3) # Aumentar probabilidad de elegirlo
        
        # 3. Buscar ayudante con cooldown
        ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular and self.candidato_valido_maestros(a, semana_num)]
        if not ayudantes_disp:
            ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular]
            
        # 4. Aplicar regla de menores/familias
        ayudantes_validos = []
        for a in ayudantes_disp:
            if self.es_menor(a) and not self.es_menor(titular):
                if not self.comparten_apellido(titular, a):
                    continue
            ayudantes_validos.append(a)
            
        if not ayudantes_validos:
            asignados_semana.add(titular)
            self.historial_maestros[titular] = semana_num
            return f"{titular} // __________________"
            
        ayudante = random.choice(ayudantes_validos)
        
        asignados_semana.add(titular)
        asignados_semana.add(ayudante)
        self.historial_maestros[titular] = semana_num
        self.historial_maestros[ayudante] = semana_num
        
        return f"{titular}{separador}{ayudante}"

    def asignar_estudiante_solo(self, asignados_semana, semana_num, es_hermanas=True):
        # Para "Explique sus creencias" (solo 1 persona)
        pool = self.hermanas if es_hermanas else self.todos_varones
        disp = [p for p in pool if p not in asignados_semana and self.candidato_valido_maestros(p, semana_num)]
        if not disp:
            disp = [p for p in pool if p not in asignados_semana]
        if not disp: return "__________________"
        p = random.choice(disp)
        asignados_semana.add(p)
        self.historial_maestros[p] = semana_num
        return p

    def generar_semana(self, semana_info, index_semana, fecha, lectura, asignaciones_maestros):
        """Genera las asignaciones para una semana específica usando índices consecutivos para el cooldown"""
        asignados = set()
        
        # TESOROS (siempre) - Usa el sistema de Pool
        presidente = self.asignar_persona(self.ancianos, asignados)
        oracion = self.asignar_persona(self.ancianos + self.siervos, asignados)
        num1_tesoros = self.asignar_desde_pool(self.pool_tesoros, self.ancianos + self.siervos, asignados)
        num2_tesoros = self.asignar_desde_pool(self.pool_tesoros, self.ancianos + self.siervos, asignados)
        
        # LECTURA DE LA BIBLIA: Excluir a Gilmer de Jesús González y Javier Alvarado
        excluidos_lectura = ['Gilmer de Jesús González', 'Javier Alvarado']
        candidatos_lectura = [v for v in self.todos_varones if v not in excluidos_lectura]
        lectura_biblia = self.asignar_desde_pool(self.pool_tesoros, candidatos_lectura, asignados)
        
        # MAESTROS (según lo que tenga la semana)
        maestros_asignaciones = {}
        for asig in asignaciones_maestros:
            # Regla 80% hermanas para asignaciones donde aplique
            usar_hermanas = (random.random() < 0.8) and bool(self.hermanas)
            asig_lower = asig.lower()
            
            if "empiece conversaciones" in asig_lower or "escenificaci" in asig_lower:
                maestros_asignaciones[asig] = self.asignar_pareja_generica(
                    asignados, index_semana, es_hermanas=usar_hermanas, permite_familiar_opuesto=True
                )
            elif "haga revisitas" in asig_lower or "haga discípulos" in asig_lower or "haga discipulos" in asig_lower:
                maestros_asignaciones[asig] = self.asignar_pareja_generica(
                    asignados, index_semana, es_hermanas=usar_hermanas, permite_familiar_opuesto=False
                )
            else:
                # Discurso o "Explique sus creencias" (cuando no dice escenificación)
                maestros_asignaciones[asig] = self.asignar_estudiante_solo(asignados, index_semana, es_hermanas=False)
        
        # VIDA CRISTIANA - Usa el sistema de Pool
        # Según lineamientos: las partes de Vida Cristiana (excepto estudio) son para Ancianos o Siervos
        num1_vida = self.asignar_desde_pool(self.pool_vida, self.ancianos + self.siervos, asignados)
        estudio_biblico = self.asignar_persona(self.ancianos, asignados)
        
        # LECTOR DEL ESTUDIO BÍBLICO: Solo usar los hermanos autorizados
        autorizados_claves = [
            "Javier Alvarado", "Gilmer de", "Hurtado", "Patiño", 
            "Santiago Goyo", "Contreras", "Roberth", "Yonnel", "Adán", "Saavedra"
        ]
        candidatos_lector = [v for v in self.todos_varones if any(clave.lower() in v.lower() for clave in autorizados_claves)]
        
        # Fallback de seguridad si no detecta a ninguno por nombres mal escritos
        if not candidatos_lector:
            excluidos_lector = self.ancianos + ['Rafael Torrealba']
            candidatos_lector = [v for v in self.todos_varones if v not in excluidos_lector]
            
        lector = self.asignar_persona(candidatos_lector, asignados, evitar=estudio_biblico)
        
        # CONSEJERO AUXILIAR (Presidencia auxiliar): Según lineamientos, debe ser un Anciano capaz de aconsejar a otros
        presidencia_aux = self.asignar_persona(self.ancianos, asignados)
        
        return {
            'semana': semana_info['numero'],
            'fecha': fecha,
            'lectura': lectura,
            'presidente': presidente,
            'oracion': oracion,
            'num1_tesoros': num1_tesoros,
            'num2_tesoros': num2_tesoros,
            'lectura_biblia': lectura_biblia,
            'maestros': maestros_asignaciones,
            'num1_vida': num1_vida,
            'estudio_biblico': estudio_biblico,
            'lector': lector,
            'presidencia_aux': presidencia_aux,
            'total_asignados': len(asignados)
        }
    
    def generar_bimestre_completo(self):
        """Genera todas las semanas del bimestre"""
        if not self.semanas:
            messagebox.showwarning("Advertencia", "Primero configure las semanas del bimestre.")
            return
        
        self.bimestre_data = []
        
        # Reiniciar variables globales para el bimestre
        self.inicializar_pools()
        self.historial_maestros = {}
        
        for idx_semana, semana in enumerate(self.semanas):
            index = idx_semana + 1  # Para contar consecutivamente y calcular cooldowns
            
            fecha = semana['fecha_entry'].get().strip()
            if not fecha:
                fecha = f"Semana {semana['numero']}"
            
            lectura = semana['lectura_entry'].get().strip()
            if not lectura:
                lectura = "Lectura semanal"
            
            # Obtener asignaciones de maestros seleccionadas
            asignaciones_maestros = []
            nombres_maestros = ["Empiece conversaciones", "Haga revisitas", "Haga discípulos", 
                               "Explique sus creencias", "Discurso"]
            for i, var in enumerate(semana['maestros_vars']):
                if var.get():
                    asignaciones_maestros.append(nombres_maestros[i])
            
            # Generar semana (pasamos el index correlativo en lugar de semana['numero'])
            datos_semana = self.generar_semana(
                semana, index, fecha, lectura, asignaciones_maestros
            )
            self.bimestre_data.append(datos_semana)
        
        # Mostrar en pantalla
        self.mostrar_en_pantalla()
        
        messagebox.showinfo("Éxito", f"✅ Bimestre generado correctamente.\n📊 Total semanas: {len(self.bimestre_data)}")
    
    def mostrar_en_pantalla(self):
        """Muestra el bimestre en una ventana con diseño de 3 columnas estilo Excel"""
        if not self.bimestre_data:
            return
        
        ventana = ctk.CTkToplevel(self.root)
        ventana.title("Vista previa del bimestre (Formato Excel)")
        ventana.geometry("1100x700")
        
        scroll_frame = ctk.CTkScrollableFrame(ventana, width=1050, height=650)
        scroll_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        for semana in self.bimestre_data:
            week_frame = ctk.CTkFrame(scroll_frame, corner_radius=0, border_width=1, border_color="black")
            week_frame.pack(padx=10, pady=10, fill="x")
            
            # Row 1: Titulo y Fecha
            header_frame = ctk.CTkFrame(week_frame, fg_color="transparent")
            header_frame.pack(fill="x", padx=5, pady=5)
            
            ctk.CTkLabel(header_frame, text=f"VIDA Y MINISTERIO CRISTIANOS  |  {semana['lectura'].upper()}", 
                         font=("Arial", 16, "bold")).pack(side="left", padx=5)
            ctk.CTkLabel(header_frame, text=semana['fecha'], font=("Arial", 12, "bold")).pack(side="right", padx=5)
            
            # Row 2: Franja Morada
            ctk.CTkFrame(week_frame, fg_color="#5E005E", height=10).pack(fill="x")
            
            # Row 3: Presidente y Oración
            sub_frame = ctk.CTkFrame(week_frame, fg_color="transparent")
            sub_frame.pack(fill="x", padx=5, pady=5)
            ctk.CTkLabel(sub_frame, text=f"PRESIDENTE DE LA REUNIÓN: {semana['presidente']}", font=("Arial", 12)).pack(side="left", padx=5)
            ctk.CTkLabel(sub_frame, text=f"ORACIÓN: {semana['oracion']}", font=("Arial", 12)).pack(side="right", padx=5)
            
            # Row 4: Columnas
            cols_frame = ctk.CTkFrame(week_frame, fg_color="transparent")
            cols_frame.pack(fill="x")
            
            cols_frame.grid_columnconfigure(0, weight=1, uniform="col")
            cols_frame.grid_columnconfigure(1, weight=1, uniform="col")
            cols_frame.grid_columnconfigure(2, weight=1, uniform="col")
            
            # Col 1: Tesoros
            col1 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col1.grid(row=0, column=0, sticky="nsew", padx=1)
            ctk.CTkLabel(col1, text="TESOROS DE LA BIBLIA", fg_color="#808080", text_color="white", font=("Arial", 12, "bold"), corner_radius=0).pack(fill="x")
            ctk.CTkLabel(col1, text=f"NUM 1: {semana['num1_tesoros']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            ctk.CTkLabel(col1, text=f"NUM 2: {semana['num2_tesoros']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            ctk.CTkLabel(col1, text=f"NUM 3: {semana['lectura_biblia']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            
            # Col 2: Maestros
            col2 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col2.grid(row=0, column=1, sticky="nsew", padx=1)
            ctk.CTkLabel(col2, text="SEAMOS MEJORES MAESTROS", fg_color="#FF8C00", text_color="white", font=("Arial", 12, "bold"), corner_radius=0).pack(fill="x")
            num_maestros = 4
            for key, val in semana['maestros'].items():
                ctk.CTkLabel(col2, text=f"NUM {num_maestros}: {val}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
                num_maestros += 1
                
            # Col 3: Vida Cristiana
            col3 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col3.grid(row=0, column=2, sticky="nsew", padx=1)
            ctk.CTkLabel(col3, text="NUESTRA VIDA CRISTIANA", fg_color="#B22222", text_color="white", font=("Arial", 12, "bold"), corner_radius=0).pack(fill="x")
            ctk.CTkLabel(col3, text=f"NUM 7: {semana['num1_vida']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            ctk.CTkLabel(col3, text=f"ESTUDIO: {semana['estudio_biblico']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            ctk.CTkLabel(col3, text=f"LECTOR: {semana['lector']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            ctk.CTkLabel(col3, text=f"Cons Aux: {semana['presidencia_aux']}", font=("Arial", 11)).pack(anchor="w", padx=5, pady=2)
            
            # Fila de servicios
            servicios_frame = ctk.CTkFrame(week_frame, fg_color="#EAEAEA", corner_radius=0)
            servicios_frame.pack(fill="x", pady=(5,0))
            ctk.CTkLabel(servicios_frame, text="Sonido: \t\t\t Plataforma: \t\t\t Acomodadores:", font=("Arial", 11)).pack(anchor="w", padx=10, pady=2)
            ctk.CTkLabel(servicios_frame, text="\t\t\t Micrófonos:", font=("Arial", 11)).pack(anchor="w", padx=10, pady=2)
    
    def exportar_excel(self):
        """Exporta el bimestre a Excel en una sola hoja con formato avanzado"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos. Genere el bimestre primero.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"bimestre_vmc_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not filename:
            return
            
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bimestre"
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 48
        ws.column_dimensions['C'].width = 40
        
        # Definir estilos
        font_title = Font(name='Arial', size=18, bold=True)
        font_date = Font(name='Arial', size=12, bold=True)
        font_header_blanco = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_normal = Font(name='Arial', size=10)
        
        fill_purple = PatternFill(start_color='5E005E', end_color='5E005E', fill_type='solid') # Línea divisoria
        fill_tesoros = PatternFill(start_color='808080', end_color='808080', fill_type='solid') # Gris
        fill_maestros = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid') # Naranja
        fill_vida = PatternFill(start_color='B22222', end_color='B22222', fill_type='solid') # Rojo oscuro
        fill_light_grey = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
        
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
                             
        current_row = 1
        
        for semana in self.bimestre_data:
            # Fila 1: Título y Fecha
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell_title = ws.cell(row=current_row, column=1, value=f"VIDA Y MINISTERIO CRISTIANOS  |  {semana['lectura'].upper()}")
            cell_title.font = font_title
            cell_title.alignment = Alignment(horizontal='center', vertical='center')
            
            cell_date = ws.cell(row=current_row, column=3, value=semana['fecha'])
            cell_date.font = font_date
            cell_date.alignment = Alignment(horizontal='left', vertical='center')
            
            current_row += 1
            
            # Fila 2: Franja morada
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_purple
            ws.row_dimensions[current_row].height = 15
            current_row += 1
            
            # Fila 3: Presidente y Oración
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell_pres = ws.cell(row=current_row, column=1, value=f"PRESIDENTE DE LA REUNIÓN: {semana['presidente']}")
            cell_pres.font = font_normal
            cell_pres.border = border_thin
            
            cell_ora = ws.cell(row=current_row, column=3, value=f"ORACIÓN: {semana['oracion']}")
            cell_ora.font = font_normal
            cell_ora.border = border_thin
            
            current_row += 1
            
            # Fila 4: Encabezados de secciones
            headers = [
                ("TESOROS DE LA BIBLIA", fill_tesoros),
                ("SEAMOS MEJORES MAESTROS", fill_maestros),
                ("NUESTRA VIDA CRISTIANA", fill_vida)
            ]
            for col, (text, fill) in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=text)
                cell.font = font_header_blanco
                cell.fill = fill
                cell.border = border_thin
            
            current_row += 1
            
            # Preparar datos de las columnas
            col_tesoros = [
                f"NUM 1: {semana['num1_tesoros']}",
                f"NUM 2: {semana['num2_tesoros']}",
                f"NUM 3: {semana['lectura_biblia']}"
            ]
            
            col_maestros = []
            num_maestros = 4
            for key, val in semana['maestros'].items():
                col_maestros.append(f"NUM {num_maestros}: {val}")
                num_maestros += 1
                
            col_vida = [
                f"NUM 7: {semana['num1_vida']}",
                f"NUM 8: ",
                f"ESTUDIO BÍBLICO: {semana['estudio_biblico']}",
                f"LECTURA ESTUDIO BÍBLICO: {semana['lector']}"
            ]
            
            max_rows = max(len(col_tesoros), len(col_maestros), len(col_vida))
            
            # Filas de asignaciones
            for i in range(max_rows):
                # Tesoros
                val_t = col_tesoros[i] if i < len(col_tesoros) else ""
                cell_t = ws.cell(row=current_row, column=1, value=val_t)
                cell_t.font = font_normal
                cell_t.border = border_thin
                
                # Maestros
                val_m = col_maestros[i] if i < len(col_maestros) else ""
                cell_m = ws.cell(row=current_row, column=2, value=val_m)
                cell_m.font = font_normal
                cell_m.border = border_thin
                
                # Vida
                val_v = col_vida[i] if i < len(col_vida) else ""
                cell_v = ws.cell(row=current_row, column=3, value=val_v)
                cell_v.font = font_normal
                cell_v.border = border_thin
                
                current_row += 1
                
            # Fila de servicios (Sonido, Plataforma, etc.)
            servicios = [
                ("Sonido:", "Plataforma:", "Acomodadores:"),
                ("", "Microfonos:", "")
            ]
            for serv_row in servicios:
                for col, text in enumerate(serv_row, start=1):
                    cell = ws.cell(row=current_row, column=col, value=text)
                    cell.font = font_normal
                    cell.fill = fill_light_grey
                    cell.border = border_thin
                current_row += 1
                
            # Espacio extra entre semanas
            current_row += 1
            
        try:
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Excel exportado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar Excel:\n{str(e)}")
    
    def exportar_pdf(self):
        """Exporta el bimestre a PDF con formato idéntico al de Excel (3 columnas)"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos. Genere el bimestre primero.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"bimestre_vmc_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        
        if not filename:
            return
            
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(filename, pagesize=landscape(letter),
                                rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        for semana in self.bimestre_data:
            data = []
            
            # Row 1: Título y Fecha/Lectura
            data.append([f"VIDA Y MINISTERIO CRISTIANOS", "", f"{semana['fecha']}  |  {semana['lectura']}"])
            
            # Row 2: Franja morada
            data.append(["", "", ""])
            
            # Row 3: Presidente y Oración
            data.append([f"PRESIDENTE DE LA REUNIÓN: {semana['presidente']}", "", f"ORACIÓN: {semana['oracion']}"])
            
            # Row 4: Encabezados
            data.append(["TESOROS DE LA BIBLIA", "SEAMOS MEJORES MAESTROS", "NUESTRA VIDA CRISTIANA"])
            
            # Preparar columnas
            col_tesoros = [
                f"NUM 1: {semana['num1_tesoros']}",
                f"NUM 2: {semana['num2_tesoros']}",
                f"NUM 3: {semana['lectura_biblia']}"
            ]
            
            col_maestros = []
            num_maestros = 4
            for key, val in semana['maestros'].items():
                col_maestros.append(f"NUM {num_maestros}: {val}")
                num_maestros += 1
                
            col_vida = [
                f"NUM 7: {semana['num1_vida']}",
                f"NUM 8: ",
                f"ESTUDIO BÍBLICO: {semana['estudio_biblico']}",
                f"LECTURA ESTUDIO BÍBLICO: {semana['lector']}"
            ]
            
            max_rows = max(len(col_tesoros), len(col_maestros), len(col_vida))
            
            for i in range(max_rows):
                val_t = col_tesoros[i] if i < len(col_tesoros) else ""
                val_m = col_maestros[i] if i < len(col_maestros) else ""
                val_v = col_vida[i] if i < len(col_vida) else ""
                data.append([val_t, val_m, val_v])
                
            # Servicios
            data.append(["Sonido:", "Plataforma:", "Acomodadores:"])
            data.append(["", "Microfonos:", ""])
            
            col_widths = [3.5 * inch, 3.5 * inch, 3.5 * inch]
            table = Table(data, colWidths=col_widths)
            
            style = TableStyle([
                # Row 1 (Title)
                ('SPAN', (0, 0), (1, 0)),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (1, 0), 16),
                ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (2, 0), (2, 0), 10),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Row 2 (Purple separator)
                ('SPAN', (0, 1), (2, 1)),
                ('BACKGROUND', (0, 1), (2, 1), colors.HexColor('#5E005E')),
                
                # Row 3 (President / Prayer)
                ('SPAN', (0, 2), (1, 2)),
                ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
                ('FONTSIZE', (0, 2), (-1, 2), 10),
                ('BOTTOMPADDING', (0, 2), (-1, 2), 6),
                ('TOPPADDING', (0, 2), (-1, 2), 6),
                
                # Row 4 (Headers)
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 3), (-1, 3), 10),
                ('TEXTCOLOR', (0, 3), (-1, 3), colors.whitesmoke),
                ('BACKGROUND', (0, 3), (0, 3), colors.HexColor('#808080')), # Grey
                ('BACKGROUND', (1, 3), (1, 3), colors.HexColor('#FF8C00')), # Orange
                ('BACKGROUND', (2, 3), (2, 3), colors.HexColor('#B22222')), # Red
                ('ALIGN', (0, 3), (-1, 3), 'LEFT'),
                
                # Content rows
                ('FONTNAME', (0, 4), (-1, -3), 'Helvetica'),
                ('FONTSIZE', (0, 4), (-1, -3), 9),
                ('VALIGN', (0, 4), (-1, -3), 'TOP'),
                ('TOPPADDING', (0, 4), (-1, -3), 4),
                ('BOTTOMPADDING', (0, 4), (-1, -3), 4),
                
                # Service rows (last 2 rows)
                ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#EAEAEA')),
                ('FONTNAME', (0, -2), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, -2), (-1, -1), 9),
                
                # Global Grid
                ('GRID', (0, 2), (-1, -1), 0.5, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ])
            
            table.setStyle(style)
            story.append(table)
            story.append(Spacer(1, 0.4 * inch))
            
        try:
            doc.build(story)
            messagebox.showinfo("Éxito", f"PDF exportado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{e}")

def main():
    root = ctk.CTk()
    app = CoordinacionVMC(root)
    root.mainloop()

if __name__ == "__main__":
    main()