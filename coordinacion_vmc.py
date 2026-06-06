import customtkinter as ctk
import pandas as pd
import random
from tkinter import messagebox, filedialog, ttk
import os
from datetime import datetime
import threading
import json
import copy
import re
try:
    from jw_scraper import JWScraper
    JW_AVAILABLE = True
except ImportError:
    JW_AVAILABLE = False
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

# Configuración
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# Colores Premium
COLOR_TREASURES = "#546E7A" # Gris
COLOR_MINISTRY = "#F57C00"  # Naranja
COLOR_LIFE = "#D32F2F"      # Rojo oscuro
COLOR_PURPLE = "#512DA8"    # Morado Teocrático
COLOR_SUCCESS = "#00C853"   # Verde
COLOR_ACCENT = "#2962FF"    # Azul

class CoordinacionVMC:
    def __init__(self, root):
        self.root = root
        self.root.title("Coordinación VMC - El Araguaney (Bimestral)")
        self.root.geometry("1300x800")
        
        # Datos de la congregación
        self.df = None
        self.ancianos = []
        self.siervos = []
        self.publicadores_varones = []
        self.hermanas = []
        self.todos_varones = []
        
        # Datos del bimestre
        self.semanas = []
        self.bimestre_data = []
        
        # Historial y reglas globales
        self.historial_maestros = {}
        self.pool_tesoros = []
        self.pool_vida = []
        
        import sys
        # Detectar directorio base (para script o ejecutable)
        if getattr(sys, 'frozen', False):
            self.base_dir = os.path.dirname(sys.executable)
        else:
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
            
        # Rutas de archivos de datos
        self.archivo_excel = os.path.join(self.base_dir, "Congregacion_Araguaney.xlsx")
        self.historial_archivo = os.path.join(self.base_dir, "vmc_historial.json")
        self.cache_jw_archivo = os.path.join(self.base_dir, "jw_cache.json")
        
        # Datos de JW.org
        self.datos_jw = []
        if JW_AVAILABLE:
            self.scraper = JWScraper()
            # Forzar ruta de cache en el scraper si fuera necesario (se manejará en el init del scraper también)
        else:
            self.scraper = None
            
        # Cargar datos
        self.cargar_datos()
        self.cargar_historial()
        
        # Crear interfaz
        self.crear_widgets()
        
    def cargar_datos(self):
        """Carga el archivo Excel de la congregación"""
        try:
            archivo = self.archivo_excel
            
            if not os.path.exists(archivo):
                self.crear_datos_ejemplo()
                return
            
            self.df = pd.read_excel(archivo)
            
            # 1. Exclusiones: Eliminar familia Spolzino y Saucedo
            self.df = self.df[~self.df['Nombre'].str.contains('Spolzino|Saucedo', case=False, na=False)]
            
            # Asegurar nuevas columnas si no existen (Granularidad avanzada)
            nuevas_cols = {
                'Edad': 30,
                'Hab_Lectura': 'Si',
                'Hab_Puede_Presidir': 'No',
                'Hab_Tes_Discurso': 'No',
                'Hab_Tes_Perlas': 'No',
                'Hab_Mae_Conversacion_Enc': 'Si',
                'Hab_Mae_Conversacion_Ayu': 'Si',
                'Hab_Mae_Revisita_Enc': 'Si',
                'Hab_Mae_Revisita_Ayu': 'Si',
                'Hab_Mae_Discipulos_Enc': 'Si',
                'Hab_Mae_Discipulos_Ayu': 'Si',
                'Hab_Mae_Creencias_Esc_Enc': 'Si',
                'Hab_Mae_Creencias_Esc_Ayu': 'Si',
                'Hab_Mae_Creencias_Dis': 'Si',
                'Hab_Mae_Discurso': 'Si',
                'Hab_Vida_Parte1': 'No',
                'Hab_Vida_Parte2': 'No',
                'Hab_Vida_Locales': 'No',
                'Hab_Oracion': 'Si',
                'Hab_Estudio_Conductor': 'No',
                'Hab_Estudio_Lector': 'No',
                'Hab_Sonido': 'No',
                'Hab_Mics': 'Si',
                'Hab_Plataforma': 'Si',
                'Hab_Acomodador': 'Si',
                'Telefono': ''
            }
            
            # Migración de datos antiguos si existen
            if 'Hab_Maestros' in self.df.columns:
                # Si ya existía la habilidad genérica, propagar a las específicas
                for c in [c for c in nuevas_cols if c.startswith('Hab_Mae_')]:
                    if c not in self.df.columns:
                        self.df[c] = self.df['Hab_Maestros']
            
            if 'Hab_Vida_Encargado' in self.df.columns:
                for c in ['Hab_Vida_Parte1', 'Hab_Vida_Parte2', 'Hab_Vida_Locales']:
                    if c not in self.df.columns:
                        self.df[c] = self.df['Hab_Vida_Encargado']

            if 'Hab_Lector' in self.df.columns:
                self.df['Hab_Estudio_Lector'] = self.df['Hab_Lector']
            
            if 'Hab_Servicios' in self.df.columns:
                for c in ['Hab_Sonido', 'Hab_Mics', 'Hab_Plataforma', 'Hab_Acomodador']:
                    if c not in self.df.columns:
                        self.df[c] = self.df['Hab_Servicios']

            if 'Hab_Puede_Presidir' not in self.df.columns:
                self.df['Hab_Puede_Presidir'] = self.df.apply(
                    lambda r: 'Si' if r.get('Privilegio') == 'Anciano' and r.get('Genero') == 'M' else 'No',
                    axis=1
                )

            for col, default in nuevas_cols.items():
                if col not in self.df.columns:
                    self.df[col] = default
            
            self.df['Es_Menor'] = self.df['Es_Menor'].astype(str).str.strip().str.title()
            
            # Limpieza y formateo de teléfonos
            def clean_phone(t):
                if pd.isna(t) or str(t).strip().lower() in ['nan', '']: return ""
                t_str = str(t).split('.')[0].strip() # Quitar .0 si viene de Excel como float
                digits = "".join(filter(str.isdigit, t_str))
                if not digits: return ""
                if not digits.startswith('58'): digits = "58" + digits
                return "+" + digits
            
            self.df['Telefono'] = self.df['Telefono'].apply(clean_phone)
            
            # Recargar listas filtradas por Habilidades
            self.ancianos = self.df[(self.df['Privilegio'] == 'Anciano') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.siervos = self.df[(self.df['Privilegio'] == 'Siervo Min.') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.publicadores_varones = self.df[(self.df['Privilegio'] == 'Publicador') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
            self.hermanas = self.df[self.df['Genero'] == 'F']['Nombre'].tolist()
            self.todos_varones = self.ancianos + self.siervos + self.publicadores_varones
            
            messagebox.showinfo("Éxito", f"Datos cargados:\n👴 Ancianos: {len(self.ancianos)}\n📖 Siervos: {len(self.siervos)}\n👨 Varones: {len(self.publicadores_varones)}\n👩 Hermanas: {len(self.hermanas)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")

    def cargar_historial(self):
        """Carga el historial persistente desde un archivo JSON"""
        if os.path.exists(self.historial_archivo):
            try:
                with open(self.historial_archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.historial_maestros = data.get('maestros', {})
                    # Convertir llaves de fecha de string a datetime si fuera necesario, 
                    # pero aquí usamos índices de semana.
            except Exception:
                self.historial_maestros = {}
        else:
            self.historial_maestros = {}

    def guardar_historial(self):
        """Guarda el historial actual en un archivo JSON"""
        try:
            with open(self.historial_archivo, 'w', encoding='utf-8') as f:
                json.dump({
                    'maestros': self.historial_maestros,
                    'ultima_actualizacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Error al guardar historial: {e}")
            
    def crear_datos_ejemplo(self):
        """Crea datos de ejemplo"""
        datos = {
            'Nombre': ['Gilmer González', 'David Cordero', 'Raúl Cordero', 'Jhovanny Suárez',
                       'Gonzalo Sayago', 'Engelberth Oviol',
                       'Javier Alvarado', 'Gilmer de Jesús González', 'Rafael Torrealba',
                       'Willians Sivira', 'Samuel Sivira', 'Alís Fernández'],
            'Privilegio': ['Anciano','Anciano','Anciano','Anciano','Anciano','Anciano',
                          'Siervo Min.','Siervo Min.','Siervo Min.',
                          'Publicador','Publicador','Publicador'],
            'Genero': ['M','M','M','M','M','M','M','M','M','M','M','M'],
            'Es_Menor': ['No','No','No','No','No','No','No','No','No','No','Si','Si'],
            'Edad': [45, 50, 48, 40, 55, 38, 35, 32, 28, 42, 15, 12],
            'Hab_Lectura': ['Si']*12,
            'Hab_Puede_Presidir': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Tes_Discurso': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Tes_Perlas': ['Si']*12,
            'Hab_Mae_Conversacion_Enc': ['Si']*12,
            'Hab_Mae_Conversacion_Ayu': ['Si']*12,
            'Hab_Mae_Revisita_Enc': ['Si']*12,
            'Hab_Mae_Revisita_Ayu': ['Si']*12,
            'Hab_Mae_Discipulos_Enc': ['Si']*12,
            'Hab_Mae_Discipulos_Ayu': ['Si']*12,
            'Hab_Mae_Creencias_Esc_Enc': ['Si']*12,
            'Hab_Mae_Creencias_Esc_Ayu': ['Si']*12,
            'Hab_Mae_Creencias_Dis': ['Si']*12,
            'Hab_Mae_Discurso': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Vida_Parte1': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Vida_Parte2': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Vida_Locales': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Estudio_Conductor': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Estudio_Lector': ['Si', 'Si', 'Si', 'Si', 'Si', 'Si', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Hab_Oracion': ['Si']*12,
            'Hab_Sonido': ['Si', 'Si', 'No', 'Si', 'No', 'No', 'No', 'Si', 'No', 'No', 'No', 'No'],
            'Hab_Mics': ['Si']*12,
            'Hab_Plataforma': ['Si']*12,
            'Hab_Acomodador': ['Si']*12,
            'Telefono': ['+584121234567']*12
        }
        # Agregar hermanas
        for i in range(20):
            nombre = f'Hermana {i+1}'
            datos['Nombre'].append(nombre)
            datos['Privilegio'].append('Publicador')
            datos['Genero'].append('F')
            datos['Es_Menor'].append('No')
            
            # Completar el resto de columnas con valores por defecto para que todos tengan el mismo largo
            for key in datos.keys():
                if len(datos[key]) < len(datos['Nombre']):
                    if key == 'Edad': datos[key].append(30)
                    elif key == 'Telefono': datos[key].append('')
                    elif key.startswith('Hab_Mae_'): datos[key].append('Si') # Por defecto las hermanas son estudiantes
                    else: datos[key].append('No')
        
        self.df = pd.DataFrame(datos)
        self.ancianos = ['Gilmer González', 'David Cordero', 'Raúl Cordero', 'Jhovanny Suárez',
                        'Gonzalo Sayago', 'Engelberth Oviol']
        self.siervos = ['Javier Alvarado', 'Gilmer de Jesús González', 'Rafael Torrealba']
        self.publicadores_varones = ['Willians Sivira', 'Samuel Sivira', 'Alís Fernández']
        self.hermanas = [f'Hermana {i+1}' for i in range(20)]
        self.todos_varones = self.ancianos + self.siervos + self.publicadores_varones
        
        messagebox.showinfo("Datos de ejemplo", f"Datos de ejemplo creados:\n👴 Ancianos: 6\n📖 Siervos: 3\n👨 Varones: {len(self.todos_varones)}\n👩 Hermanas: 20")
        
    def crear_widgets(self):
        # Frame principal
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
        
        # TabView Principal
        self.tabview = ctk.CTkTabview(self.root, fg_color="transparent", segmented_button_selected_color=COLOR_PURPLE)
        self.tabview.pack(padx=10, pady=10, fill="both", expand=True)
        
        self.tab_prog = self.tabview.add("📅 PROGRAMACIÓN")
        self.tab_cong = self.tabview.add("👥 CONGREGACIÓN")
        self.tab_stats = self.tabview.add("📊 ESTADÍSTICAS")
        self.tab_config = self.tabview.add("⚙️ AJUSTES")
        
        self._setup_tab_programacion()
        self._setup_tab_congregacion()
        self._setup_tab_estadisticas()
        self._setup_tab_configuracion()

    def _setup_tab_programacion(self):
        main_frame = self.tab_prog
        
        # Título con estilo premium
        header_frame = ctk.CTkFrame(main_frame, fg_color=COLOR_PURPLE, height=80, corner_radius=10)
        header_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header_frame, text="✨ COORDINACIÓN VIDA Y MINISTERIO", 
                             font=("Segoe UI", 24, "bold"), text_color="white").pack(pady=(10, 0))
        ctk.CTkLabel(header_frame, text="Desarrollado por Gilmer Gonzalez", 
                             font=("Segoe UI", 12, "italic"), text_color="#E0E0E0").pack(pady=(0, 10))
        
        # Dashboard superior
        dash_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        dash_frame.pack(fill="x", padx=10, pady=5)
        
        # Panel Izquierdo: JW.org
        jw_panel = ctk.CTkFrame(dash_frame, border_width=1, border_color="#E0E0E0")
        jw_panel.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        
        ctk.CTkLabel(jw_panel, text="📥 SINCRONIZACIÓN JW.ORG", font=("Arial", 12, "bold"), text_color=COLOR_PURPLE).pack(pady=5)
        
        controls = ctk.CTkFrame(jw_panel, fg_color="transparent")
        controls.pack(pady=5)
        
        current_year = datetime.now().year
        self.year_combo = ctk.CTkComboBox(controls, values=[str(current_year - 1), str(current_year), str(current_year + 1)], width=90)
        self.year_combo.pack(side="left", padx=5)
        self.year_combo.set(str(current_year))
        
        bimestres_nombres = [b[0] for b in JWScraper.get_bimestres()] if JW_AVAILABLE else ["Mayo-Junio"]
        self.bimestre_combo = ctk.CTkComboBox(controls, values=bimestres_nombres, width=160)
        self.bimestre_combo.pack(side="left", padx=5)
        if JW_AVAILABLE:
            nombre_actual, _ = JWScraper.detectar_bimestre_actual()
            self.bimestre_combo.set(nombre_actual)
            
        self.btn_descargar = ctk.CTkButton(jw_panel, text="Sincronizar Guía", 
                     command=self.descargar_desde_jw,
                     fg_color=COLOR_PURPLE, hover_color="#4A148C",
                     font=("Arial", 13, "bold"), height=35)
        self.btn_descargar.pack(pady=10, padx=20, fill="x")
        
        self.status_label = ctk.CTkLabel(jw_panel, text="Listo para descargar", font=("Arial", 11), text_color="gray")
        self.status_label.pack(pady=2)

        # Panel Derecho: Acciones
        action_panel = ctk.CTkFrame(dash_frame, border_width=1, border_color="#E0E0E0")
        action_panel.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        
        ctk.CTkLabel(action_panel, text="⚙️ ACCIONES Y EXPORTACIÓN", font=("Arial", 12, "bold"), text_color=COLOR_ACCENT).pack(pady=5)
        
        config_manual = ctk.CTkFrame(action_panel, fg_color="transparent")
        config_manual.pack(pady=2)
        ctk.CTkLabel(config_manual, text="Semanas:", font=("Arial", 11)).pack(side="left", padx=5)
        self.num_semanas = ctk.CTkComboBox(config_manual, values=["8", "9", "10"], width=70, height=25)
        self.num_semanas.pack(side="left", padx=5)
        self.num_semanas.set("8")
        ctk.CTkButton(config_manual, text="+ Config Manual", command=self.configurar_semanas, 
                     fg_color="transparent", text_color=COLOR_ACCENT, border_width=1, border_color=COLOR_ACCENT,
                     width=100, height=25).pack(side="left", padx=5)

        btn_grid = ctk.CTkFrame(action_panel, fg_color="transparent")
        btn_grid.pack(expand=True)
        
        self.btn_generar = ctk.CTkButton(btn_grid, text="📊 GENERAR BIMESTRE", 
                                       command=self.generar_bimestre_completo,
                                       fg_color=COLOR_SUCCESS, hover_color="#1B5E20",
                                       font=("Arial", 13, "bold"), width=180, height=40)
        self.btn_generar.grid(row=0, column=0, padx=5, pady=5)
        
        ctk.CTkButton(btn_grid, text="📁 EXCEL", command=self.exportar_excel,
                     fg_color="#FF6F00", hover_color="#E65100", width=85).grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkButton(btn_grid, text="📄 PDF", command=self.exportar_pdf,
                     fg_color="#C62828", hover_color="#B71C1C", width=85).grid(row=0, column=2, padx=5, pady=5)
        
        ctk.CTkButton(btn_grid, text="📄 BOLSILLO", command=self.exportar_pdf_bolsillo,
                     fg_color="#455A64", hover_color="#263238", width=180).grid(row=1, column=0, padx=5, pady=5)
        
        ctk.CTkButton(btn_grid, text="📄 VALES S-89", command=self.exportar_s89_pdf,
                     fg_color="#388E3C", hover_color="#2E7D32", width=180).grid(row=1, column=1, columnspan=2, padx=5, pady=5)
                     
        ctk.CTkButton(btn_grid, text="📂 CARGAR GUARDADO", command=self.cargar_estado_bimestre,
                     fg_color="#00695C", hover_color="#004D40", width=180).grid(row=2, column=0, columnspan=3, padx=5, pady=5)

        # Barra de progreso moderna
        progress_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        progress_frame.pack(fill="x", padx=15, pady=5)
        self.progress_bar = ctk.CTkProgressBar(progress_frame, height=12)
        self.progress_bar.pack(fill="x", pady=2)
        self.progress_bar.set(0)
        
        # Frame para las semanas (scrollable)
        self.semanas_frame = ctk.CTkScrollableFrame(main_frame, fg_color="transparent", height=450)
        self.semanas_frame.pack(pady=10, fill="both", expand=True, padx=15)

    def _setup_tab_congregacion(self):
        """Pestaña para gestionar la lista de hermanos"""
        parent = self.tab_cong
        
        controls = ctk.CTkFrame(parent, fg_color="transparent")
        controls.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(controls, text="📋 LISTA DE PUBLICADORES", font=("Arial", 18, "bold")).pack(side="left", padx=5)
        
        ctk.CTkButton(controls, text="+ AÑADIR HERMANO", command=self.modal_añadir_hermano,
                     fg_color=COLOR_SUCCESS, width=150).pack(side="right", padx=5)
        ctk.CTkButton(controls, text="💾 GUARDAR CAMBIOS", command=self.guardar_excel_congregacion,
                     fg_color=COLOR_ACCENT, width=150).pack(side="right", padx=5)

        # Tabla de hermanos (Simulada con scrollable frame)
        self.table_frame = ctk.CTkScrollableFrame(parent, fg_color="transparent", border_width=1, border_color="#DDD")
        self.table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.actualizar_tabla_hermanos()

    def actualizar_tabla_hermanos(self):
        for widget in self.table_frame.winfo_children():
            widget.destroy()
            
        if self.df is None: return
        
        # Definir anchos de columna fijos para asegurar alineación
        anchos = [240, 120, 70, 170, 100, 40]
        headers = ["Nombre", "Privilegio", "Género", "Teléfono", "Habilidades", "Acc."]
        
        # Encabezados
        h_frame = ctk.CTkFrame(self.table_frame, fg_color="transparent", corner_radius=0)
        h_frame.grid(row=0, column=0, columnspan=len(headers), sticky="ew", pady=(0, 5))
        
        for i, h in enumerate(headers):
            ctk.CTkLabel(h_frame, text=h, font=("Arial", 11, "bold"), width=anchos[i]).grid(row=0, column=i, padx=5, pady=5)

        # Contenido
        for idx, row in self.df.iterrows():
            # Nombre
            e1 = ctk.CTkEntry(self.table_frame, width=anchos[0], height=28)
            e1.insert(0, row['Nombre'])
            e1.grid(row=idx+1, column=0, padx=5, pady=2)
            e1.bind("<KeyRelease>", lambda e, i=idx, w=e1: self._update_df_val(i, 'Nombre', w.get()))
            
            # Privilegio
            e2 = ctk.CTkComboBox(self.table_frame, values=["Anciano", "Siervo Min.", "Publicador"], 
                                width=anchos[1], height=28, 
                                command=lambda val, i=idx: self._update_df_val(i, 'Privilegio', val))
            e2.set(row['Privilegio'])
            e2.grid(row=idx+1, column=1, padx=5, pady=2)
            
            # Género
            e3 = ctk.CTkComboBox(self.table_frame, values=["M", "F"], width=anchos[2], height=28,
                                command=lambda val, i=idx: self._update_df_val(i, 'Genero', val))
            e3.set(row['Genero'])
            e3.grid(row=idx+1, column=2, padx=5, pady=2)
            
            # Teléfono
            e_tel = ctk.CTkEntry(self.table_frame, width=anchos[3], height=28)
            e_tel.insert(0, str(row.get('Telefono', '')))
            e_tel.grid(row=idx+1, column=3, padx=5, pady=2)
            e_tel.bind("<KeyRelease>", lambda e, i=idx, w=e_tel: self._update_df_val(i, 'Telefono', w.get()))
            
            # Botón Habilidades
            ctk.CTkButton(self.table_frame, text="⚙️ ASIG.", width=anchos[4], height=28, fg_color=COLOR_ACCENT,
                         command=lambda i=idx: self.modal_habilidades_hermano(i)).grid(row=idx+1, column=4, padx=5, pady=2)
            
            # Botón eliminar
            ctk.CTkButton(self.table_frame, text="🗑️", width=anchos[5], height=28, fg_color="#C62828", hover_color="#B71C1C",
                         command=lambda i=idx: self.eliminar_hermano(i)).grid(row=idx+1, column=5, padx=5, pady=2)

    def _update_df_val(self, index, column, value):
        """Actualiza un valor específico en el DataFrame (sin guardar a disco inmediatamente)"""
        self.df.at[index, column] = value

    def modal_añadir_hermano(self):
        modal = ctk.CTkToplevel(self.root)
        modal.title("Añadir Publicador")
        modal.geometry("500x750")
        modal.after(100, lambda: modal.focus_force())
        modal.grab_set() 
        
        ctk.CTkLabel(modal, text="Añadir Nuevo Publicador", font=("Arial", 18, "bold")).pack(pady=10)
        
        main_f = ctk.CTkFrame(modal, fg_color="transparent")
        main_f.pack(pady=5, padx=20, fill="both", expand=True, side="top")
        
        top_section = ctk.CTkFrame(main_f, fg_color="transparent")
        top_section.pack(fill="x", pady=5)
        
        inputs_frame = ctk.CTkFrame(top_section, fg_color="transparent")
        inputs_frame.pack(side="left", fill="both", expand=True)
        
        button_frame = ctk.CTkFrame(top_section, fg_color="transparent")
        button_frame.pack(side="right", padx=10, fill="y")
        
        ctk.CTkLabel(inputs_frame, text="Nombre Completo:").pack(anchor="w")
        e_nom = ctk.CTkEntry(inputs_frame, width=400)
        e_nom.pack(pady=5, anchor="w")
        
        row1 = ctk.CTkFrame(inputs_frame, fg_color="transparent")
        row1.pack(fill="x", pady=5)
        
        ctk.CTkLabel(row1, text="Privilegio:").pack(side="left")
        e_priv = ctk.CTkComboBox(row1, values=["Anciano", "Siervo Min.", "Publicador"], width=150)
        e_priv.set("Publicador")
        e_priv.pack(side="left", padx=5)
        
        ctk.CTkLabel(row1, text="Género:").pack(side="left", padx=5)
        e_gen = ctk.CTkComboBox(row1, values=["M", "F"], width=70)
        e_gen.set("M")
        e_gen.pack(side="left", padx=5)
        
        row2 = ctk.CTkFrame(inputs_frame, fg_color="transparent")
        row2.pack(fill="x", pady=5)
        
        ctk.CTkLabel(row2, text="Edad:").pack(side="left")
        e_edad = ctk.CTkEntry(row2, width=60)
        e_edad.insert(0, "30")
        e_edad.pack(side="left", padx=5)
        
        ctk.CTkLabel(row2, text="¿Menor?:").pack(side="left", padx=5)
        e_men = ctk.CTkComboBox(row2, values=["No", "Si"], width=70)
        e_men.set("No")
        e_men.pack(side="left", padx=5)

        ctk.CTkLabel(inputs_frame, text="Teléfono (WhatsApp):").pack(anchor="w", pady=(5,0))
        e_tel_new = ctk.CTkEntry(inputs_frame, width=400, placeholder_text="Ej: +584121234567")
        e_tel_new.pack(pady=5, anchor="w")
        
        btn_save = ctk.CTkButton(button_frame, text="AÑADIR PUBLICADOR", fg_color=COLOR_SUCCESS, height=45, width=150)
        btn_save.pack(pady=40)
        
        ctk.CTkLabel(main_f, text="Habilidades / Permisos (Sugeridos según perfil):", font=("Arial", 12, "bold")).pack(anchor="w", pady=(10, 5))
        
        # Frame scrollable para habilidades
        scroll_f = ctk.CTkScrollableFrame(main_f, width=420, height=300)
        scroll_f.pack(pady=5, fill="both", expand=True)
        
        hab_vars = {}
        hab_chks = {}
        habilidades = [
            ("TES_SEP", "--- TESOROS DE LA BIBLIA ---"),
            ("Hab_Puede_Presidir", "Puede presidir"),
            ("Hab_Tes_Discurso", "Discurso"),
            ("Hab_Tes_Perlas", "Busquemos perlas escondidas"),
            ("Hab_Lectura", "Lectura de la biblia"),
            
            ("MAE_SEP", "--- SEAMOS MEJORES MAESTROS ---"),
            ("Hab_Mae_Conversacion_Enc", "Empiece conversaciones (Encargado)"),
            ("Hab_Mae_Conversacion_Ayu", "Empiece conversaciones (Ayudante)"),
            ("Hab_Mae_Revisita_Enc", "Haga revisitas (Encargado)"),
            ("Hab_Mae_Revisita_Ayu", "Haga revisitas (Ayudante)"),
            ("Hab_Mae_Discipulos_Enc", "Haga discípulos (Encargado)"),
            ("Hab_Mae_Discipulos_Ayu", "Haga discípulos (Ayudante)"),
            ("Hab_Mae_Creencias_Esc_Enc", "Explique sus creencias (Escenif. Encargado)"),
            ("Hab_Mae_Creencias_Esc_Ayu", "Explique sus creencias (Escenif. Ayudante)"),
            ("Hab_Mae_Creencias_Dis", "Explique sus creencias (Discurso)"),
            ("Hab_Mae_Discurso", "Discurso"),
            
            ("VIDA_SEP", "--- NUESTRA VIDA CRISTIANA ---"),
            ("Hab_Vida_Parte1", "Vida Cristiana"),
            ("Hab_Vida_Parte2", "Vida Cristiana"),
            ("Hab_Vida_Locales", "Necesidades locales"),
            ("Hab_Estudio_Conductor", "Estudio bíblico (Conductor)"),
            ("Hab_Estudio_Lector", "Estudio bíblico (Lector)"),
            ("Hab_Oracion", "Puede hacer Oraciones"),
            
            ("SERV_SEP", "--- SERVICIOS MECÁNICOS ---"),
            ("Hab_Sonido", "Audio / Sonido"),
            ("Hab_Mics", "Micrófonos"),
            ("Hab_Plataforma", "Plataforma"),
            ("Hab_Acomodador", "Acomodador / Ujier")
        ]
        
        for key, text in habilidades:
            if "_SEP" in key:
                ctk.CTkLabel(scroll_f, text=text, font=("Arial", 10, "bold"), text_color="gray").pack(anchor="w", pady=(5, 2))
                continue
            
            var = ctk.BooleanVar(value=False)
            chk = ctk.CTkCheckBox(scroll_f, text=text, variable=var, font=("Arial", 11))
            chk.pack(anchor="w", pady=1)
            hab_vars[key] = var
            hab_chks[key] = chk
 
        def actualizar_sugerencias(*args):
            priv = e_priv.get()
            gen = e_gen.get()
            menor = e_men.get() == "Si"
            
            # Reset
            for v in hab_vars.values(): v.set(False)
            
            # Restricción de presidir según privilegio/género
            if "Hab_Puede_Presidir" in hab_chks:
                if gen == "M" and priv in ["Anciano", "Siervo Min."] and not menor:
                    hab_chks["Hab_Puede_Presidir"].configure(state="normal")
                else:
                    hab_chks["Hab_Puede_Presidir"].configure(state="disabled")
                    hab_vars["Hab_Puede_Presidir"].set(False)
            
            if gen == "F":
                # Hermanas: Solo maestros (estudiantes)
                for k in [k for k in hab_vars if k.startswith('Hab_Mae_')]:
                    hab_vars[k].set(True)
                hab_vars["Hab_Mae_Discurso"].set(False) # Hermanas no dan discurso público/VMC usualmente solo
                hab_vars["Hab_Mae_Creencias_Dis"].set(False)
            else:
                # Varones: Lectura y Maestros siempre (estudiantes)
                hab_vars["Hab_Lectura"].set(True)
                for k in [k for k in hab_vars if k.startswith('Hab_Mae_')]:
                    hab_vars[k].set(True)
                
                if not menor:
                    # Varones adultos: Oración y servicios base
                    hab_vars["Hab_Oracion"].set(True)
                    hab_vars["Hab_Mics"].set(True)
                    hab_vars["Hab_Plataforma"].set(True)
                    hab_vars["Hab_Acomodador"].set(True)
                    
                    if priv == "Anciano":
                        hab_vars["Hab_Puede_Presidir"].set(True)
                        hab_vars["Hab_Tes_Discurso"].set(True)
                        hab_vars["Hab_Tes_Perlas"].set(True)
                        hab_vars["Hab_Vida_Parte1"].set(True)
                        hab_vars["Hab_Vida_Parte2"].set(True)
                        hab_vars["Hab_Vida_Locales"].set(True)
                        hab_vars["Hab_Estudio_Conductor"].set(True)
                        hab_vars["Hab_Estudio_Lector"].set(True)
                        hab_vars["Hab_Sonido"].set(True)
                    elif priv == "Siervo Min.":
                        hab_vars["Hab_Tes_Perlas"].set(True)
                        hab_vars["Hab_Estudio_Lector"].set(True)
                        hab_vars["Hab_Sonido"].set(True)
                        # Algunos siervos dan partes de VC, se puede marcar manual si aplica
        e_priv.configure(command=actualizar_sugerencias)
        e_gen.configure(command=actualizar_sugerencias)
        e_men.configure(command=actualizar_sugerencias)
        
        # Inicializar sugerencias
        actualizar_sugerencias()

        def save():
            nombre = e_nom.get().strip()
            if not nombre:
                messagebox.showwarning("Falta nombre", "Debe ingresar un nombre.")
                return
            
            new_row = {
                'Nombre': nombre,
                'Privilegio': e_priv.get(),
                'Genero': e_gen.get(),
                'Edad': int(e_edad.get()) if e_edad.get().isdigit() else 30,
                'Es_Menor': e_men.get(),
                'Telefono': e_tel_new.get().strip()
            }
            for key, var in hab_vars.items():
                new_row[key] = "Si" if var.get() else "No"
            
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
            self.actualizar_tabla_hermanos()
            self.guardar_excel_congregacion() # Guardar automáticamente
            modal.destroy()

        btn_save.configure(command=save)

    def modal_habilidades_hermano(self, index):
        """Modal para editar habilidades de un hermano existente"""
        row = self.df.iloc[index]
        modal = ctk.CTkToplevel(self.root)
        modal.title(f"Habilidades - {row['Nombre']}")
        modal.geometry("450x600") # Un poco más alto para las categorías
        modal.after(100, lambda: modal.focus_force())
        modal.grab_set()
        
        ctk.CTkLabel(modal, text=f"Configurar Asignaciones: {row['Nombre']}", font=("Arial", 16, "bold")).pack(pady=10)
        
        tel_frame = ctk.CTkFrame(modal, fg_color="transparent")
        tel_frame.pack(fill="x", padx=40)
        ctk.CTkLabel(tel_frame, text="Teléfono:").pack(side="left")
        e_tel_edit = ctk.CTkEntry(tel_frame, width=200)
        e_tel_edit.insert(0, row.get('Telefono', ''))
        e_tel_edit.pack(side="left", padx=10)
        
        # Frame scrollable interno para que quepan todas las opciones
        scroll_f = ctk.CTkScrollableFrame(modal, width=400, height=450)
        scroll_f.pack(pady=5, padx=10, fill="both", expand=True)
        
        hab_vars = {}
        habilidades = [
            ("TES_SEP", "--- TESOROS DE LA BIBLIA ---"),
            ("Hab_Puede_Presidir", "Puede presidir"),
            ("Hab_Tes_Discurso", "Discurso"),
            ("Hab_Tes_Perlas", "Busquemos perlas escondidas"),
            ("Hab_Lectura", "Lectura de la biblia"),
            
            ("MAE_SEP", "--- SEAMOS MEJORES MAESTROS ---"),
            ("Hab_Mae_Conversacion_Enc", "Empiece conversaciones (Encargado)"),
            ("Hab_Mae_Conversacion_Ayu", "Empiece conversaciones (Ayudante)"),
            ("Hab_Mae_Revisita_Enc", "Haga revisitas (Encargado)"),
            ("Hab_Mae_Revisita_Ayu", "Haga revisitas (Ayudante)"),
            ("Hab_Mae_Discipulos_Enc", "Haga discípulos (Encargado)"),
            ("Hab_Mae_Discipulos_Ayu", "Haga discípulos (Ayudante)"),
            ("Hab_Mae_Creencias_Esc_Enc", "Explique sus creencias (Escenif. Encargado)"),
            ("Hab_Mae_Creencias_Esc_Ayu", "Explique sus creencias (Escenif. Ayudante)"),
            ("Hab_Mae_Creencias_Dis", "Explique sus creencias (Discurso)"),
            ("Hab_Mae_Discurso", "Discurso"),
            
            ("VIDA_SEP", "--- NUESTRA VIDA CRISTIANA ---"),
            ("Hab_Vida_Parte1", "Vida Cristiana"),
            ("Hab_Vida_Parte2", "Vida Cristiana"),
            ("Hab_Vida_Locales", "Necesidades locales"),
            ("Hab_Estudio_Conductor", "Estudio bíblico (Conductor)"),
            ("Hab_Estudio_Lector", "Estudio bíblico (Lector)"),
            ("Hab_Oracion", "Puede hacer Oraciones"),
            
            ("SERV_SEP", "--- SERVICIOS MECÁNICOS ---"),
            ("Hab_Sonido", "Audio / Sonido"),
            ("Hab_Mics", "Micrófonos"),
            ("Hab_Plataforma", "Plataforma"),
            ("Hab_Acomodador", "Acomodador / Ujier")
        ]
        
        for key, text in habilidades:
            if "_SEP" in key:
                ctk.CTkLabel(scroll_f, text=text, font=("Arial", 11, "bold"), text_color="gray").pack(anchor="w", padx=20, pady=(10, 2))
                continue
                
            val = True if row.get(key, "No") == "Si" else False
            var = ctk.BooleanVar(value=val)
            chk = ctk.CTkCheckBox(scroll_f, text=text, variable=var)
            chk.pack(anchor="w", padx=40, pady=2)
            
            # Deshabilitar si no es Anciano ni Siervo Min., o si es una hermana
            if key == "Hab_Puede_Presidir":
                if row.get('Genero') != 'M' or row.get('Privilegio') not in ["Anciano", "Siervo Min."]:
                    var.set(False)
                    chk.configure(state="disabled")
                    
            hab_vars[key] = var
            
        def save_habs():
            for key, var in hab_vars.items():
                self.df.at[index, key] = "Si" if var.get() else "No"
            self.df.at[index, 'Telefono'] = e_tel_edit.get().strip()
            self.actualizar_tabla_hermanos()
            self.guardar_excel_congregacion() # Guardar automáticamente
            modal.destroy()
            
        ctk.CTkButton(modal, text="GUARDAR HABILIDADES", command=save_habs, fg_color=COLOR_ACCENT, height=40).pack(pady=15)

    def eliminar_hermano(self, index):
        if messagebox.askyesno("Confirmar", "¿Eliminar a este publicador?"):
            self.df = self.df.drop(index).reset_index(drop=True)
            self.actualizar_tabla_hermanos()
            self.guardar_excel_congregacion() # Guardar automáticamente

    def guardar_excel_congregacion(self, silent=False):
        # Lógica para guardar el DF de vuelta al archivo
        try:
            archivo = self.archivo_excel
            self.df.to_excel(archivo, index=False)
            if not silent:
                messagebox.showinfo("Éxito", "Lista de congregación actualizada correctamente.")
            self.inicializar_listas_habilidades() # Actualizar listas en memoria sin recargar todo el archivo
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"No se pudo guardar: {e}")
            print(f"Error al guardar excel: {e}")

    def _setup_tab_estadisticas(self):
        parent = self.tab_stats
        for widget in parent.winfo_children(): widget.destroy()
        
        ctk.CTkLabel(parent, text="📊 RANKING DE PARTICIPACIÓN", font=("Arial", 18, "bold")).pack(pady=10)
        ctk.CTkLabel(parent, text="Muestra cuántas veces ha participado cada hermano (según historial)", font=("Arial", 12)).pack()
        
        stats_frame = ctk.CTkScrollableFrame(parent, fg_color="transparent", width=600, height=500)
        stats_frame.pack(pady=10, padx=20, fill="both", expand=True)
        
        # Calcular conteos desde el historial
        conteos = {}
        for nombre, ultima_semana in self.historial_maestros.items():
            # Aquí el historial solo guarda la 'última vez'. Para estadísticas reales, 
            # necesitaríamos un historial de todas las participaciones.
            # Por ahora, simularemos o usaremos lo que hay.
            conteos[nombre] = conteos.get(nombre, 0) + 1
            
        sorted_stats = sorted(conteos.items(), key=lambda x: x[1], reverse=True)
        
        for nombre, cant in sorted_stats:
            row = ctk.CTkFrame(stats_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=nombre, width=250, anchor="w").pack(side="left", padx=10)
            
            # Barra de progreso visual para la estadística
            progress = ctk.CTkProgressBar(row, width=200)
            progress.pack(side="left", padx=10)
            progress.set(min(cant / 10, 1.0)) # Normalizado a 10 participaciones
            
            ctk.CTkLabel(row, text=f"{cant} veces").pack(side="left", padx=10)

    def _setup_tab_configuracion(self):
        parent = self.tab_config
        
        ctk.CTkLabel(parent, text="⚙️ CONFIGURACIÓN", font=("Arial", 18, "bold")).pack(pady=10)
        
        # Modo Oscuro
        self.dark_mode_var = ctk.BooleanVar(value=False)
        ctk.CTkSwitch(parent, text="Modo Oscuro Premium", variable=self.dark_mode_var,
                     command=self.toggle_dark_mode).pack(pady=10)
        
        # WhatsApp Settings
        ws_frame = ctk.CTkFrame(parent, border_width=1, border_color="#25D366")
        ws_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        ctk.CTkLabel(ws_frame, text="📱 CONFIGURACIÓN DE WHATSAPP", font=("Arial", 14, "bold"), text_color="#25D366").pack(pady=8)
        
        # Selector de método de envío
        metodo_frame = ctk.CTkFrame(ws_frame, fg_color="transparent")
        metodo_frame.pack(fill="x", padx=15, pady=5)
        ctk.CTkLabel(metodo_frame, text="Método de envío:", font=("Arial", 11)).pack(side="left")
        self.ws_metodo = ctk.CTkComboBox(metodo_frame, values=["WhatsApp Desktop (App)", "WhatsApp Web (Navegador)"], width=250)
        self.ws_metodo.set("WhatsApp Desktop (App)")
        self.ws_metodo.pack(side="left", padx=10)
        
        # Delay entre mensajes
        delay_frame = ctk.CTkFrame(ws_frame, fg_color="transparent")
        delay_frame.pack(fill="x", padx=15, pady=5)
        ctk.CTkLabel(delay_frame, text="Segundos entre mensajes (envío masivo):", font=("Arial", 11)).pack(side="left")
        self.ws_delay = ctk.CTkEntry(delay_frame, width=60)
        self.ws_delay.insert(0, "3")
        self.ws_delay.pack(side="left", padx=10)
        
        # Humanizar mensajes
        humanizar_frame = ctk.CTkFrame(ws_frame, fg_color="transparent")
        humanizar_frame.pack(fill="x", padx=15, pady=5)
        self.ws_humanizar_var = ctk.BooleanVar(value=True)
        self.ws_humanizar_switch = ctk.CTkSwitch(humanizar_frame, text="Humanizar mensajes (varía saludos, despedidas y emojis de forma aleatoria)", 
                                                 variable=self.ws_humanizar_var, progress_color="#25D366")
        self.ws_humanizar_switch.pack(side="left")
        
        ctk.CTkLabel(ws_frame, text="Plantilla de mensaje:", font=("Arial", 12, "bold")).pack(pady=(10, 5), padx=15, anchor="w")
        ctk.CTkLabel(ws_frame, text="Variables: {nombre}, {parte}, {fecha}, {seccion}", font=("Arial", 10), text_color="gray").pack(padx=15, anchor="w")
        self.ws_template = ctk.CTkTextbox(ws_frame, height=100, width=500)
        self.ws_template.pack(pady=5, padx=15)
        self.ws_template.insert("0.0", "Hola {nombre}, 🙂\n\nTe informamos que tienes la siguiente asignación:\n\n📋 *{parte}*\n📅 *{fecha}*\n📖 Sección: {seccion}\n\n¡Muchas gracias por tu servicio! 🙏")
        
        ctk.CTkLabel(ws_frame, text="", height=5).pack()  # Spacer
        
        # Backups
        ctk.CTkButton(parent, text="📂 Crear Copia de Seguridad", command=self.crear_backup,
                     fg_color=COLOR_ACCENT).pack(pady=20)

    def toggle_dark_mode(self):
        mode = "dark" if self.dark_mode_var.get() else "light"
        ctk.set_appearance_mode(mode)

    def crear_backup(self):
        try:
            import shutil
            backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
            if not os.path.exists(backup_dir): os.makedirs(backup_dir)
            
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Backup Excel
            shutil.copy2(self.archivo_excel, os.path.join(backup_dir, f"Congregacion_{fecha_str}.xlsx"))
            # Backup Historial
            if os.path.exists(self.historial_archivo):
                shutil.copy2(self.historial_archivo, os.path.join(backup_dir, f"Historial_{fecha_str}.json"))
                
            messagebox.showinfo("Éxito", f"Copia de seguridad creada en:\n{backup_dir}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el backup: {e}")

    def descargar_desde_jw(self):
        """Descarga datos de JW.org en un hilo separado"""
        if not JW_AVAILABLE or not self.scraper:
            messagebox.showerror("Error", "El módulo jw_scraper no está disponible.\nVerifique que jw_scraper.py esté en la misma carpeta.")
            return
        
        self.btn_descargar.configure(state="disabled", text="Descargando...")
        self.status_label.configure(text="Conectando a JW.org...", text_color="orange")
        self.progress_bar.set(0)
        
        def _descargar():
            try:
                year = int(self.year_combo.get())
                bimestre_nombre = self.bimestre_combo.get()
                
                # Buscar el mes de inicio del bimestre seleccionado
                mes_inicio = 5  # default
                for nombre, mes in JWScraper.get_bimestres():
                    if nombre == bimestre_nombre:
                        mes_inicio = mes
                        break
                
                semanas_iso = JWScraper.calcular_semanas_bimestre(year, mes_inicio)
                
                def update_progress(current, total, info):
                    self.progress_bar.set(current / total)
                    self.status_label.configure(text=f"Descargando {current}/{total}: {info}")
                
                resultados = []
                total = len(semanas_iso)
                for i, (iso_year, iso_week) in enumerate(semanas_iso):
                    try:
                        datos = self.scraper.obtener_semana(iso_year, iso_week)
                        resultados.append(datos)
                    except Exception as e:
                        resultados.append({
                            'error': str(e), 'fecha': f'Semana {iso_week} (error)',
                            'lectura_biblica': '', 'maestros': [],
                        })
                    self.root.after(0, update_progress, i + 1, total, 
                                   resultados[-1].get('fecha', f'Semana {iso_week}'))
                
                self.datos_jw = resultados
                self.root.after(0, self._aplicar_datos_jw)
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error al descargar:\n{str(e)}"))
            finally:
                self.root.after(0, lambda: self.btn_descargar.configure(state="normal", text="⬇ DESCARGAR GUÍA"))
        
        threading.Thread(target=_descargar, daemon=True).start()
    
    def _aplicar_datos_jw(self):
        """Aplica los datos descargados a los formularios de semanas"""
        if not self.datos_jw:
            return
        
        # Actualizar número de semanas y configurar
        num = len(self.datos_jw)
        if num <= 10:
            self.num_semanas.set(str(num))
        
        # Crear formularios con datos pre-llenados
        self._crear_formularios_semanas(self.datos_jw)
        
        exitosas = sum(1 for d in self.datos_jw if 'error' not in d)
        self.status_label.configure(
            text=f"✅ {exitosas}/{num} semanas descargadas", text_color="green"
        )
        self.progress_bar.set(1)
    
    def configurar_semanas(self):
        """Crea formularios vacíos para entrada manual"""
        self._crear_formularios_semanas(None)
    
    def _crear_formularios_semanas(self, datos_precargados=None):
        """Crea los formularios para cada semana del bimestre"""
        # Limpiar frame
        for widget in self.semanas_frame.winfo_children():
            widget.destroy()
        
        self.semanas = []
        num = len(datos_precargados) if datos_precargados else int(self.num_semanas.get())
        
        for i in range(num):
            jw_data = datos_precargados[i] if datos_precargados and i < len(datos_precargados) else None
            tiene_error = jw_data and 'error' in jw_data
            
            semana_frame = ctk.CTkFrame(self.semanas_frame)
            semana_frame.pack(pady=8, fill="x", padx=10)
            
            # Encabezado de semana
            header_text = f"📅 SEMANA {i+1}"
            if jw_data and not tiene_error:
                header_text += f"  •  {jw_data.get('fecha', '')}"
                if jw_data.get('cancion_inicial'):
                    header_text += f"  |  🎵 {jw_data['cancion_inicial']}, {jw_data.get('cancion_intermedia','')}, {jw_data.get('cancion_final','')}"
            
            header_color = "red" if tiene_error else None
            ctk.CTkLabel(semana_frame, text=header_text, font=("Arial", 13, "bold"),
                        text_color=header_color).grid(row=0, column=0, columnspan=6, pady=5, sticky="w")
            
            # Fecha
            ctk.CTkLabel(semana_frame, text="Fecha:").grid(row=1, column=0, padx=5, pady=2, sticky="e")
            fecha_entry = ctk.CTkEntry(semana_frame, width=180, placeholder_text="Ej: 2-8 de octubre")
            fecha_entry.grid(row=1, column=1, padx=5, pady=2, sticky="w")
            if jw_data and not tiene_error:
                fecha_entry.insert(0, jw_data.get('fecha', ''))
            
            # Lectura
            ctk.CTkLabel(semana_frame, text="Lectura:").grid(row=1, column=2, padx=5, pady=2, sticky="e")
            lectura_entry = ctk.CTkEntry(semana_frame, width=180, placeholder_text="Ej: Isaías 58, 59")
            lectura_entry.grid(row=1, column=3, padx=5, pady=2, sticky="w")
            if jw_data and not tiene_error:
                lectura_entry.insert(0, jw_data.get('lectura_biblica', ''))
            
            # Asignaciones de Maestros
            ctk.CTkLabel(semana_frame, text="SEAMOS MEJORES MAESTROS:", 
                        font=("Arial", 11, "bold")).grid(row=2, column=0, columnspan=6, pady=(8,2), sticky="w")
            
            nombres_maestros = [
                "Empiece conversaciones", "Haga revisitas", "Haga discípulos",
                "Explique sus creencias", "Discurso"
            ]
            
            vars_maestros = []
            if jw_data and not tiene_error and jw_data.get('maestros'):
                # Si hay datos de JW.org, crear un checkbox por cada parte real para permitir duplicados
                for idx, m in enumerate(jw_data['maestros']):
                    tipo = m['tipo']
                    var = ctk.BooleanVar(value=True)
                    # Mostrar el número y mins para identificarla bien
                    txt = f"{m['numero']}. {tipo} ({m['mins']} min)"
                    chk = ctk.CTkCheckBox(semana_frame, text=txt, variable=var, font=("Arial", 10))
                    chk.grid(row=3, column=idx, padx=5, pady=2, sticky="w")
                    vars_maestros.append((m['titulo'], var))
            else:
                # Si es manual o no hay datos, usar el set estándar de 5 tipos
                for idx, asig in enumerate(nombres_maestros):
                    val = True if idx < 3 else False
                    var = ctk.BooleanVar(value=val)
                    chk = ctk.CTkCheckBox(semana_frame, text=asig, variable=var)
                    chk.grid(row=3, column=idx, padx=10, pady=2, sticky="w")
                    vars_maestros.append((asig, var))
            
            # Mostrar detalle de asignaciones si hay datos de JW.org
            if jw_data and not tiene_error and jw_data.get('maestros'):
                detalles = "  |  ".join([f"{m['tipo']} ({m['mins']}min)" for m in jw_data['maestros']])
                ctk.CTkLabel(semana_frame, text=f"→ Partes detectadas: {detalles}", 
                            font=("Arial", 10), text_color="gray").grid(
                    row=4, column=0, columnspan=6, padx=20, sticky="w")
            
            # Mostrar info de Vida Cristiana si hay datos
            if jw_data and not tiene_error and jw_data.get('vida_cristiana'):
                vc_info = "  |  ".join([f"{v['titulo']} ({v['mins']}min)" for v in jw_data['vida_cristiana']])
                ctk.CTkLabel(semana_frame, text=f"VIDA CRISTIANA: {vc_info}", 
                            font=("Arial", 10), text_color="#555").grid(
                    row=5, column=0, columnspan=6, padx=20, pady=(2,0), sticky="w")
            
            # --- Cancelación de semana ---
            cancel_frame = ctk.CTkFrame(semana_frame, fg_color="transparent")
            cancel_frame.grid(row=6, column=0, columnspan=6, padx=10, pady=(5, 5), sticky="w")
            
            no_reunion_var = ctk.BooleanVar(value=False)
            chk_cancelar = ctk.CTkCheckBox(cancel_frame, text="No hay reunión esta semana", variable=no_reunion_var, text_color="red")
            chk_cancelar.pack(side="left", padx=5)
            
            ctk.CTkLabel(cancel_frame, text="Motivo:").pack(side="left", padx=5)
            motivo_entry = ctk.CTkEntry(cancel_frame, width=250, placeholder_text="Ej. Asamblea de Circuito", state="disabled")
            motivo_entry.pack(side="left", padx=5)
            
            def toggle_cancel(var=no_reunion_var, m_ent=motivo_entry, f_ent=fecha_entry, l_ent=lectura_entry, m_vars=vars_maestros):
                if var.get():
                    m_ent.configure(state="normal")
                    f_ent.configure(state="disabled")
                    l_ent.configure(state="disabled")
                    for _, mv in m_vars:
                        pass # No podemos deshabilitar variables facilmente sin la ref del widget, pero al menos activamos motivo
                else:
                    m_ent.configure(state="disabled")
                    m_ent.delete(0, 'end')
                    f_ent.configure(state="normal")
                    l_ent.configure(state="normal")
            
            chk_cancelar.configure(command=toggle_cancel)

            self.semanas.append({
                'numero': i+1,
                'fecha_entry': fecha_entry,
                'lectura_entry': lectura_entry,
                'maestros_vars': vars_maestros,
                'frame': semana_frame,
                'jw_data': jw_data,
                'no_reunion_var': no_reunion_var,
                'motivo_entry': motivo_entry
            })
        
        modo = "desde JW.org" if datos_precargados else "manual"
        ctk.CTkLabel(self.semanas_frame, text=f"✅ {num} semanas configuradas ({modo}). Presione 'GENERAR BIMESTRE'", 
                    font=("Arial", 12, "bold"), text_color="green").pack(pady=10)
    
    def es_menor(self, nombre):
        if self.df is None or self.df.empty:
            return False
        match = self.df[self.df['Nombre'] == nombre]
        if not match.empty:
            return match.iloc[0].get('Es_Menor', 'No') == 'Si'
        return False
        
    def comparten_apellido(self, nombre1, nombre2):
        if not nombre1 or not nombre2 or nombre1 == "__________________" or nombre2 == "__________________":
            return False
        ap1 = nombre1.split()[-1].lower() if len(nombre1.split()) > 1 else nombre1.lower()
        ap2 = nombre2.split()[-1].lower() if len(nombre2.split()) > 1 else nombre2.lower()
        return ap1 == ap2

    def inicializar_listas_habilidades(self):
        """Prepara listas filtradas por habilidades para la generación"""
        if self.df is None: return
        
        # --- TESOROS ---
        self.list_puede_presidir = self.df[(self.df['Hab_Puede_Presidir'] == 'Si') & (self.df['Genero'] == 'M') & (self.df['Privilegio'].isin(['Anciano', 'Siervo Min.']))]['Nombre'].tolist()
        self.list_tes_discurso = self.df[(self.df['Hab_Tes_Discurso'] == 'Si') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        self.list_tes_perlas = self.df[(self.df['Hab_Tes_Perlas'] == 'Si') & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        self.list_lectores_biblia = self.df[self.df['Hab_Lectura'] == 'Si']['Nombre'].tolist()
        
        # --- MAESTROS ---
        # Listas para cada tipo específico
        self.list_mae_conv_enc = self.df[self.df['Hab_Mae_Conversacion_Enc'] == 'Si']['Nombre'].tolist()
        self.list_mae_conv_ayu = self.df[self.df['Hab_Mae_Conversacion_Ayu'] == 'Si']['Nombre'].tolist()
        self.list_mae_rev_enc = self.df[self.df['Hab_Mae_Revisita_Enc'] == 'Si']['Nombre'].tolist()
        self.list_mae_rev_ayu = self.df[self.df['Hab_Mae_Revisita_Ayu'] == 'Si']['Nombre'].tolist()
        self.list_mae_disc_enc = self.df[self.df['Hab_Mae_Discipulos_Enc'] == 'Si']['Nombre'].tolist()
        self.list_mae_disc_ayu = self.df[self.df['Hab_Mae_Discipulos_Ayu'] == 'Si']['Nombre'].tolist()
        self.list_mae_cree_esc_enc = self.df[self.df['Hab_Mae_Creencias_Esc_Enc'] == 'Si']['Nombre'].tolist()
        self.list_mae_cree_esc_ayu = self.df[self.df['Hab_Mae_Creencias_Esc_Ayu'] == 'Si']['Nombre'].tolist()
        self.list_mae_cree_dis = self.df[self.df['Hab_Mae_Creencias_Dis'] == 'Si']['Nombre'].tolist()
        self.list_mae_discurso = self.df[self.df['Hab_Mae_Discurso'] == 'Si']['Nombre'].tolist()

        # Listas genéricas (compatibilidad)
        cols_mae = [c for c in self.df.columns if c.startswith('Hab_Mae_')]
        mask_mae = (self.df[cols_mae] == 'Si').any(axis=1)
        self.list_maestros_varones = self.df[mask_mae & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        self.list_maestros_mujeres = self.df[mask_mae & (self.df['Genero'] == 'F')]['Nombre'].tolist()
            
        self.list_oracion = self.df[self.df['Hab_Oracion'] == 'Si']['Nombre'].tolist()
        
        # --- VIDA CRISTIANA ---
        self.list_vida_p1 = self.df[self.df['Hab_Vida_Parte1'] == 'Si']['Nombre'].tolist()
        self.list_vida_p2 = self.df[self.df['Hab_Vida_Parte2'] == 'Si']['Nombre'].tolist()
        self.list_vida_locales = self.df[self.df['Hab_Vida_Locales'] == 'Si']['Nombre'].tolist()
        
        self.list_estudio_conductor = self.df[self.df['Hab_Estudio_Conductor'] == 'Si']['Nombre'].tolist()
        self.list_estudio_lector = self.df[self.df['Hab_Estudio_Lector'] == 'Si']['Nombre'].tolist()
        
        # --- SERVICIOS ---
        self.list_serv_sonido = self.df[self.df['Hab_Sonido'] == 'Si']['Nombre'].tolist()
        self.list_serv_mics = self.df[self.df['Hab_Mics'] == 'Si']['Nombre'].tolist()
        self.list_serv_plataforma = self.df[self.df['Hab_Plataforma'] == 'Si']['Nombre'].tolist()
        self.list_serv_acomodador = self.df[self.df['Hab_Acomodador'] == 'Si']['Nombre'].tolist()
        
        # Pools para rotación circular
        cols_tes = ['Hab_Tes_Discurso', 'Hab_Tes_Perlas', 'Hab_Lectura']
        mask_tes = (self.df[cols_tes] == 'Si').any(axis=1)
        self.list_tesoros = self.df[mask_tes & (self.df['Genero'] == 'M')]['Nombre'].tolist()
        
        cols_vida = ['Hab_Vida_Parte1', 'Hab_Vida_Parte2', 'Hab_Vida_Locales']
        mask_vida = (self.df[cols_vida] == 'Si').any(axis=1)
        self.list_vida = self.df[mask_vida & (self.df['Genero'] == 'M')]['Nombre'].tolist()

    def inicializar_pools(self):
        self.inicializar_listas_habilidades()
        
        elegibles_tesoros = list(self.list_tesoros)
        random.shuffle(elegibles_tesoros)
        self.pool_tesoros = elegibles_tesoros
        
        elegibles_vida = list(self.list_vida)
        random.shuffle(elegibles_vida)
        self.pool_vida = elegibles_vida

    def asignar_desde_pool(self, pool, lista_candidatos_base, asignados_semana, evitar=None):
        """Extrae del pool garantizando que todos pasen antes de repetir."""
        # Filtrar pool actual
        disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
        
        if not disponibles:
            # Si se agotaron los elegibles, rellenar con la lista base
            base = list(lista_candidatos_base)
            random.shuffle(base)
            pool.extend(base)
            disponibles = [p for p in pool if p not in asignados_semana and p != evitar and p in lista_candidatos_base]
            
        if not disponibles:
            return "__________________"
            
        seleccionado = disponibles[0]
        pool.remove(seleccionado)
        asignados_semana.add(seleccionado)
        return seleccionado

    def asignar_persona(self, lista_candidatos, asignados_semana, evitar=None):
        """Asigna aleatorio normal (sin pool)"""
        disponibles = [p for p in lista_candidatos if p not in asignados_semana and p != evitar]
        if not disponibles:
            return "__________________"
        seleccionado = random.choice(disponibles)
        asignados_semana.add(seleccionado)
        return seleccionado

    def candidato_valido_maestros(self, persona, semana_num):
        ultima = self.historial_maestros.get(persona, -99)
        return (semana_num - ultima) >= 6

    def asignar_pareja_generica(self, asignados_semana, semana_num, pool_titulares, pool_ayudantes_base, permite_familiar_opuesto=False):
        separador = " // "
        
        # 1. Buscar titular con cooldown de 6 semanas
        titulares_disp = [t for t in pool_titulares if t not in asignados_semana and self.candidato_valido_maestros(t, semana_num)]
        if not titulares_disp:
            titulares_disp = [t for t in pool_titulares if t not in asignados_semana]
            
        if not titulares_disp: return "__________________ // __________________"
        titular = random.choice(titulares_disp)
        
        # 2. Determinar pool de ayudantes permitido
        pool_ayudantes = list(pool_ayudantes_base)
        
        if permite_familiar_opuesto:
            # 30% de probabilidad de buscar familiar de sexo opuesto (mismo sexo o familiar)
            if random.random() < 0.3:
                es_mujer = titular in self.hermanas
                pool_opuesto = self.todos_varones if es_mujer else self.hermanas
                fam_opuestos = [p for p in pool_opuesto if self.comparten_apellido(titular, p)]
                if fam_opuestos:
                    pool_ayudantes.extend(fam_opuestos * 3) # Aumentar probabilidad
        
        # 3. Buscar ayudante con cooldown
        ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular and self.candidato_valido_maestros(a, semana_num)]
        if not ayudantes_disp:
            ayudantes_disp = [a for a in pool_ayudantes if a not in asignados_semana and a != titular]
            
        # 4. Aplicar regla de menores/familias
        ayudantes_validos = []
        for a in ayudantes_disp:
            if self.es_menor(a) and not self.es_menor(titular):
                if not self.comparten_apellido(titular, a):
                    continue
            ayudantes_validos.append(a)
            
        if not ayudantes_validos:
            asignados_semana.add(titular)
            self.historial_maestros[titular] = semana_num
            return f"{titular} // __________________"
            
        ayudante = random.choice(ayudantes_validos)
        
        asignados_semana.add(titular)
        asignados_semana.add(ayudante)
        self.historial_maestros[titular] = semana_num
        self.historial_maestros[ayudante] = semana_num
        
        return f"{titular}{separador}{ayudante}"

    def asignar_estudiante_solo_con_pool(self, asignados_semana, semana_num, pool):
        disp = [p for p in pool if p not in asignados_semana and self.candidato_valido_maestros(p, semana_num)]
        if not disp:
            disp = [p for p in pool if p not in asignados_semana]
        if not disp: return "__________________"
        p = random.choice(disp)
        asignados_semana.add(p)
        self.historial_maestros[p] = semana_num
        return p

    def asignar_estudiante_solo(self, asignados_semana, semana_num, es_hermanas=True):
        # Para "Explique sus creencias" (solo 1 persona) - Versión antigua/genérica
        pool = self.list_maestros_mujeres if es_hermanas else self.list_maestros_varones
        return self.asignar_estudiante_solo_con_pool(asignados_semana, semana_num, pool)

    def generar_semana(self, semana_info, index_semana, fecha, lectura, asignaciones_maestros):
        """Genera las asignaciones para una semana específica usando índices consecutivos para el cooldown"""
        asignados = set()
        
        # TESOROS (siempre) - Usa el sistema de Pool
        pool_presidir = self.list_puede_presidir if hasattr(self, 'list_puede_presidir') and self.list_puede_presidir else self.ancianos
        presidente = self.asignar_persona(pool_presidir, asignados)
        oracion = self.asignar_persona(self.list_oracion, asignados)
        
        # Tesoros 1 y 2 (Discurso y Perlas usualmente)
        num1_tesoros = self.asignar_desde_pool(self.pool_tesoros, self.list_tes_discurso if self.list_tes_discurso else self.list_tesoros, asignados)
        num2_tesoros = self.asignar_desde_pool(self.pool_tesoros, self.list_tes_perlas if self.list_tes_perlas else self.list_tesoros, asignados)
        
        # LECTURA DE LA BIBLIA
        lectura_biblia = self.asignar_desde_pool(self.pool_tesoros, self.list_lectores_biblia, asignados)
        
        # MAESTROS (según lo que tenga la semana)
        maestros_asignaciones = []
        for asig in asignaciones_maestros:
            asig_lower = asig.lower()
            
            # Determinar pools según el tipo de parte
            if "convers" in asig_lower:
                p_enc, p_ayu = self.list_mae_conv_enc, self.list_mae_conv_ayu
            elif "revisita" in asig_lower:
                p_enc, p_ayu = self.list_mae_rev_enc, self.list_mae_rev_ayu
            elif "discipulo" in asig_lower or "discípulo" in asig_lower:
                p_enc, p_ayu = self.list_mae_disc_enc, self.list_mae_disc_ayu
            elif "escenificaci" in asig_lower:
                p_enc, p_ayu = self.list_mae_cree_esc_enc, self.list_mae_cree_esc_ayu
            else:
                p_enc, p_ayu = None, None

            if p_enc:
                # Partes que requieren Pareja (Titular + Ayudante)
                # Regla 80% hermanas para estas partes
                usar_hermanas = (random.random() < 0.8) and bool(self.hermanas)
                
                # Filtrar pools por género si aplica
                f_enc = [n for n in p_enc if (n in self.hermanas if usar_hermanas else n in self.todos_varones)]
                f_ayu = [n for n in p_ayu if (n in self.hermanas if usar_hermanas else n in self.todos_varones)]
                
                if not f_enc: f_enc = p_enc # Fallback
                if not f_ayu: f_ayu = p_ayu # Fallback
                
                res = self.asignar_pareja_generica(asignados, index_semana, f_enc, f_ayu, permite_familiar_opuesto=True)
                maestros_asignaciones.append((asig, res))
            elif "creencias" in asig_lower:
                # Explique sus creencias (Discurso)
                pool = self.list_mae_cree_dis if self.list_mae_cree_dis else self.list_maestros_varones
                res = self.asignar_estudiante_solo_con_pool(asignados, index_semana, pool)
                maestros_asignaciones.append((asig, res))
            else:
                # Discurso estándar de maestros
                pool = self.list_mae_discurso if self.list_mae_discurso else self.list_maestros_varones
                res = self.asignar_estudiante_solo_con_pool(asignados, index_semana, pool)
                maestros_asignaciones.append((asig, res))
        
        # VIDA CRISTIANA - Soporte para múltiples partes (según guía JW.org)
        partes_vida = []
        jw_vc = semana_info.get('jw_data') or {}
        vida_cristiana_partes = [p for p in jw_vc.get('vida_cristiana', []) if 'estudio' not in p.get('tipo', '').lower()]

        if vida_cristiana_partes:
            for pv_idx, parte_info in enumerate(vida_cristiana_partes):
                tipo_pv_original = parte_info.get('tipo', f'Parte {pv_idx + 1}')
                if 'necesidades' in tipo_pv_original.lower():
                    # Necesidades de la congregación: solo ancianos
                    tipo_pv_label = 'Necesidades locales'
                    pool_pv = self.ancianos if self.ancianos else (
                        self.list_puede_presidir if hasattr(self, 'list_puede_presidir') and self.list_puede_presidir else self.list_vida
                    )
                    asignado_pv = self.asignar_persona(pool_pv, asignados)
                elif pv_idx == 0:
                    tipo_pv_label = 'Parte 1'
                    pool_pv = self.list_vida_p1 if self.list_vida_p1 else self.list_vida
                    asignado_pv = self.asignar_desde_pool(self.pool_vida, pool_pv, asignados)
                else:
                    tipo_pv_label = f'Parte {pv_idx + 1}'
                    pool_pv = self.list_vida_p2 if self.list_vida_p2 else self.list_vida
                    asignado_pv = self.asignar_desde_pool(self.pool_vida, pool_pv, asignados)
                partes_vida.append((tipo_pv_label, asignado_pv))
        else:
            # Sin datos JW: generar 1 parte por defecto
            pool_v1 = self.list_vida_p1 if self.list_vida_p1 else self.list_vida
            asignado_pv = self.asignar_desde_pool(self.pool_vida, pool_v1, asignados)
            partes_vida.append(('Parte 1', asignado_pv))
        
        # Estudio Bíblico
        estudio_biblico = self.asignar_persona(self.list_estudio_conductor, asignados)
        lector = self.asignar_persona(self.list_estudio_lector, asignados, evitar=estudio_biblico)
        
        # CONSEJERO AUXILIAR
        pool_presidir = self.list_puede_presidir if hasattr(self, 'list_puede_presidir') and self.list_puede_presidir else self.ancianos
        presidencia_aux = self.asignar_persona(pool_presidir, asignados)
        
        # SERVICIOS MECÁNICOS
        sonido1 = self.asignar_persona(self.list_serv_sonido, asignados)
        sonido2 = self.asignar_persona(self.list_serv_sonido, asignados, evitar=sonido1)
        plataforma = self.asignar_persona(self.list_serv_plataforma, asignados)
        mic1 = self.asignar_persona(self.list_serv_mics, asignados)
        mic2 = self.asignar_persona(self.list_serv_mics, asignados, evitar=mic1)
        acomodador1 = self.asignar_persona(self.list_serv_acomodador, asignados)
        acomodador2 = self.asignar_persona(self.list_serv_acomodador, asignados, evitar=acomodador1)
        
        return {
            'semana': semana_info['numero'],
            'fecha': fecha,
            'lectura': lectura,
            'presidente': presidente,
            'oracion': oracion,
            'num1_tesoros': num1_tesoros,
            'num2_tesoros': num2_tesoros,
            'lectura_biblia': lectura_biblia,
            'maestros': maestros_asignaciones,
            'partes_vida': partes_vida,
            'estudio_biblico': estudio_biblico,
            'lector': lector,
            'presidencia_aux': presidencia_aux,
            'sonido': f"{sonido1} / {sonido2}",
            'plataforma': plataforma,
            'microfonos': f"{mic1} / {mic2}",
            'acomodadores': f"{acomodador1} / {acomodador2}",
            'total_asignados': len(asignados)
        }
    
    def generar_bimestre_completo(self):
        """Genera todas las semanas del bimestre"""
        if not self.semanas:
            messagebox.showwarning("Advertencia", "Primero configure las semanas del bimestre.")
            return
        
        self.bimestre_data = []
        
        # Reiniciar variables para la sesión de generación actual
        self.inicializar_pools()
        # Recargar el historial desde el archivo guardado para ignorar intentos de generación previos no confirmados
        self.cargar_historial()
        
        for idx_semana, semana in enumerate(self.semanas):
            index = idx_semana + 1  # Para contar consecutivamente y calcular cooldowns
            
            fecha = semana['fecha_entry'].get().strip()
            if not fecha:
                fecha = f"Semana {semana['numero']}"
            
            lectura = semana['lectura_entry'].get().strip()
            if not lectura:
                lectura = "Lectura semanal"
            
            # Obtener asignaciones de maestros seleccionadas
            asignaciones_maestros = []
            for tipo, var in semana['maestros_vars']:
                if var.get():
                    asignaciones_maestros.append(tipo)
            
            # Generar semana (pasamos el index correlativo en lugar de semana['numero'])
            no_reunion = semana.get('no_reunion_var').get() if 'no_reunion_var' in semana else False
            motivo = semana.get('motivo_entry').get() if 'motivo_entry' in semana else ""
            
            if no_reunion:
                datos_semana = {
                    'no_reunion': True,
                    'motivo': motivo,
                    'fecha': fecha,
                    'semana_num': semana['numero'],
                    'lectura': lectura
                }
            else:
                datos_semana = self.generar_semana(
                    semana, index, fecha, lectura, asignaciones_maestros
                )
            self.bimestre_data.append(datos_semana)
        
        # Mostrar en pantalla
        self.mostrar_en_pantalla()
        
        messagebox.showinfo("Éxito", f"✅ Bimestre generado correctamente.\n📊 Total semanas: {len(self.bimestre_data)}")
        
    def cargar_estado_bimestre(self):
        """Carga el estado del bimestre guardado previamente y abre la ventana de edición"""
        ruta_estado = os.path.join(self.base_dir, "vmc_estado_bimestre.json")
        if not os.path.exists(ruta_estado):
            messagebox.showwarning("Advertencia", "No se encontró ningún bimestre guardado.")
            return
            
        try:
            with open(ruta_estado, 'r', encoding='utf-8') as f:
                self.bimestre_data = json.load(f)
            self.mostrar_en_pantalla()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el bimestre guardado: {e}")
    
    def _obtener_seccion_parte(self, parte):
        """Determina la sección de la reunión según el nombre de la parte"""
        parte_lower = parte.lower()
        if any(x in parte_lower for x in ['tesoro', 'perla', 'lectura', 'presidente', 'oración']):
            return "Tesoros de la Biblia"
        elif any(x in parte_lower for x in ['convers', 'revisita', 'discípulo', 'discipulo', 'creencia', 'maestro', 'ministerio']):
            return "Seamos Mejores Maestros"
        elif any(x in parte_lower for x in ['vida', 'estudio', 'lector', 'auxiliar', 'cristiana']):
            return "Nuestra Vida Cristiana"
        elif any(x in parte_lower for x in ['sonido', 'micrófono', 'plataforma', 'acomodador']):
            return "Servicios Mecánicos"
        return "Reunión VMC"

    def _obtener_telefono(self, nombre):
        """Busca el teléfono de un hermano en el DataFrame"""
        if self.df is None:
            return ""
        match = self.df[self.df['Nombre'] == nombre]
        if not match.empty:
            tel = str(match.iloc[0].get('Telefono', '')).strip()
            if tel and tel.lower() not in ['nan', 'none']:
                return tel
        return ""

    def _humanizar_mensaje(self, nombre, parte, fecha, seccion):
        """Genera un mensaje con saludos, despedidas y emojis dinámicamente variados para evitar que parezca de bot"""
        import random
        
        # Variantes de saludos cariñosos y teocráticos
        saludos = [
            "Hola {nombre}, 🙂",
            "Hola, hermano {nombre}. Espero que te encuentres muy bien. 🌟",
            "Saludos, {nombre} 👋",
            "¡Hola, {nombre}! Qué gusto saludarte. 😊",
            "Buen día, {nombre}. ✨",
            "Hola {nombre}, ¿cómo estás? Espero que todo esté excelente. 🤗",
            "Estimado {nombre}, un saludo muy cordial. 🌸",
            "Hola {nombre}. ¡Espero que tengas un feliz día! ☀️",
            "Saludos cariñosos, {nombre}. ❤️"
        ]
        
        # Variantes de introducciones para las asignaciones
        introducciones = [
            "Te informamos que tienes la siguiente asignación para la reunión:",
            "Te escribo para comentarte que tienes la siguiente asignación:",
            "Paso por aquí a recordarte tu asignación programada:",
            "Te comparto los detalles de tu asignación para esta semana:",
            "Aquí tienes la información sobre tu parte en la reunión:",
            "Queremos notificarte de tu participación para la reunión de esta semana:",
            "Te dejamos por acá los detalles de tu asignación VMC:"
        ]
        
        # Variantes de agradecimientos/despedidas
        despedidas = [
            "¡Muchas gracias por tu servicio! 🙏",
            "¡Muchas gracias por tu excelente disposición de siempre! 🙏",
            "Que Jehová bendiga mucho tu buena actitud y tus esfuerzos. ✨",
            "Agradecemos de corazón tu valioso apoyo. ¡Un gran abrazo! 🤗",
            "Gracias por tu buena disposición de siempre. 👍",
            "¡Que tengas una bendecida y excelente semana! Saludos. 😊",
            "Agradecemos mucho tu valioso servicio a favor de la congregación. 🤝",
            "Que Jehová te acompañe en tu preparación. ¡Saludos! 📖"
        ]
        
        # Elegir de forma aleatoria
        saludo = random.choice(saludos).format(nombre=nombre)
        intro = random.choice(introducciones)
        despedida = random.choice(despedidas)
        
        # Diseños visuales alternativos
        disenos = [
            f"{saludo}\n\n{intro}\n\n📋 *{parte}*\n📅 *{fecha}*\n📖 Sección: {seccion}\n\n{despedida}",
            f"{saludo}\n\n{intro}\n\n👉 *{parte}*\n📅 _Fecha: {fecha}_\n📌 Sección: {seccion}\n\n{despedida}",
            f"{saludo}\n\n{intro}\n\n✨ Asignación: *{parte}*\n📅 Semana: *{fecha}*\n📖 Sección: *{seccion}*\n\n{despedida}"
        ]
        
        return random.choice(disenos)

    def _construir_url_whatsapp(self, telefono, mensaje):
        """Construye la URL de WhatsApp según el método seleccionado"""
        import urllib.parse
        tel_clean = "".join(filter(str.isdigit, telefono))
        texto_encoded = urllib.parse.quote(mensaje)
        
        metodo = self.ws_metodo.get() if hasattr(self, 'ws_metodo') else "WhatsApp Desktop (App)"
        
        if "Desktop" in metodo:
            # Protocolo whatsapp:// abre la app de escritorio directamente
            return f"whatsapp://send?phone={tel_clean}&text={texto_encoded}"
        else:
            # WhatsApp Web en navegador
            return f"https://web.whatsapp.com/send?phone={tel_clean}&text={texto_encoded}"

    def copiar_whatsapp(self, nombre, parte, fecha):
        """Genera el texto de WhatsApp, lo copia y abre WhatsApp si hay teléfono"""
        try:
            import webbrowser
            
            plantilla = self.ws_template.get("0.0", "end").strip()
            
            # Manejar parejas (titular // ayudante o participante1 / participante2) - enviar al primero
            if " // " in nombre:
                solo_nombre = nombre.split(" // ")[0].strip()
            elif " / " in nombre:
                solo_nombre = nombre.split(" / ")[0].strip()
            elif "//" in nombre:
                solo_nombre = nombre.split("//")[0].strip()
            elif "/" in nombre:
                solo_nombre = nombre.split("/")[0].strip()
            else:
                solo_nombre = nombre.strip()
            seccion = self._obtener_seccion_parte(parte)
            
            # Buscar teléfono
            telefono = self._obtener_telefono(solo_nombre)
            
            # Generar mensaje (humanizado o con plantilla)
            if hasattr(self, 'ws_humanizar_var') and self.ws_humanizar_var.get():
                mensaje = self._humanizar_mensaje(solo_nombre, parte, fecha, seccion)
            else:
                mensaje = plantilla.format(
                    nombre=solo_nombre, 
                    parte=parte, 
                    fecha=fecha,
                    seccion=seccion
                )
            
            # Copiar al portapapeles siempre
            self.root.clipboard_clear()
            self.root.clipboard_append(mensaje)
            self.root.update()
            
            if telefono:
                url = self._construir_url_whatsapp(telefono, mensaje)
                webbrowser.open(url)
                messagebox.showinfo("WhatsApp", f"✅ Abriendo chat de {solo_nombre}...\n(Mensaje copiado al portapapeles)")
            else:
                messagebox.showinfo("WhatsApp", f"📋 Mensaje para {solo_nombre} copiado al portapapeles.\n⚠️ No se encontró teléfono para envío directo.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al enviar por WhatsApp: {e}")

    def _recopilar_asignaciones_whatsapp(self):
        """Recopila todas las asignaciones del bimestre para envío masivo de WhatsApp"""
        envios = []
        if not self.bimestre_data:
            return envios
        
        for semana in self.bimestre_data:
            if semana.get('no_reunion', False):
                continue
                
            fecha = semana['fecha']
            
            # Mapa de claves -> nombre de la parte para el mensaje
            partes_simples = [
                ('presidente', 'Presidencia'),
                ('oracion', 'Oración'),
                ('num1_tesoros', 'Tesoro 1 (Discurso)'),
                ('num2_tesoros', 'Tesoro 2 (Perlas)'),
                ('lectura_biblia', 'Lectura de la Biblia'),
                ('estudio_biblico', 'Estudio Bíblico (Conductor)'),
                ('lector', 'Estudio Bíblico (Lector)'),
                ('presidencia_aux', 'Presidencia auxiliar'),
                ('plataforma', 'Plataforma'),
            ]
            
            for clave, nombre_parte in partes_simples:
                persona = semana.get(clave, '')
                if persona and persona != '__________________':
                    for p in persona.split('//'):
                        p = p.strip()
                        if p and p != '__________________':
                            envios.append({'nombre': p, 'parte': nombre_parte, 'fecha': fecha})
            
            # Partes de Vida Cristiana (dinámicas, pueden ser múltiples)
            for pv_tipo, pv_val in semana.get('partes_vida', []):
                if pv_val and pv_val != '__________________':
                    envios.append({'nombre': pv_val, 'parte': f'Vida Cristiana: {pv_tipo}', 'fecha': fecha})
            
            # Maestros (pueden tener titular // ayudante)
            for tipo, val in semana.get('maestros', []):
                if val and val != '__________________':
                    partes = val.split('//')
                    for idx_p, p in enumerate(partes):
                        p = p.strip()
                        if p and p != '__________________':
                            rol = "(Encargado)" if idx_p == 0 and len(partes) > 1 else "(Ayudante)" if idx_p == 1 else ""
                            envios.append({'nombre': p, 'parte': f"{tipo} {rol}".strip(), 'fecha': fecha})
            
            # Servicios con múltiples asignados (formato "nombre1 / nombre2")
            for clave, nombre_parte in [('sonido', 'Audio y Video'), ('microfonos', 'Micrófonos'), ('acomodadores', 'Acomodador')]:
                val = semana.get(clave, '')
                if val:
                    for p in val.split('/'):
                        p = p.strip()
                        if p and p != '__________________':
                            envios.append({'nombre': p, 'parte': nombre_parte, 'fecha': fecha})
        
        return envios

    def enviar_todos_whatsapp(self):
        """Envía WhatsApp a todos los asignados del bimestre con intervalo configurable"""
        envios = self._recopilar_asignaciones_whatsapp()
        
        if not envios:
            messagebox.showwarning("Sin datos", "No hay asignaciones para enviar.")
            return
        
        # Agrupar envíos por persona
        envios_agrupados = {}
        for e in envios:
            nombre = e['nombre']
            if nombre not in envios_agrupados:
                envios_agrupados[nombre] = []
            envios_agrupados[nombre].append({'parte': e['parte'], 'fecha': e['fecha']})
            
        # Filtrar solo los que tienen teléfono y armar datos
        envios_con_tel = []
        envios_sin_tel = []
        for nombre, asignaciones in envios_agrupados.items():
            tel = self._obtener_telefono(nombre)
            
            # Construir texto de las partes
            texto_partes = []
            for a in asignaciones:
                texto_partes.append(f"• {a['parte']} ({a['fecha']})")
            texto_unido = "\n".join(texto_partes)
            
            parte_resumen = f"{len(asignaciones)} asignaciones" if len(asignaciones) > 1 else asignaciones[0]['parte']
            
            datos = {
                'nombre': nombre,
                'parte_resumen': parte_resumen,
                'asignaciones': asignaciones,
                'texto_partes': texto_unido,
                'cantidad': len(asignaciones)
            }
            
            if tel:
                datos['telefono'] = tel
                envios_con_tel.append(datos)
            else:
                envios_sin_tel.append(datos)
        
        # Mostrar ventana de confirmación con resumen
        modal = ctk.CTkToplevel(self.root)
        modal.title("📱 Envío Masivo de WhatsApp")
        modal.geometry("650x600")
        modal.after(100, lambda: modal.focus_force())
        modal.grab_set()
        
        ctk.CTkLabel(modal, text="📱 ENVÍO MASIVO DE WHATSAPP", font=("Arial", 18, "bold"), text_color="#25D366").pack(pady=10)
        
        # Resumen
        resumen_frame = ctk.CTkFrame(modal, border_width=1, border_color="#25D366")
        resumen_frame.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(resumen_frame, text=f"✅ Con teléfono: {len(envios_con_tel)} hermanos", 
                     font=("Arial", 12), text_color="#25D366").pack(anchor="w", padx=15, pady=2)
        ctk.CTkLabel(resumen_frame, text=f"⚠️ Sin teléfono: {len(envios_sin_tel)} (se omitirán)", 
                     font=("Arial", 12), text_color="orange").pack(anchor="w", padx=15, pady=2)
        
        try:
            delay_val = int(self.ws_delay.get())
        except:
            delay_val = 3
        ctk.CTkLabel(resumen_frame, text=f"⏱️ Intervalo: {delay_val} segundos entre cada mensaje", 
                     font=("Arial", 12)).pack(anchor="w", padx=15, pady=2)
        
        # Lista detallada
        scroll = ctk.CTkScrollableFrame(modal, height=300)
        scroll.pack(fill="both", expand=True, padx=20, pady=5)
        
        for i, e in enumerate(envios_con_tel):
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=f"{i+1}.", width=30).pack(side="left")
            ctk.CTkLabel(row, text=e['nombre'], width=180, anchor="w", font=("Arial", 11, "bold")).pack(side="left")
            ctk.CTkLabel(row, text=e['parte_resumen'], width=250, anchor="w", font=("Arial", 10)).pack(side="left")
        
        # Progreso
        progress_frame = ctk.CTkFrame(modal, fg_color="transparent")
        progress_frame.pack(fill="x", padx=20, pady=5)
        
        progress_bar = ctk.CTkProgressBar(progress_frame, height=15, progress_color="#25D366")
        progress_bar.pack(fill="x", pady=2)
        progress_bar.set(0)
        
        status_label = ctk.CTkLabel(progress_frame, text="Listo para enviar", font=("Arial", 11))
        status_label.pack(pady=2)
        
        # Botones
        btn_frame = ctk.CTkFrame(modal, fg_color="transparent")
        btn_frame.pack(pady=10)
        
        cancelado = {'valor': False}
        
        def iniciar_envio():
            import webbrowser
            import time
            import random
            
            btn_enviar.configure(state="disabled")
            btn_cancelar.configure(state="normal", fg_color="#C62828")
            plantilla = self.ws_template.get("0.0", "end").strip()
            
            try:
                delay = int(self.ws_delay.get())
            except:
                delay = 3
            
            def enviar_siguiente(idx):
                if cancelado['valor'] or idx >= len(envios_con_tel):
                    if cancelado['valor']:
                        status_label.configure(text=f"❌ Cancelado. Se enviaron {idx} de {len(envios_con_tel)}", text_color="red")
                    else:
                        status_label.configure(text=f"✅ ¡Completado! {len(envios_con_tel)} mensajes enviados", text_color="#25D366")
                        progress_bar.set(1)
                    btn_enviar.configure(state="normal")
                    btn_cancelar.configure(state="disabled")
                    return
                
                e = envios_con_tel[idx]
                
                try:
                    if hasattr(self, 'ws_humanizar_var') and self.ws_humanizar_var.get():
                        saludos = [
                            f"Hola {e['nombre']}, 🙂",
                            f"Hola, hermano {e['nombre']}. Espero que te encuentres muy bien. 🌟",
                            f"Saludos, {e['nombre']} 👋",
                            f"¡Hola, {e['nombre']}! Qué gusto saludarte. 😊",
                            f"Buen día, {e['nombre']}. ✨",
                        ]
                        intros = [
                            "Te informamos que tienes las siguientes asignaciones programadas:",
                            "Te escribo para comentarte tus participaciones para este bimestre:",
                            "Paso por aquí a recordarte tus asignaciones:",
                            "Te comparto los detalles de tus partes en la reunión:",
                        ]
                        despedidas = [
                            "¡Muchas gracias por tu excelente disposición de siempre! 🙏",
                            "Que Jehová bendiga mucho tu buena actitud y tus esfuerzos. ✨",
                            "Agradecemos de corazón tu valioso apoyo. ¡Un gran abrazo! 🤗",
                            "¡Que tengas una bendecida y excelente semana! Saludos. 😊",
                        ]
                        saludo = random.choice(saludos)
                        intro = random.choice(intros)
                        despedida = random.choice(despedidas)
                        mensaje = f"{saludo}\n\n{intro}\n\n{e['texto_partes']}\n\n{despedida}"
                    else:
                        mensaje = plantilla.format(
                            nombre=e['nombre'],
                            parte=f"\n{e['texto_partes']}",
                            fecha="este bimestre",
                            seccion="Varias" if len(e['asignaciones']) > 1 else self._obtener_seccion_parte(e['asignaciones'][0]['parte'])
                        )
                except KeyError:
                    mensaje = f"Hola {e['nombre']}, tienes las siguientes asignaciones:\n{e['texto_partes']}"
                
                url = self._construir_url_whatsapp(e['telefono'], mensaje)
                webbrowser.open(url)
                
                progreso = (idx + 1) / len(envios_con_tel)
                progress_bar.set(progreso)
                status_label.configure(text=f"📱 Enviando {idx+1}/{len(envios_con_tel)}: {e['nombre']}")
                
                # Programar siguiente envío con delay
                modal.after(delay * 1000, lambda: enviar_siguiente(idx + 1))
            
            enviar_siguiente(0)
        
        def cancelar_envio():
            cancelado['valor'] = True
        
        btn_enviar = ctk.CTkButton(btn_frame, text="🚀 INICIAR ENVÍO", fg_color="#25D366", hover_color="#128C7E",
                                    font=("Arial", 14, "bold"), width=200, height=40, command=iniciar_envio)
        btn_enviar.pack(side="left", padx=10)
        
        btn_cancelar = ctk.CTkButton(btn_frame, text="⛔ CANCELAR", fg_color="gray", state="disabled",
                                      width=120, height=40, command=cancelar_envio)
        btn_cancelar.pack(side="left", padx=10)
        
        ctk.CTkButton(btn_frame, text="CERRAR", fg_color="#666", width=100, height=40,
                      command=modal.destroy).pack(side="left", padx=10)

    def mostrar_en_pantalla(self):
        """Muestra el bimestre en una ventana interactiva donde se pueden editar los nombres"""
        if not self.bimestre_data:
            return
        
        ventana = ctk.CTkToplevel(self.root)
        ventana.title("Edición Final del Bimestre - El Araguaney")
        ventana.geometry("1200x800")
        ventana.after(100, lambda: ventana.focus_force())
        
        # Panel superior con controles
        top_panel = ctk.CTkFrame(ventana, height=50)
        top_panel.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(top_panel, text="✍️ Puedes editar los nombres directamente antes de exportar", 
                     font=("Arial", 12, "italic")).pack(side="left", padx=10)
        
        def guardar_cambios():
            # Los cambios ya se guardan en el diccionario mediante los callbacks de los combos
            self.guardar_historial()
            # Guardar estado del bimestre persistente
            try:
                ruta_estado = os.path.join(self.base_dir, "vmc_estado_bimestre.json")
                with open(ruta_estado, 'w', encoding='utf-8') as f:
                    json.dump(self.bimestre_data, f, ensure_ascii=False, indent=4)
            except Exception as e:
                print(f"Error al guardar estado: {e}")
            
            # Guardar datos combinados para la app móvil (sin dependencias complejas)
            try:
                ruta_movil = os.path.join(self.base_dir, "vmc_datos_movil.json")
                datos_movil = {
                    "bimestre": self.bimestre_data,
                    "congregacion": self.df[['Nombre', 'Privilegio', 'Genero', 'Telefono']].to_dict(orient='records') if self.df is not None else []
                }
                with open(ruta_movil, 'w', encoding='utf-8') as f:
                    json.dump(datos_movil, f, ensure_ascii=False, indent=4)
            except Exception as e:
                print(f"Error al guardar datos móvil: {e}")

            messagebox.showinfo("Éxito", "Cambios guardados localmente (y datos sincronizados para móvil). Ahora puede exportar a Excel o PDF.")
            ventana.destroy()
            
        ctk.CTkButton(top_panel, text="💾 CONFIRMAR CAMBIOS", command=guardar_cambios,
                     fg_color="#2E7D32", hover_color="#1B5E20").pack(side="right", padx=10)
        
        ctk.CTkButton(top_panel, text="📱 ENVIAR TODOS POR WHATSAPP", command=self.enviar_todos_whatsapp,
                     fg_color="#25D366", hover_color="#128C7E", font=("Arial", 12, "bold")).pack(side="right", padx=10)
        
        ctk.CTkButton(top_panel, text="📄 GENERAR S-89", command=self.exportar_s89_pdf,
                     fg_color="#388E3C", hover_color="#2E7D32").pack(side="right", padx=10)

        scroll_frame = ctk.CTkScrollableFrame(ventana, width=1150, height=700)
        scroll_frame.pack(padx=10, pady=5, fill="both", expand=True)
        
        # Listas genéricas base
        lista_v = sorted(list(set(self.todos_varones)))
        lista_h = sorted(list(set(self.hermanas)))
        lista_t = sorted(list(set(self.todos_varones + self.hermanas)))
        lista_anc = sorted(list(set(self.ancianos)))
        lista_anc_sie = sorted(list(set(self.ancianos + self.siervos)))
        lista_presidir = sorted(list(set(self.list_puede_presidir))) if hasattr(self, 'list_puede_presidir') and self.list_puede_presidir else lista_anc
        
        # Listas granulares específicas de habilidades
        lista_oracion = sorted(list(set(self.list_oracion))) if hasattr(self, 'list_oracion') and self.list_oracion else lista_anc_sie
        lista_tes_1 = sorted(list(set(self.list_tes_discurso))) if hasattr(self, 'list_tes_discurso') and self.list_tes_discurso else lista_anc_sie
        lista_tes_2 = sorted(list(set(self.list_tes_perlas))) if hasattr(self, 'list_tes_perlas') and self.list_tes_perlas else lista_anc_sie
        lista_lectura = sorted(list(set(self.list_lectores_biblia))) if hasattr(self, 'list_lectores_biblia') and self.list_lectores_biblia else lista_v
        
        lista_vida_1 = sorted(list(set(self.list_vida_p1))) if hasattr(self, 'list_vida_p1') and self.list_vida_p1 else lista_anc_sie
        lista_vida_2 = sorted(list(set(self.list_vida_p2))) if hasattr(self, 'list_vida_p2') and self.list_vida_p2 else lista_anc_sie
        lista_est_cond = sorted(list(set(self.list_estudio_conductor))) if hasattr(self, 'list_estudio_conductor') and self.list_estudio_conductor else lista_anc
        lista_est_lect = sorted(list(set(self.list_estudio_lector))) if hasattr(self, 'list_estudio_lector') and self.list_estudio_lector else lista_v
        
        lista_sonido = sorted(list(set(self.list_serv_sonido))) if hasattr(self, 'list_serv_sonido') and self.list_serv_sonido else []
        lista_plat = sorted(list(set(self.list_serv_plataforma))) if hasattr(self, 'list_serv_plataforma') and self.list_serv_plataforma else []
        lista_mics = sorted(list(set(self.list_serv_mics))) if hasattr(self, 'list_serv_mics') and self.list_serv_mics else []
        lista_acom = sorted(list(set(self.list_serv_acomodador))) if hasattr(self, 'list_serv_acomodador') and self.list_serv_acomodador else []
        
        def create_editable_label(parent, label_text, current_val, key_in_dict, week_dict, pool, is_maestro=False, maestro_idx=None):
            frame = ctk.CTkFrame(parent, fg_color="transparent")
            frame.pack(anchor="w", fill="x", padx=5, pady=1)
            
            ctk.CTkLabel(frame, text=label_text, font=("Arial", 10, "bold"), width=100, anchor="w").pack(side="left")
            
            # Botón de WhatsApp al lado de cada asignación
            btn_ws = ctk.CTkButton(frame, text="📱", width=25, height=20, fg_color="#25D366", hover_color="#128C7E",
                                  command=lambda n=current_val, p=label_text, f=week_dict['fecha']: self.copiar_whatsapp(n, p, f))
            btn_ws.pack(side="right", padx=2)

            # Si es pareja (titular // ayudante o participante1 / participante2), separar para editar
            sep = None
            if " // " in str(current_val):
                sep = " // "
            elif " / " in str(current_val):
                sep = " / "
            elif "//" in str(current_val):
                sep = "//"
            elif "/" in str(current_val):
                sep = "/"
            
            # Si no hay separador pero es una parte que puede ser pareja (por ejemplo "Explique sus creencias")
            if sep is None and is_maestro and any(x in label_text.lower() for x in ["creencia", "convers", "revis", "discip", "escenificaci"]):
                sep = " // "
                current_val = f"{current_val}{sep}__________________"

            # Asegurar que la lista de opciones tenga "__________________" para poder vaciar
            pool_con_vacio = list(pool)
            if "__________________" not in pool_con_vacio:
                pool_con_vacio.insert(0, "__________________")

            if sep is not None:
                t, a = current_val.split(sep, 1)
                t, a = t.strip(), a.strip()
                
                combo_t = ctk.CTkComboBox(frame, values=pool_con_vacio, width=120, height=24, font=("Arial", 10))
                combo_t.set(t)
                combo_t.pack(side="left", padx=2, fill="x", expand=True)
                
                ctk.CTkLabel(frame, text=sep).pack(side="left")
                
                combo_a = ctk.CTkComboBox(frame, values=pool_con_vacio, width=120, height=24, font=("Arial", 10))
                combo_a.set(a)
                combo_a.pack(side="left", padx=2, fill="x", expand=True)
                
                def update_couple(*args, w=week_dict, k=key_in_dict, idx=maestro_idx, ct=combo_t, ca=combo_a, sp=sep):
                    new_val = f"{ct.get()}{sp}{ca.get()}"
                    if idx is not None: w[k][idx] = (w[k][idx][0], new_val)
                    else: w[k] = new_val
                
                combo_t.configure(command=update_couple)
                combo_a.configure(command=update_couple)
            else:
                combo = ctk.CTkComboBox(frame, values=pool_con_vacio, width=180, height=24, font=("Arial", 10))
                combo.set(str(current_val))
                combo.pack(side="left", padx=2)
                
                def update_single(val, w=week_dict, k=key_in_dict, idx=maestro_idx):
                    if idx is not None: w[k][idx] = (w[k][idx][0], val)
                    else: w[k] = val
                
                combo.configure(command=update_single)

        for week_idx, semana in enumerate(self.bimestre_data):
            week_frame = ctk.CTkFrame(scroll_frame, corner_radius=5, border_width=1, border_color="#CCC")
            week_frame.pack(padx=10, pady=10, fill="x")
            
            # Row 1: Titulo y Fecha
            header_frame = ctk.CTkFrame(week_frame, fg_color="transparent", height=40)
            header_frame.pack(fill="x", padx=1, pady=1)
            
            ctk.CTkLabel(header_frame, text=f"📅 {semana['fecha'].upper()}  |  {semana['lectura']}", 
                         font=("Arial", 14, "bold")).pack(side="left", padx=15)
            
            # Row 2: Cuerpo (3 columnas)
            cols_frame = ctk.CTkFrame(week_frame, fg_color="transparent")
            cols_frame.pack(fill="x", pady=5)
            
            if semana.get('no_reunion', False):
                motivo = semana.get('motivo', 'Semana cancelada')
                ctk.CTkLabel(cols_frame, text=f"🚫 {motivo.upper()}", 
                             font=("Arial", 16, "bold"), text_color="#D32F2F", pady=20).pack(fill="x")
                continue # Saltar al siguiente week_frame
            
            cols_frame.grid_columnconfigure(0, weight=1, minsize=280)
            cols_frame.grid_columnconfigure(1, weight=2, minsize=380)
            cols_frame.grid_columnconfigure(2, weight=2, minsize=340)
            
            # Col 1: Tesoros y General
            col1 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col1.grid(row=0, column=0, sticky="nsew", padx=5)
            ctk.CTkLabel(col1, text="💎 TESOROS", fg_color="#808080", text_color="white", font=("Arial", 11, "bold")).pack(fill="x", pady=2)
            
            create_editable_label(col1, "Presidente:", semana['presidente'], 'presidente', semana, lista_presidir)
            create_editable_label(col1, "Oración:", semana['oracion'], 'oracion', semana, lista_oracion)
            create_editable_label(col1, "Tesoro 1:", semana['num1_tesoros'], 'num1_tesoros', semana, lista_tes_1)
            create_editable_label(col1, "Tesoro 2:", semana['num2_tesoros'], 'num2_tesoros', semana, lista_tes_2)
            create_editable_label(col1, "Lectura Biblia:", semana['lectura_biblia'], 'lectura_biblia', semana, lista_lectura)
            
            # Col 2: Maestros
            col2 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col2.grid(row=0, column=1, sticky="nsew", padx=5)
            ctk.CTkLabel(col2, text="🤝 SEAMOS MEJORES MAESTROS", fg_color="#FF8C00", text_color="white", font=("Arial", 11, "bold")).pack(fill="x", pady=2)
            
            for i, (tipo, val) in enumerate(semana['maestros']):
                pool = lista_t if ("//" in str(val) or any(x in tipo.lower() for x in ["convers", "revis", "discip", "creencia"])) else lista_v
                create_editable_label(col2, f"{tipo}:", val, 'maestros', semana, pool, is_maestro=True, maestro_idx=i)
            
            # Col 3: Vida Cristiana
            col3 = ctk.CTkFrame(cols_frame, fg_color="transparent")
            col3.grid(row=0, column=2, sticky="nsew", padx=5)
            ctk.CTkLabel(col3, text="🏠 VIDA CRISTIANA", fg_color="#B22222", text_color="white", font=("Arial", 11, "bold")).pack(fill="x", pady=2)
            
            for pv_idx, (pv_tipo, pv_val) in enumerate(semana.get('partes_vida', [('Parte 1', '__________________')])):
                if 'necesidades' in pv_tipo.lower():
                    pool_pv = lista_anc
                elif pv_idx == 0:
                    pool_pv = lista_vida_1
                else:
                    pool_pv = lista_vida_2
                create_editable_label(col3, f"{pv_tipo}:", pv_val, 'partes_vida', semana, pool_pv, is_maestro=True, maestro_idx=pv_idx)
            create_editable_label(col3, "Estudio Bíblico:", semana['estudio_biblico'], 'estudio_biblico', semana, lista_est_cond)
            create_editable_label(col3, "Lector:", semana['lector'], 'lector', semana, lista_est_lect)
            create_editable_label(col3, "Presidencia auxiliar:", semana['presidencia_aux'], 'presidencia_aux', semana, lista_presidir)
            
            # Fila de servicios (Editable también)
            serv_frame = ctk.CTkFrame(week_frame, fg_color="transparent")
            serv_frame.pack(fill="x", pady=5, padx=5)
            
            s_cols = ctk.CTkFrame(serv_frame, fg_color="transparent")
            s_cols.pack(fill="x")
            
            # Dividir servicios en 2 filas pequeñas
            f1 = ctk.CTkFrame(s_cols, fg_color="transparent")
            f1.pack(fill="x")
            create_editable_label(f1, "Sonido y Video:", semana['sonido'], 'sonido', semana, lista_sonido)
            create_editable_label(f1, "Plataforma:", semana['plataforma'], 'plataforma', semana, lista_plat)
            
            f2 = ctk.CTkFrame(s_cols, fg_color="transparent")
            f2.pack(fill="x")
            create_editable_label(f2, "Micrófonos:", semana['microfonos'], 'microfonos', semana, lista_mics)
            create_editable_label(f2, "Acomodadores:", semana['acomodadores'], 'acomodadores', semana, lista_acom)
    
    def exportar_excel(self):
        """Exporta el bimestre a Excel en una sola hoja con formato avanzado"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos. Genere el bimestre primero.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"bimestre_vmc_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not filename:
            return
            
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bimestre"
        
        # Eliminar líneas de cuadrícula para un look más limpio
        ws.sheet_view.showGridLines = False
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 52
        ws.column_dimensions['C'].width = 42
        
        # Definir estilos
        font_title = Font(name='Segoe UI', size=16, bold=True, color='333333')
        font_date = Font(name='Segoe UI', size=12, bold=True, color='5E005E')
        font_header_blanco = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        font_normal = Font(name='Segoe UI', size=10, color='444444')
        font_cancel = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
        
        fill_purple = PatternFill(start_color='5E005E', end_color='5E005E', fill_type='solid') # Línea divisoria
        fill_tesoros = PatternFill(start_color='607D8B', end_color='607D8B', fill_type='solid') # Azul Grisaceo
        fill_maestros = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid') # Naranja Premium
        fill_vida = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid') # Rojo oscuro
        fill_cancel = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid') # Rojo cancelado
        
        # Cebras
        fill_light_grey = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        border_soft = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                             top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
                             
        current_row = 1
        
        for semana in self.bimestre_data:
            # Fila 1: Título y Fecha
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell_title = ws.cell(row=current_row, column=1, value=f"VIDA Y MINISTERIO CRISTIANOS  |  {semana['lectura'].upper()}")
            cell_title.font = font_title
            cell_title.alignment = Alignment(horizontal='center', vertical='center')
            
            cell_date = ws.cell(row=current_row, column=3, value=semana['fecha'])
            cell_date.font = font_date
            cell_date.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[current_row].height = 25
            
            current_row += 1
            
            # Fila 2: Franja morada
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_purple
            ws.row_dimensions[current_row].height = 5
            current_row += 1
            
            if semana.get('no_reunion', False):
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                cell = ws.cell(row=current_row, column=1, value=f"🚫 {semana.get('motivo', 'SEMANA CANCELADA').upper()}")
                cell.font = font_cancel
                cell.fill = fill_cancel
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[current_row].height = 40
                current_row += 2 # Espacio extra
                continue
            
            # Fila 3: Presidente y Oración
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell_pres = ws.cell(row=current_row, column=1, value=f"PRESIDENTE DE LA REUNIÓN: {semana['presidente']}")
            cell_pres.font = font_normal
            cell_pres.border = border_soft
            
            cell_ora = ws.cell(row=current_row, column=3, value=f"ORACIÓN: {semana['oracion']}")
            cell_ora.font = font_normal
            cell_ora.border = border_soft
            ws.row_dimensions[current_row].height = 20
            
            current_row += 1
            
            # Fila 4: Encabezados de secciones
            headers = [
                ("TESOROS DE LA BIBLIA", fill_tesoros),
                ("SEAMOS MEJORES MAESTROS", fill_maestros),
                ("NUESTRA VIDA CRISTIANA", fill_vida)
            ]
            for col, (text, fill) in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=text)
                cell.font = font_header_blanco
                cell.fill = fill
                cell.border = border_soft
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[current_row].height = 22
            
            current_row += 1
            
            # Preparar datos de las columnas
            col_tesoros = [
                f"NUM 1: {semana['num1_tesoros']}",
                f"NUM 2: {semana['num2_tesoros']}",
                f"NUM 3: {semana['lectura_biblia']}"
            ]
            
            col_maestros = []
            num_actual = 4
            for tipo, val in semana['maestros']:
                col_maestros.append(f"NUM {num_actual}: {val}")
                num_actual += 1
                
            col_vida = []
            for pv_tipo, pv_val in semana.get('partes_vida', []):
                col_vida.append(f"NUM {num_actual}: [{pv_tipo}] {pv_val}")
                num_actual += 1
            col_vida.append(f"NUM {num_actual}: (Estudio) {semana['estudio_biblico']}")
            col_vida.append(f"LECTOR ESTUDIO: {semana['lector']}")
            
            max_rows = max(len(col_tesoros), len(col_maestros), len(col_vida))
            
            # Filas de asignaciones
            for i in range(max_rows):
                fill_row = fill_light_grey if i % 2 == 0 else fill_white
                
                # Tesoros
                val_t = col_tesoros[i] if i < len(col_tesoros) else ""
                cell_t = ws.cell(row=current_row, column=1, value=val_t)
                cell_t.font = font_normal
                cell_t.border = border_soft
                cell_t.fill = fill_row
                
                # Maestros
                val_m = col_maestros[i] if i < len(col_maestros) else ""
                cell_m = ws.cell(row=current_row, column=2, value=val_m)
                cell_m.font = font_normal
                cell_m.border = border_soft
                cell_m.fill = fill_row
                
                # Vida
                val_v = col_vida[i] if i < len(col_vida) else ""
                cell_v = ws.cell(row=current_row, column=3, value=val_v)
                cell_v.font = font_normal
                cell_v.border = border_soft
                cell_v.fill = fill_row
                
                ws.row_dimensions[current_row].height = 18
                current_row += 1
                
            # Fila de servicios (Sonido, Plataforma, etc.)
            servicios_data = [
                (f"Sonido: {semana.get('sonido', '')}", f"Plataforma: {semana.get('plataforma', '')}", f"Acomodadores: {semana.get('acomodadores', '')}"),
                (f"Presidencia Aux: {semana.get('presidencia_aux', '')}", f"Micrófonos: {semana.get('microfonos', '')}", "")
            ]
            for serv_row in servicios_data:
                for col, text in enumerate(serv_row, start=1):
                    cell = ws.cell(row=current_row, column=col, value=text)
                    cell.font = Font(name='Arial', size=9)
                    cell.fill = fill_light_grey
                    cell.border = border_soft
                current_row += 1
                
            # Espacio extra entre semanas
            current_row += 1
            
        try:
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Excel exportado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar Excel:\n{str(e)}")
    
    def exportar_pdf(self):
        """Exporta el bimestre a PDF con formato idéntico al de Excel (3 columnas)"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos. Genere el bimestre primero.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"bimestre_vmc_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        
        if not filename:
            return
            
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(filename, pagesize=landscape(letter),
                                rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        for semana in self.bimestre_data:
            data = []
            
            # Row 1: Título y Fecha/Lectura
            data.append([f"VIDA Y MINISTERIO CRISTIANOS", "", f"{semana['fecha']}  |  {semana['lectura']}"])
            
            # Row 2: Franja morada
            data.append(["", "", ""])
            
            if semana.get('no_reunion', False):
                motivo = semana.get('motivo', 'SEMANA CANCELADA').upper()
                data.append([f"🚫 {motivo}", "", ""])
                
                col_widths = [3.5 * inch, 3.5 * inch, 3.5 * inch]
                table = Table(data, colWidths=col_widths)
                style = TableStyle([
                    ('SPAN', (0, 0), (1, 0)),
                    ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (1, 0), 16),
                    ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (2, 0), (2, 0), 10),
                    ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                    ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    
                    ('SPAN', (0, 1), (2, 1)),
                    ('BACKGROUND', (0, 1), (2, 1), colors.HexColor('#5E005E')),
                    
                    ('SPAN', (0, 2), (2, 2)),
                    ('BACKGROUND', (0, 2), (2, 2), colors.HexColor('#D32F2F')),
                    ('TEXTCOLOR', (0, 2), (2, 2), colors.whitesmoke),
                    ('FONTNAME', (0, 2), (2, 2), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 2), (2, 2), 16),
                    ('ALIGN', (0, 2), (2, 2), 'CENTER'),
                    ('VALIGN', (0, 2), (2, 2), 'MIDDLE'),
                    ('TOPPADDING', (0, 2), (2, 2), 20),
                    ('BOTTOMPADDING', (0, 2), (2, 2), 20),
                    
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#DDDDDD')),
                ])
                table.setStyle(style)
                story.append(table)
                story.append(Spacer(1, 0.4 * inch))
                continue
            
            # Row 3: Presidente y Oración
            data.append([f"PRESIDENTE DE LA REUNIÓN: {semana['presidente']}", "", f"ORACIÓN: {semana['oracion']}"])
            
            # Row 4: Encabezados
            data.append(["TESOROS DE LA BIBLIA", "SEAMOS MEJORES MAESTROS", "NUESTRA VIDA CRISTIANA"])
            
            # Preparar columnas
            col_tesoros = [
                f"NUM 1: {semana['num1_tesoros']}",
                f"NUM 2: {semana['num2_tesoros']}",
                f"NUM 3: {semana['lectura_biblia']}"
            ]
            
            col_maestros = []
            num_actual = 4
            for tipo, val in semana['maestros']:
                col_maestros.append(f"NUM {num_actual}: {val}")
                num_actual += 1
                
            col_vida = []
            for pv_tipo, pv_val in semana.get('partes_vida', []):
                col_vida.append(f"NUM {num_actual}: [{pv_tipo}] {pv_val}")
                num_actual += 1
            col_vida.append(f"NUM {num_actual}: (Estudio) {semana['estudio_biblico']}")
            col_vida.append(f"LECTOR ESTUDIO: {semana['lector']}")
            
            max_rows = max(len(col_tesoros), len(col_maestros), len(col_vida))
            
            for i in range(max_rows):
                val_t = col_tesoros[i] if i < len(col_tesoros) else ""
                val_m = col_maestros[i] if i < len(col_maestros) else ""
                val_v = col_vida[i] if i < len(col_vida) else ""
                data.append([val_t, val_m, val_v])
                
            # Servicios
            data.append([f"Sonido: {semana.get('sonido','')}", f"Plataforma: {semana.get('plataforma','')}", f"Acomodadores: {semana.get('acomodadores','')}"])
            data.append([f"Presidencia Aux: {semana.get('presidencia_aux','')}", f"Micrófonos: {semana.get('microfonos','')}", ""])
            
            col_widths = [3.5 * inch, 3.5 * inch, 3.5 * inch]
            table = Table(data, colWidths=col_widths)
            
            style_cmds = [
                # Row 1 (Title)
                ('SPAN', (0, 0), (1, 0)),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (1, 0), 16),
                ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (2, 0), (2, 0), 10),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Row 2 (Purple separator)
                ('SPAN', (0, 1), (2, 1)),
                ('BACKGROUND', (0, 1), (2, 1), colors.HexColor('#5E005E')),
                
                # Row 3 (President / Prayer)
                ('SPAN', (0, 2), (1, 2)),
                ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
                ('FONTSIZE', (0, 2), (-1, 2), 10),
                ('BOTTOMPADDING', (0, 2), (-1, 2), 6),
                ('TOPPADDING', (0, 2), (-1, 2), 6),
                
                # Row 4 (Headers)
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 3), (-1, 3), 10),
                ('TEXTCOLOR', (0, 3), (-1, 3), colors.whitesmoke),
                ('BACKGROUND', (0, 3), (0, 3), colors.HexColor('#607D8B')), # Azul Grisáceo
                ('BACKGROUND', (1, 3), (1, 3), colors.HexColor('#E65100')), # Naranja Premium
                ('BACKGROUND', (2, 3), (2, 3), colors.HexColor('#B71C1C')), # Rojo Oscuro
                ('ALIGN', (0, 3), (-1, 3), 'CENTER'),
                
                # Content rows
                ('FONTNAME', (0, 4), (-1, -3), 'Helvetica'),
                ('FONTSIZE', (0, 4), (-1, -3), 9),
                ('VALIGN', (0, 4), (-1, -3), 'TOP'),
                ('TOPPADDING', (0, 4), (-1, -3), 6),
                ('BOTTOMPADDING', (0, 4), (-1, -3), 6),
                
                # Service rows (last 2 rows)
                ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#F5F5F5')),
                ('FONTNAME', (0, -2), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, -2), (-1, -1), 9),
                
                # Global Grid
                ('GRID', (0, 2), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ]
            
            # Cebra
            for idx_row in range(4, len(data)-2):
                if idx_row % 2 == 0:
                    style_cmds.append(('BACKGROUND', (0, idx_row), (-1, idx_row), colors.HexColor('#F9F9F9')))
            
            style = TableStyle(style_cmds)
            
            table.setStyle(style)
            story.append(table)
            story.append(Spacer(1, 0.4 * inch))
            
        try:
            doc.build(story)
            messagebox.showinfo("Éxito", f"PDF exportado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{e}")

    def exportar_pdf_bolsillo(self):
        """Genera un PDF con programas individuales pequeños (2 por página)"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos.")
            return
            
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=f"programas_bolsillo_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        if not filename: return
        
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        for i, semana in enumerate(self.bimestre_data):
            if semana.get('no_reunion', False):
                motivo = semana.get('motivo', 'SEMANA CANCELADA').upper()
                data = [
                    [f"REUNIÓN: {semana['fecha']}"],
                    [f"🚫 {motivo}"]
                ]
                t = Table(data, colWidths=[4*inch])
                t.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                    ('BACKGROUND', (0,0), (0,0), colors.HexColor('#5E005E')),
                    ('TEXTCOLOR', (0,0), (0,0), colors.white),
                    ('FONTNAME', (0,0), (0,0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0,1), (0,1), colors.HexColor('#D32F2F')),
                    ('TEXTCOLOR', (0,1), (0,1), colors.whitesmoke),
                    ('FONTNAME', (0,1), (0,1), 'Helvetica-Bold'),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 8)
                ]))
                story.append(t)
                story.append(Spacer(1, 0.5*inch))
                if i % 2 != 0: story.append(PageBreak())
                continue

            data = [
                [f"REUNIÓN: {semana['fecha']}"],
                [f"LECTURA: {semana['lectura']}"],
                [f"PRESIDENTE: {semana['presidente']}"],
                ["TESOROS:"],
                [f"1. {semana['num1_tesoros']}"],
                [f"2. {semana['num2_tesoros']}"],
                [f"3. {semana['lectura_biblia']}"],
                ["MINISTERIO:"],
            ]
            for t, v in semana['maestros']:
                data.append([f"- {t}: {v}"])
                
            data.append(["VIDA CRISTIANA:"])
            for pv_tipo, pv_val in semana.get('partes_vida', []):
                data.append([f"- {pv_tipo}: {pv_val}"])
            data.append([f"- Estudio: {semana['estudio_biblico']}"])
            data.append([f"- Lector: {semana['lector']}"])
            
            t = Table(data, colWidths=[4*inch])
            t.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('BACKGROUND', (0,0), (0,2), colors.lightgrey),
                ('FONTNAME', (0,0), (0,2), 'Helvetica-Bold'),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.5*inch))
            
            # Cada 2 programas, salto de página
            if (i + 1) % 2 == 0:
                story.append(PageBreak())
                
        doc.build(story)
        messagebox.showinfo("Éxito", "Programas de bolsillo generados.")

    def exportar_s89_pdf(self):
        """Exporta los formularios S-89 para todos los estudiantes en un solo PDF"""
        if not self.bimestre_data:
            messagebox.showwarning("Advertencia", "No hay datos. Genere el bimestre primero.")
            return
            
        try:
            filename = filedialog.asksaveasfilename(
                parent=self.root,
                title="Guardar Vales S-89",
                defaultextension=".pdf",
                filetypes=[("Documentos PDF", "*.pdf")],
                initialfile=f"S89_Bimestre_{datetime.now().strftime('%Y%m%d')}"
            )
            
            if not filename:
                return
        
            c = canvas.Canvas(filename, pagesize=letter)
            width, height = letter
            
            # Dimensiones del vale S-89 (aprox 10.5cm x 13.5cm para que quepan 4 en una carta)
            vale_w = 10.5 * cm
            vale_h = 13.5 * cm
            
            # Margen y espaciado para centrar 2x2 en la página
            margin_x = (width - 2 * vale_w) / 3
            margin_y = (height - 2 * vale_h) / 3
            
            vales_por_pagina = 4
            contador = 0
            
            for semana in self.bimestre_data:
                if semana.get('no_reunion', False):
                    continue
                    
                fecha = semana['fecha']
                
                # Lista de asignaciones S-89
                asignaciones_estudiante = []
                
                # 1. Lectura de la Biblia (Num 3)
                lectura_val = semana.get('lectura_biblia', '')
                if lectura_val and lectura_val != "__________________":
                    asignaciones_estudiante.append({
                        'nombre': lectura_val,
                        'ayudante': '',
                        'num': '3',
                        'tipo': 'Lectura de la Biblia'
                    })
                
                # 2. Maestros (Num 4+)
                for i, (tipo_orig, val) in enumerate(semana['maestros']):
                    num = str(i + 4) # Fallback
                    # Intentar extraer número del tipo si viene de JW.org (ej: "4. Empiece...")
                    m_num = re.match(r'(\d+)\.', tipo_orig)
                    if m_num:
                        num = m_num.group(1)
                        tipo = tipo_orig.split('.', 1)[1].strip()
                    else:
                        tipo = tipo_orig
                    
                    nombre = val
                    ayudante = ""
                    if "//" in val:
                        partes = val.split("//")
                        nombre = partes[0].strip()
                        ayudante = partes[1].strip()
                    
                    if nombre and nombre != "__________________":
                        asignaciones_estudiante.append({
                            'nombre': nombre,
                            'ayudante': ayudante,
                            'num': num,
                            'tipo': tipo
                        })
                
                # Dibujar cada vale encontrado para esta semana
                for asig in asignaciones_estudiante:
                    if contador >= vales_por_pagina:
                        c.showPage()
                        contador = 0
                    
                    # Calcular posición (2x2 grid)
                    col = contador % 2
                    row = 1 - (contador // 2) # 1 es arriba, 0 es abajo
                    
                    x_pos = margin_x + col * (vale_w + margin_x)
                    y_pos = margin_y + row * (vale_h + margin_y)
                    
                    self._dibujar_s89(c, x_pos, y_pos, vale_w, vale_h, asig, fecha)
                    contador += 1
                    
            c.save()
            messagebox.showinfo("Éxito", f"✅ Formularios S-89 exportados correctamente en:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF de los vales:\n{str(e)}")

    def _dibujar_s89(self, c, x, y, w, h, asig, fecha):
        """Dibuja un solo vale S-89 en las coordenadas dadas"""
        
        # Guardar estado
        c.saveState()
        
        # Marco del vale
        c.setLineWidth(0.8)
        c.setStrokeColor(colors.black)
        c.rect(x, y, w, h)
        
        # Título principal
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(x + w/2, y + h - 0.8*cm, "ASIGNACIÓN PARA LA REUNIÓN")
        c.drawCentredString(x + w/2, y + h - 1.3*cm, "VIDA Y MINISTERIO CRISTIANOS")
        
        # Etiquetas de campos
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 0.6*cm, y + h - 2.5*cm, "Nombre:")
        c.drawString(x + 0.6*cm, y + h - 3.5*cm, "Ayudante:")
        c.drawString(x + 0.6*cm, y + h - 4.5*cm, "Fecha:")
        c.drawString(x + 0.6*cm, y + h - 5.8*cm, "Intervención núm.:")
        
        # Datos de los campos
        c.setFont("Helvetica", 10)
        c.drawString(x + 2.3*cm, y + h - 2.5*cm, asig['nombre'])
        c.drawString(x + 2.3*cm, y + h - 3.5*cm, asig['ayudante'] or "---------------------------------------")
        c.drawString(x + 2.3*cm, y + h - 4.5*cm, fecha)
        c.drawString(x + 4.0*cm, y + h - 5.8*cm, asig['num'])
        
        # Líneas decorativas (punteadas) para los campos
        c.setDash(1, 2)
        c.setLineWidth(0.5)
        c.line(x + 2.2*cm, y + h - 2.6*cm, x + w - 0.6*cm, y + h - 2.6*cm)
        c.line(x + 2.2*cm, y + h - 3.6*cm, x + w - 0.6*cm, y + h - 3.6*cm)
        c.line(x + 2.2*cm, y + h - 4.6*cm, x + w - 0.6*cm, y + h - 4.6*cm)
        c.line(x + 3.9*cm, y + h - 5.9*cm, x + w - 0.6*cm, y + h - 5.9*cm)
        c.setDash([]) # Reset dash
        
        # Tipo de asignación (en cursiva y resaltado)
        c.setFont("Helvetica-Oblique", 12)
        # Limpiar tipo (quitar minutos si los tiene)
        tipo_texto = re.sub(r'\(.*?\)', '', asig['tipo']).strip()
        c.drawCentredString(x + w/2, y + h - 7.0*cm, f"( {tipo_texto} )")
        
        # Ubicación
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(x + 0.6*cm, y + h - 8.0*cm, "Se presentará en:")
        
        c.setFont("Helvetica", 9.5)
        # Checkboxes de Salas
        # Sala principal
        c.rect(x + 1.2*cm, y + h - 8.6*cm, 0.35*cm, 0.35*cm)
        c.drawString(x + 1.3*cm, y + h - 8.5*cm, "✓") # Marcado por defecto en la principal
        c.drawString(x + 1.8*cm, y + h - 8.6*cm, "Sala principal")
        
        # Sala Aux 1
        c.rect(x + 1.2*cm, y + h - 9.2*cm, 0.35*cm, 0.35*cm)
        c.drawString(x + 1.8*cm, y + h - 9.2*cm, "Sala auxiliar núm. 1")
        
        # Sala Aux 2
        c.rect(x + 1.2*cm, y + h - 9.8*cm, 0.35*cm, 0.35*cm)
        c.drawString(x + 1.8*cm, y + h - 9.8*cm, "Sala auxiliar núm. 2")
        
        # Nota al estudiante (Bloque de texto formateado)
        style = getSampleStyleSheet()['BodyText']
        style.fontSize = 8
        style.leading = 9
        style.alignment = 0 # Left
        
        texto_nota = (
            "<b>Nota al estudiante:</b> En la <i>Guía de actividades</i> encontrará la información "
            "que necesita para su intervención. Repase también las indicaciones que se describen "
            "en las <i>Instrucciones para la reunión Vida y Ministerio Cristianos (S-38)</i>."
        )
        
        p = Paragraph(texto_nota, style)
        p_w, p_h = p.wrap(w - 1.2*cm, 3*cm)
        p.drawOn(c, x + 0.6*cm, y + 1.8*cm)
        
        # Pie de página
        c.setFont("Helvetica", 7.5)
        c.drawString(x + 0.6*cm, y + 0.6*cm, "S-89-S 11/23")
        c.drawRightString(x + w - 0.6*cm, y + 0.6*cm, "Impreso en Colombia")
        
        # Restaurar estado
        c.restoreState()

def main():
    root = ctk.CTk()
    app = CoordinacionVMC(root)
    root.mainloop()

if __name__ == "__main__":
    main()